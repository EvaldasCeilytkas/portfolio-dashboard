from __future__ import annotations

import argparse
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHEMA_VERSION = 1
EPSILON = 0.000001

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = PROJECT_ROOT / "excel" / "SEB Robo.xlsx"
DEFAULT_OUTPUT = (
    PROJECT_ROOT / "public" / "data" / "platforms" / "seb-robo.json"
)

OVERVIEW_SHEET = "Apžvalga"
MONTH_SHEET_RE = re.compile(r"^\d{4}\.\d{2}$")


def finite_number(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if math.isfinite(number) else default


def rounded(value: Any, digits: int = 2) -> float:
    return round(finite_number(value), digits)


def nullable_number(value: Any, digits: int = 4) -> float | None:
    if value is None:
        return None
    number = finite_number(value, float("nan"))
    if not math.isfinite(number):
        return None
    return round(number, digits)


def month_key(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}-{value.month:02d}"

    if isinstance(value, (int, float)):
        text = f"{float(value):.2f}"
    else:
        text = str(value or "").strip().replace(",", ".")

    match = re.fullmatch(r"(\d{4})\.(\d{1,2})", text)
    if not match:
        return None

    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}"


def month_start(value: Any) -> str | None:
    key = month_key(value)
    return f"{key}-01" if key else None


def month_end(value: Any) -> str | None:
    key = month_key(value)
    if not key:
        return None

    year, month = map(int, key.split("-"))
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)

    return date.fromordinal(next_month.toordinal() - 1).isoformat()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def find_last_data_row(ws, start_row: int = 3) -> int:
    last_row = 0
    for row in range(start_row, ws.max_row + 1):
        if month_key(ws.cell(row, 1).value):
            last_row = row

    if not last_row:
        raise ValueError(f"Lape „{ws.title}“ nerasta mėnesinių duomenų.")

    return last_row


def read_overview(ws) -> tuple[list[dict[str, Any]], list[str], int]:
    tickers: list[str] = []
    ticker_columns: list[int] = []

    # SEB Robo ETF pozicijos yra E:I stulpeliuose.
    for col in range(5, 10):
        ticker = str(ws.cell(2, col).value or "").strip()
        if ticker:
            tickers.append(ticker)
            ticker_columns.append(col)

    last_row = find_last_data_row(ws)
    history: list[dict[str, Any]] = []

    previous_contributed = 0.0

    for row in range(3, last_row + 1):
        period = month_start(ws.cell(row, 1).value)
        if not period:
            continue

        contributed = finite_number(ws.cell(row, 2).value)
        invested = finite_number(ws.cell(row, 3).value)
        cash = finite_number(ws.cell(row, 4).value)
        holdings_value = finite_number(ws.cell(row, 10).value)
        current_value = finite_number(ws.cell(row, 11).value)
        profit = finite_number(ws.cell(row, 12).value)
        return_raw = ws.cell(row, 13).value
        dividends = finite_number(ws.cell(row, 14).value)
        fees = finite_number(ws.cell(row, 15).value)

        active_count = sum(
            1
            for col in ticker_columns
            if finite_number(ws.cell(row, col).value) > EPSILON
        )

        history.append({
            "date": period,
            "month": period[:7],
            "invested": rounded(contributed),
            "netInvested": rounded(invested),
            "holdingsValue": rounded(holdings_value),
            "currentValue": rounded(current_value),
            "profit": rounded(profit),
            "returnRate": (
                round(finite_number(return_raw) * 100, 4)
                if return_raw is not None
                else None
            ),
            "cash": rounded(cash),
            "income": rounded(dividends),
            "fees": rounded(fees),
            "contributions": rounded(contributed - previous_contributed),
            "withdrawals": 0.0,
            "activeInvestments": active_count,
            "delayedInvestments": 0,
            "completedInvestments": 0,
        })

        previous_contributed = contributed

    return history, tickers, last_row


def scan_monthly_activity(
    wb,
    tickers: list[str],
) -> dict[str, dict[str, float]]:
    activity: dict[str, dict[str, float]] = {
        ticker: defaultdict(float) for ticker in tickers
    }

    for sheet_name in wb.sheetnames:
        if not MONTH_SHEET_RE.fullmatch(sheet_name):
            continue

        ws = wb[sheet_name]

        for row in range(1, ws.max_row + 1):
            title = str(ws.cell(row, 1).value or "").strip()
            if not title:
                continue

            ticker = next(
                (
                    candidate
                    for candidate in tickers
                    if title == candidate or title.startswith(candidate + " ")
                ),
                None,
            )
            if not ticker:
                continue

            total_row = None
            for offset in range(1, 12):
                candidate_row = row + offset
                if candidate_row > ws.max_row:
                    break

                if (
                    str(ws.cell(candidate_row, 1).value or "").strip()
                    == "Viso:"
                ):
                    total_row = candidate_row
                    break

            if not total_row:
                continue

            item = activity[ticker]
            item["purchased"] += finite_number(ws.cell(total_row, 2).value)
            item["purchasedUnits"] += finite_number(
                ws.cell(total_row, 4).value
            )
            item["sold"] += finite_number(ws.cell(total_row, 7).value)
            item["soldUnits"] += finite_number(ws.cell(total_row, 9).value)
            item["dividends"] += finite_number(
                ws.cell(total_row, 12).value
            )

    return activity


def read_investment(
    ws,
    ticker: str,
    current_value: float,
    activity: dict[str, float],
) -> dict[str, Any]:
    rows = [
        row
        for row in range(5, ws.max_row + 1)
        if month_key(ws.cell(row, 1).value)
    ]

    if not rows:
        raise ValueError(
            f"{ticker}: investicijos lape nėra mėnesinių duomenų."
        )

    first_row = rows[0]
    last_row = rows[-1]

    full_name = str(ws["A1"].value or ticker).strip()
    name = full_name

    if full_name.upper().startswith(ticker.upper() + " "):
        name = full_name[len(ticker):].strip()

    sheet_quantity = finite_number(ws.cell(last_row, 5).value)
    price = finite_number(ws.cell(last_row, 6).value)
    sheet_value = finite_number(ws.cell(last_row, 7).value)

    total_purchased = finite_number(activity.get("purchased"))
    total_sold = finite_number(activity.get("sold"))
    total_dividends = finite_number(activity.get("dividends"))
    purchased_units = finite_number(activity.get("purchasedUnits"))
    sold_units = finite_number(activity.get("soldUnits"))
    remaining_units = purchased_units - sold_units

    # Aktyvi tik realiai likusi pozicija. Techninis likutis iki 0,01 vnt.
    # laikomas pilnai parduota pozicija.
    is_active = remaining_units > 0.01
    quantity = remaining_units if is_active else 0.0
    effective_current_value = current_value if is_active else 0.0

    net_invested = total_purchased - total_sold
    profit = (
        effective_current_value
        + total_sold
        + total_dividends
        - total_purchased
    )
    return_rate = (
        profit / total_purchased * 100
        if total_purchased > EPSILON
        else None
    )

    return {
        "id": ticker.lower(),
        "slug": slugify(ticker),
        "ticker": ticker,
        "name": name,
        "fullName": full_name,
        "type": "ETF",
        "status": "active" if is_active else "completed",
        "currency": "EUR",
        "startDate": month_start(ws.cell(first_row, 1).value),
        "endDate": (
            None if is_active else month_end(ws.cell(last_row, 1).value)
        ),
        "invested": rounded(total_purchased),
        "netInvested": rounded(net_invested),
        "currentValue": rounded(effective_current_value),
        "sheetValue": rounded(sheet_value),
        "profit": rounded(profit),
        "returnRate": nullable_number(return_rate),
        "quantity": round(quantity, 8),
        "purchasedUnits": round(purchased_units, 8),
        "soldUnits": round(sold_units, 8),
        "price": rounded(price if is_active else 0.0, 6),
        "realizedProceeds": rounded(total_sold),
        "dividends": rounded(total_dividends),
        "fees": 0.0,
    }


def build_document(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        input_path,
        data_only=True,
        read_only=False,
    )

    if OVERVIEW_SHEET not in workbook.sheetnames:
        raise ValueError(f"Nerastas privalomas lapas „{OVERVIEW_SHEET}“.")

    overview = workbook[OVERVIEW_SHEET]
    history, tickers, last_row = read_overview(overview)

    missing_sheets = [
        ticker for ticker in tickers if ticker not in workbook.sheetnames
    ]
    if missing_sheets:
        raise ValueError(
            "Nerasti instrumentų lapai: " + ", ".join(missing_sheets)
        )

    activity = scan_monthly_activity(workbook, tickers)

    investments: list[dict[str, Any]] = []

    for col, ticker in enumerate(tickers, start=5):
        current_value = finite_number(overview.cell(last_row, col).value)

        investments.append(
            read_investment(
                workbook[ticker],
                ticker,
                current_value,
                activity.get(ticker, {}),
            )
        )

    investments.sort(
        key=lambda item: (-item["currentValue"], item["ticker"])
    )

    latest = history[-1]
    active = [
        item for item in investments if item["status"] == "active"
    ]
    completed = [
        item for item in investments if item["status"] == "completed"
    ]

    total_contributed = rounded(latest["invested"])
    invested = rounded(
        sum(
            max(finite_number(item.get("netInvested")), 0.0)
            for item in active
        )
    )
    holdings_value = rounded(sum(item["currentValue"] for item in active))
    cash = rounded(latest["cash"])
    if abs(cash) < 0.05:
        cash = 0.0
    current_value = rounded(holdings_value + cash)
    realized_profit = rounded(sum(item["profit"] for item in investments))
    profit = realized_profit

    latest["totalContributed"] = total_contributed
    latest["invested"] = invested
    latest["netInvested"] = invested
    latest["holdingsValue"] = holdings_value
    latest["cash"] = cash
    latest["currentValue"] = current_value
    latest["profit"] = profit
    latest["activeInvestments"] = len(active)
    latest["completedInvestments"] = len(completed)

    return_rate = (
        round(profit / total_contributed * 100, 4)
        if total_contributed > EPSILON
        else None
    )

    total_dividends = rounded(
        sum(finite_number(point["income"]) for point in history)
    )
    total_fees = rounded(
        sum(finite_number(point["fees"]) for point in history)
    )

    updated_at = month_end(overview.cell(last_row, 1).value)
    start_date = history[0]["date"]

    document = {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAt": datetime.now().astimezone().isoformat(
            timespec="seconds"
        ),
        "platform": {
            "id": "seb-robo",
            "slug": "seb-robo",
            "name": "SEB Robo",
            "group": "robo",
            "type": "robo-advisor",
            "category": "Robo Advisor ETF portfelis",
            "currency": "EUR",
            "active": True,
            "startDate": start_date,
            "updatedAt": updated_at,
            "website": "https://www.seb.lt",
        },
        "summary": {
            "invested": invested,
            "netInvested": invested,
            "totalContributed": total_contributed,
            "holdingsValue": holdings_value,
            "currentValue": current_value,
            "profit": profit,
            "realizedProfit": realized_profit,
            "returnRate": return_rate,
            "xirr": None,
            "cash": cash,
            "incomeReceived": total_dividends,
            "fees": total_fees,
            "activeInvestments": len(active),
            "delayedInvestments": 0,
            "completedInvestments": len(completed),
            "averageRate": None,
            "averageLtv": None,
            "totalInvestments": len(investments),
        },
        "history": history,
        "investments": investments,
        "distributions": {
            "status": [
                {
                    "key": "active",
                    "label": "Aktyvios",
                    "count": len(active),
                    "value": rounded(
                        sum(item["currentValue"] for item in active)
                    ),
                },
                {
                    "key": "completed",
                    "label": "Parduotos",
                    "count": len(completed),
                    "value": rounded(
                        sum(
                            item["realizedProceeds"]
                            for item in completed
                        )
                    ),
                },
            ],
            "holdings": [
                {
                    "key": item["ticker"],
                    "label": item["ticker"],
                    "name": item["name"],
                    "value": item["currentValue"],
                    "weight": (
                        round(
                            item["currentValue"]
                            / latest["holdingsValue"]
                            * 100,
                            4,
                        )
                        if latest["holdingsValue"] > EPSILON
                        else 0.0
                    ),
                }
                for item in active
            ],
        },
        "latestMonth": history[-1],
        "largestInvestment": (
            {
                "id": active[0]["id"],
                "ticker": active[0]["ticker"],
                "name": active[0]["name"],
                "currentValue": active[0]["currentValue"],
            }
            if active
            else None
        ),
        "source": {
            "file": input_path.name,
            "overviewSheet": OVERVIEW_SHEET,
            "instrumentSheets": tickers,
            "monthlySheets": [
                name
                for name in workbook.sheetnames
                if MONTH_SHEET_RE.fullmatch(name)
            ],
        },
    }

    validate_document(document)
    return document


def validate_document(document: dict[str, Any]) -> None:
    summary = document["summary"]
    investments = document["investments"]

    errors: list[str] = []

    if not document["history"]:
        errors.append("History yra tuščias.")

    if not investments:
        errors.append("Investicijų sąrašas yra tuščias.")

    active_count = sum(
        item["status"] == "active" for item in investments
    )
    completed_count = sum(
        item["status"] == "completed" for item in investments
    )

    if active_count != summary["activeInvestments"]:
        errors.append("Nesutampa aktyvių investicijų skaičius.")

    if completed_count != summary["completedInvestments"]:
        errors.append("Nesutampa parduotų investicijų skaičius.")

    holdings_value = rounded(
        sum(
            item["currentValue"]
            for item in investments
            if item["status"] == "active"
        )
    )

    if abs(holdings_value - summary["holdingsValue"]) > 0.05:
        errors.append(
            "Aktyvių pozicijų vertė nesutampa su Apžvalgos ETF verte: "
            f"{holdings_value:.2f} != "
            f"{summary['holdingsValue']:.2f}"
        )

    expected_total = rounded(
        summary["holdingsValue"] + summary["cash"]
    )

    if abs(expected_total - summary["currentValue"]) > 0.05:
        errors.append(
            "ETF vertė ir grynieji nesutampa su portfelio verte: "
            f"{expected_total:.2f} != "
            f"{summary['currentValue']:.2f}"
        )

    ids = [item["id"] for item in investments]

    if len(ids) != len(set(ids)):
        errors.append("Rasti dubliuoti investicijų ID.")

    if errors:
        raise ValueError("\n".join(errors))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "SEB Robo Excel importeris į Portfolio V2 JSON."
        )
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Excel failas. Numatyta: {DEFAULT_INPUT}",
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"JSON failas. Numatyta: {DEFAULT_OUTPUT}",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    print("=" * 64)
    print("SEB ROBO IMPORTER V1")
    print("=" * 64)
    print(f"Excel failas: {input_path}")

    if not input_path.is_file():
        raise FileNotFoundError(
            f"Excel failas nerastas: {input_path}"
        )

    document = build_document(input_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8") as file:
        json.dump(
            document,
            file,
            ensure_ascii=False,
            indent=2,
        )

    summary = document["summary"]

    print(
        f"✅ Nuskaityta pozicijų: "
        f"{summary['totalInvestments']}"
    )
    print(f"✅ Aktyvių: {summary['activeInvestments']}")
    print(f"✅ Parduotų: {summary['completedInvestments']}")
    print(f"✅ Investuota: {summary['invested']:.2f} EUR")
    print(
        f"✅ ETF vertė: {summary['holdingsValue']:.2f} EUR"
    )
    print(f"✅ Grynieji: {summary['cash']:.2f} EUR")
    print(
        f"✅ Portfelio vertė: "
        f"{summary['currentValue']:.2f} EUR"
    )
    print(f"✅ Pelnas: {summary['profit']:.2f} EUR")
    print(f"✅ JSON sukurtas: {output_path}")


if __name__ == "__main__":
    main()
