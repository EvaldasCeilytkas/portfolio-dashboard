from __future__ import annotations

import argparse
import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHEMA_VERSION = 1
EPSILON = 0.000001

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = PROJECT_ROOT / "excel" / "Synergy.xlsx"
DEFAULT_OUTPUT = (
    PROJECT_ROOT / "public" / "data" / "platforms" / "synergy.json"
)

OVERVIEW_SHEET = "Apžvalga"
MONTH_SHEET_RE = re.compile(r"^\d{4}\.\d{2}$")


def finite_number(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if math.isfinite(number) else default


def rounded(value: Any, digits: int = 2) -> float:
    return round(finite_number(value), digits)


def nullable_number(value: Any, digits: int = 4) -> float | None:
    if value is None:
        return None
    number = finite_number(value, float("nan"))
    if not math.isfinite(number):
        return None
    return round(number, digits)


def month_key(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}-{value.month:02d}"

    if isinstance(value, (int, float)):
        text = f"{float(value):.2f}"
    else:
        text = str(value or "").strip().replace(",", ".")

    match = re.fullmatch(r"(\d{4})\.(\d{1,2})", text)
    if not match:
        return None

    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        return None

    return f"{year:04d}-{month:02d}"


def month_start(value: Any) -> str | None:
    key = month_key(value)
    return f"{key}-01" if key else None


def month_end(value: Any) -> str | None:
    key = month_key(value)
    if not key:
        return None

    year, month = map(int, key.split("-"))
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)

    return date.fromordinal(next_month.toordinal() - 1).isoformat()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def find_data_rows(ws, start_row: int = 1) -> list[int]:
    return [
        row
        for row in range(start_row, ws.max_row + 1)
        if month_key(ws.cell(row, 1).value)
    ]


def get_fund_sheets(workbook) -> list[str]:
    return [
        name
        for name in workbook.sheetnames
        if name != OVERVIEW_SHEET and not MONTH_SHEET_RE.fullmatch(name)
    ]


def read_history(ws) -> list[dict[str, Any]]:
    rows = find_data_rows(ws, start_row=2)
    if not rows:
        raise ValueError("Apžvalgos lape nerasta mėnesinių duomenų.")

    history: list[dict[str, Any]] = []
    previous_contributed = 0.0

    for row in rows:
        period = month_start(ws.cell(row, 1).value)
        if not period:
            continue

        contributed = finite_number(ws.cell(row, 2).value)
        invested = finite_number(ws.cell(row, 3).value)
        current_value = finite_number(ws.cell(row, 6).value)
        profit = finite_number(ws.cell(row, 7).value)
        return_raw = ws.cell(row, 8).value
        fees = finite_number(ws.cell(row, 9).value)

        active_count = sum(
            1
            for col in range(4, 6)
            if finite_number(ws.cell(row, col).value) > EPSILON
        )

        history.append({
            "date": period,
            "month": period[:7],
            "invested": rounded(contributed),
            "netInvested": rounded(invested),
            "currentValue": rounded(current_value),
            "profit": rounded(profit),
            "returnRate": (
                round(finite_number(return_raw) * 100, 4)
                if return_raw is not None
                else None
            ),
            "cash": 0.0,
            "income": 0.0,
            "fees": rounded(fees),
            "contributions": rounded(contributed - previous_contributed),
            "withdrawals": 0.0,
            "activeInvestments": active_count,
            "delayedInvestments": 0,
            "completedInvestments": 0,
        })

        previous_contributed = contributed

    return history


def read_investment(ws, sheet_name: str) -> dict[str, Any]:
    rows = find_data_rows(ws, start_row=2)
    if not rows:
        raise ValueError(f"{sheet_name}: nerasta mėnesinių duomenų.")

    first_row = rows[0]
    last_row = rows[-1]

    ticker = str(ws["A1"].value or sheet_name).strip()
    current_quantity = finite_number(ws.cell(last_row, 7).value)
    current_price = finite_number(ws.cell(last_row, 8).value)
    current_value = finite_number(ws.cell(last_row, 9).value)

    total_contributed = finite_number(ws.cell(last_row, 3).value)
    total_invested = finite_number(ws.cell(last_row, 5).value)
    total_profit = finite_number(ws.cell(last_row, 12).value)
    return_raw = ws.cell(last_row, 11).value

    is_active = (
        current_quantity > EPSILON
        and current_value > EPSILON
    )

    return {
        "id": slugify(sheet_name),
        "slug": slugify(sheet_name),
        "ticker": ticker,
        "name": sheet_name,
        "fullName": sheet_name,
        "type": "fund",
        "status": "active" if is_active else "completed",
        "currency": "EUR",
        "startDate": month_start(ws.cell(first_row, 1).value),
        "endDate": (
            None if is_active else month_end(ws.cell(last_row, 1).value)
        ),
        "invested": rounded(total_contributed),
        "netInvested": rounded(total_invested if is_active else 0.0),
        "currentValue": rounded(current_value if is_active else 0.0),
        "profit": rounded(total_profit),
        "returnRate": (
            round(finite_number(return_raw) * 100, 4)
            if return_raw is not None
            else None
        ),
        "quantity": round(current_quantity if is_active else 0.0, 8),
        "price": rounded(current_price if is_active else 0.0, 6),
        "realizedProceeds": 0.0,
        "dividends": 0.0,
        "fees": rounded(total_contributed - total_invested),
    }


def build_document(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        input_path,
        data_only=True,
        read_only=False,
    )

    if OVERVIEW_SHEET not in workbook.sheetnames:
        raise ValueError(f"Nerastas privalomas lapas „{OVERVIEW_SHEET}“.")

    fund_sheets = get_fund_sheets(workbook)
    if not fund_sheets:
        raise ValueError("Nerasti fondų lapai.")

    overview = workbook[OVERVIEW_SHEET]
    history = read_history(overview)

    investments = [
        read_investment(workbook[name], name)
        for name in fund_sheets
    ]
    investments.sort(
        key=lambda item: (-item["currentValue"], item["name"])
    )

    active = [
        item for item in investments if item["status"] == "active"
    ]
    completed = [
        item for item in investments if item["status"] == "completed"
    ]

    latest = history[-1]

    invested = rounded(latest["invested"])
    net_invested = rounded(latest["netInvested"])
    current_value = rounded(latest["currentValue"])
    profit = rounded(latest["profit"])
    return_rate = latest["returnRate"]
    total_fees = rounded(
        sum(finite_number(point["fees"]) for point in history)
    )

    latest["activeInvestments"] = len(active)
    latest["completedInvestments"] = len(completed)

    overall_xirr_raw = overview.cell(2, 45).value  # Overview!AS2
    overall_xirr = (
        round(finite_number(overall_xirr_raw) * 100, 4)
        if overall_xirr_raw not in (None, "")
        else None
    )

    document = {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAt": datetime.now().astimezone().isoformat(
            timespec="seconds"
        ),
        "platform": {
            "id": "synergy",
            "slug": "synergy",
            "name": "Synergy",
            "group": "funds",
            "type": "funds",
            "category": "Investiciniai fondai",
            "currency": "EUR",
            "active": bool(active),
            "startDate": history[0]["date"],
            "updatedAt": month_end(history[-1]["month"]),
            "website": None,
        },
        "summary": {
            "invested": invested,
            "netInvested": net_invested,
            "totalContributed": invested,
            "currentValue": current_value,
            "profit": profit,
            "realizedProfit": 0.0,
            "returnRate": return_rate,
            "xirr": overall_xirr,
            "cash": 0.0,
            "incomeReceived": 0.0,
            "fees": total_fees,
            "activeInvestments": len(active),
            "delayedInvestments": 0,
            "completedInvestments": len(completed),
            "averageRate": None,
            "averageLtv": None,
            "totalInvestments": len(investments),
        },
        "history": history,
        "investments": investments,
        "distributions": {
            "status": [
                {
                    "key": "active",
                    "label": "Aktyvūs",
                    "count": len(active),
                    "value": rounded(
                        sum(item["currentValue"] for item in active)
                    ),
                },
                {
                    "key": "completed",
                    "label": "Parduoti",
                    "count": len(completed),
                    "value": rounded(
                        sum(
                            item["realizedProceeds"]
                            for item in completed
                        )
                    ),
                },
            ],
            "holdings": [
                {
                    "key": item["ticker"],
                    "label": item["ticker"],
                    "name": item["name"],
                    "value": item["currentValue"],
                    "weight": (
                        round(
                            item["currentValue"]
                            / current_value
                            * 100,
                            4,
                        )
                        if current_value > EPSILON
                        else 0.0
                    ),
                }
                for item in active
            ],
        },
        "latestMonth": latest,
        "largestInvestment": (
            {
                "id": active[0]["id"],
                "ticker": active[0]["ticker"],
                "name": active[0]["name"],
                "currentValue": active[0]["currentValue"],
            }
            if active
            else None
        ),
        "source": {
            "file": input_path.name,
            "overviewSheet": OVERVIEW_SHEET,
            "instrumentSheets": fund_sheets,
            "monthlySheets": [
                name
                for name in workbook.sheetnames
                if MONTH_SHEET_RE.fullmatch(name)
            ],
        },
    }

    validate_document(document)
    return document


def validate_document(document: dict[str, Any]) -> None:
    summary = document["summary"]
    investments = document["investments"]
    errors: list[str] = []

    if not document["history"]:
        errors.append("History yra tuščias.")

    if not investments:
        errors.append("Investicijų sąrašas yra tuščias.")

    active_count = sum(
        item["status"] == "active" for item in investments
    )
    completed_count = sum(
        item["status"] == "completed" for item in investments
    )

    if active_count != summary["activeInvestments"]:
        errors.append("Nesutampa aktyvių fondų skaičius.")

    if completed_count != summary["completedInvestments"]:
        errors.append("Nesutampa parduotų fondų skaičius.")

    holdings_value = rounded(
        sum(
            item["currentValue"]
            for item in investments
            if item["status"] == "active"
        )
    )

    if abs(holdings_value - summary["currentValue"]) > 0.05:
        errors.append(
            "Aktyvių fondų vertė nesutampa su Apžvalgos verte: "
            f"{holdings_value:.2f} != "
            f"{summary['currentValue']:.2f}"
        )

    position_invested = rounded(
        sum(
            item["netInvested"]
            for item in investments
            if item["status"] == "active"
        )
    )

    if abs(position_invested - summary["netInvested"]) > 0.05:
        errors.append(
            "Fondų investuota suma nesutampa su Apžvalga: "
            f"{position_invested:.2f} != "
            f"{summary['netInvested']:.2f}"
        )

    ids = [item["id"] for item in investments]
    if len(ids) != len(set(ids)):
        errors.append("Rasti dubliuoti fondų ID.")

    if errors:
        raise ValueError("\n".join(errors))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Synergy Excel importeris į Portfolio V2 JSON."
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Excel failas. Numatyta: {DEFAULT_INPUT}",
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"JSON failas. Numatyta: {DEFAULT_OUTPUT}",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    print("=" * 64)
    print("SYNERGY IMPORTER V1")
    print("=" * 64)
    print(f"Excel failas: {input_path}")

    if not input_path.is_file():
        raise FileNotFoundError(f"Excel failas nerastas: {input_path}")

    document = build_document(input_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8") as file:
        json.dump(
            document,
            file,
            ensure_ascii=False,
            indent=2,
        )

    summary = document["summary"]

    print(f"✅ Nuskaityta fondų: {summary['totalInvestments']}")
    print(f"✅ Aktyvių: {summary['activeInvestments']}")
    print(f"✅ Parduotų: {summary['completedInvestments']}")
    print(f"✅ Įnešta: {summary['invested']:.2f} EUR")
    print(f"✅ Investuota: {summary['netInvested']:.2f} EUR")
    print(f"✅ Fondų vertė: {summary['currentValue']:.2f} EUR")
    print(f"✅ Pelnas: {summary['profit']:.2f} EUR")
    print(f"✅ JSON sukurtas: {output_path}")


if __name__ == "__main__":
    main()
