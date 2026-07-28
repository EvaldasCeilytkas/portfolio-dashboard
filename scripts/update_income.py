from __future__ import annotations

import json
import math
import shutil
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "excel" / "Income.xlsx"
PORTFOLIO_PATH = ROOT / "public" / "data" / "portfolio.json"
INCOME_PREVIEW_PATH = ROOT / "public" / "data" / "income-platform.json"


def number(value: Any, default: float = 0.0) -> float:
    try:
        numeric = float(value)
        return numeric if math.isfinite(numeric) else default
    except (TypeError, ValueError):
        return default


def round_money(value: Any) -> float:
    return round(number(value), 2)


def excel_date(value: Any) -> date | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    text = str(value).strip()

    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass

    return None


def iso_date(value: Any) -> str | None:
    parsed = excel_date(value)
    return parsed.isoformat() if parsed else None


def percent(value: Any) -> float:
    numeric = number(value)
    return numeric * 100 if abs(numeric) <= 1 else numeric


def find_label_value(ws, label: str) -> Any:
    wanted = label.strip().lower()

    for row in ws.iter_rows(min_col=1, max_col=2):
        cell_label = str(row[0].value or "").strip().lower()

        if cell_label == wanted:
            return row[1].value

    return None


def loan_status(
    planned_end: date | None,
    actual_end: date | None,
    remaining_principal: float,
    delay_days: int,
) -> str:
    if actual_end or remaining_principal <= 0.005:
        return "completed"

    if delay_days > 0:
        return "late"

    return "active"


def read_payments(ws) -> list[dict[str, Any]]:
    payments: list[dict[str, Any]] = []

    for row in range(3, ws.max_row + 1):
        planned = excel_date(ws.cell(row, 4).value)
        actual = excel_date(ws.cell(row, 5).value)
        principal = round_money(ws.cell(row, 7).value)
        interest = round_money(ws.cell(row, 8).value)
        fee = round_money(ws.cell(row, 9).value)

        if not planned and not actual and principal == 0 and interest == 0 and fee == 0:
            continue

        # Tuščių „Viso“ ir formulių rezervinių eilučių neįtraukiame.
        if str(ws.cell(row, 4).value or "").strip().lower() == "viso:":
            continue

        if principal == 0 and interest == 0 and fee == 0:
            continue

        delay_days = 0
        if planned and actual:
            delay_days = (actual - planned).days

        payments.append(
            {
                "plannedDate": planned.isoformat() if planned else None,
                "actualDate": actual.isoformat() if actual else None,
                "delayDays": delay_days,
                "principal": principal,
                "interest": interest,
                "fee": fee,
                "netIncome": round(interest - fee, 2),
                "status": "paid" if actual else "scheduled",
            }
        )

    return payments


def read_loan(ws, overview_row: dict[str, Any], today: date) -> dict[str, Any]:
    loan_id = str(find_label_value(ws, "Loan ID") or ws.title)
    country = str(find_label_value(ws, "Šalis") or overview_row.get("country") or "—")
    originator = str(find_label_value(ws, "Skolintojas") or overview_row.get("originator") or "—")
    loan_type = str(find_label_value(ws, "Tipas") or "P2P paskola")
    term = str(find_label_value(ws, "Terminas") or overview_row.get("term") or "—")

    invested_date = excel_date(find_label_value(ws, "Investuota"))
    start_date = excel_date(find_label_value(ws, "Pradžia"))
    planned_end = excel_date(find_label_value(ws, "Pabaiga"))
    actual_end = excel_date(overview_row.get("actualEndDate"))

    interest_rate = percent(find_label_value(ws, "Palūkanos"))
    invested = round_money(
        find_label_value(ws, "Investuota, Eur")
        or find_label_value(ws, "Investuota viso")
        or overview_row.get("invested")
    )

    principal_repaid = round_money(overview_row.get("principalRepaid"))
    interest_received = round_money(overview_row.get("interestReceived"))
    fees = round_money(overview_row.get("fees"))
    remaining_principal = max(0.0, round(invested - principal_repaid, 2))
    payments = read_payments(ws)

    completed_payment_delays = [
        number(item.get("delayDays"))
        for item in payments
        if item.get("status") == "paid"
    ]
    active_delay = max([0, *completed_payment_delays])
    status = loan_status(planned_end, actual_end, remaining_principal, 0)

    days_remaining = (
        max(0, (planned_end - today).days)
        if planned_end and status != "completed"
        else 0
    )

    return {
        "id": loan_id,
        "loanId": loan_id,
        "country": country,
        "originator": originator,
        "type": loan_type,
        "term": term,
        "investedDate": invested_date.isoformat() if invested_date else None,
        "startDate": start_date.isoformat() if start_date else None,
        "plannedEndDate": planned_end.isoformat() if planned_end else None,
        "actualEndDate": actual_end.isoformat() if actual_end else None,
        "daysRemaining": days_remaining,
        "delayDays": 0,
        "interestRate": round(interest_rate, 2),
        "invested": invested,
        "principalRepaid": principal_repaid,
        "remainingPrincipal": remaining_principal,
        "interestReceived": interest_received,
        "fees": fees,
        "netIncome": round(interest_received - fees, 2),
        "status": status,
        "payments": payments,
    }


def read_overview(ws) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    loans: list[dict[str, Any]] = []

    for row in range(3, ws.max_row + 1):
        loan_id = ws.cell(row, 3).value

        if loan_id in (None, ""):
            continue

        loans.append(
            {
                "loanId": str(loan_id),
                "investedDate": ws.cell(row, 2).value,
                "country": ws.cell(row, 4).value,
                "originator": ws.cell(row, 5).value,
                "term": ws.cell(row, 6).value,
                "interestRate": ws.cell(row, 7).value,
                "invested": ws.cell(row, 8).value,
                "plannedEndDate": ws.cell(row, 9).value,
                "actualEndDate": ws.cell(row, 10).value,
                "daysRemaining": ws.cell(row, 11).value,
                "delayDays": ws.cell(row, 12).value,
                "principalRepaid": ws.cell(row, 13).value,
                "interestReceived": ws.cell(row, 14).value,
                "fees": ws.cell(row, 15).value,
            }
        )

    # Overview suvestinės reikšmės pateiktos pirmoje duomenų eilutėje.
    summary_row = 3
    summary = {
        "deposited": round_money(ws.cell(summary_row, 17).value),
        "invested": round_money(ws.cell(summary_row, 18).value),
        "bonuses": round_money(ws.cell(summary_row, 19).value),
        "cash": round_money(ws.cell(summary_row, 20).value),
        "principalRepaid": round_money(ws.cell(summary_row, 21).value),
        "interestReceived": round_money(ws.cell(summary_row, 22).value),
        "value": round_money(ws.cell(summary_row, 23).value),
        "fees": round_money(ws.cell(summary_row, 27).value),
    }

    return loans, summary


def read_cash_history(ws) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        event_date = excel_date(ws.cell(row, 1).value)
        deposited = round_money(ws.cell(row, 2).value)
        withdrawn = round_money(ws.cell(row, 3).value)
        bonus = round_money(ws.cell(row, 4).value)

        if not event_date:
            continue

        events.append(
            {
                "date": event_date,
                "deposited": deposited,
                "withdrawn": withdrawn,
                "bonus": bonus,
            }
        )

    return sorted(events, key=lambda item: item["date"])


def build_platform(excel_path: Path) -> dict[str, Any]:
    workbook = load_workbook(excel_path, data_only=True)
    today = date.today()

    overview_rows, overview_summary = read_overview(workbook["Overview"])
    overview_by_id = {item["loanId"]: item for item in overview_rows}

    loans: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in {"Overview", "Pinigai"}:
            continue

        if not str(sheet_name).isdigit():
            continue

        loans.append(
            read_loan(
                workbook[sheet_name],
                overview_by_id.get(str(sheet_name), {}),
                today,
            )
        )

    loans.sort(key=lambda item: item.get("investedDate") or "")

    active_loans = [item for item in loans if item["status"] == "active"]
    late_loans = [item for item in loans if item["status"] == "late"]
    completed_loans = [item for item in loans if item["status"] == "completed"]

    deposited = overview_summary["deposited"]
    bonuses = overview_summary["bonuses"]
    cash = overview_summary["cash"]
    withdrawn = 0.0
    outstanding = round(sum(item["remainingPrincipal"] for item in loans), 2)
    current_value = round_money(overview_summary["value"] or (outstanding + cash))
    profit = round(current_value + withdrawn - deposited, 2)
    return_rate = round((profit / deposited * 100) if deposited else 0, 6)

    outstanding_active = sum(item["remainingPrincipal"] for item in active_loans)
    weighted_rate = (
        sum(item["interestRate"] * item["remainingPrincipal"] for item in active_loans)
        / outstanding_active
        if outstanding_active
        else 0
    )

    largest_share = (
        max((item["remainingPrincipal"] for item in active_loans), default=0)
        / current_value
        * 100
        if current_value
        else 0
    )
    late_share = (
        sum(item["remainingPrincipal"] for item in late_loans)
        / current_value
        * 100
        if current_value
        else 0
    )
    cash_share = cash / current_value * 100 if current_value else 0

    forecast_by_month: dict[str, dict[str, float | str]] = {}
    for loan in active_loans:
        planned_end = loan.get("plannedEndDate")
        if not planned_end:
            continue

        month = planned_end[:7]
        item = forecast_by_month.setdefault(
            month,
            {
                "month": month,
                "principal": 0.0,
                "interest": 0.0,
                "fees": 0.0,
                "total": 0.0,
            },
        )
        item["principal"] = number(item["principal"]) + loan["remainingPrincipal"]
        item["total"] = number(item["total"]) + loan["remainingPrincipal"]

    income_forecast = []
    for item in sorted(forecast_by_month.values(), key=lambda value: str(value["month"])):
        income_forecast.append(
            {
                **item,
                "principal": round_money(item["principal"]),
                "interest": round_money(item["interest"]),
                "fees": round_money(item["fees"]),
                "total": round_money(item["total"]),
            }
        )

    income_history = []
    for loan in loans:
        for payment in loan["payments"]:
            if payment["status"] != "paid" or not payment["actualDate"]:
                continue

            income_history.append(
                {
                    "date": payment["actualDate"],
                    "principal": payment["principal"],
                    "interest": payment["interest"],
                    "fees": payment["fee"],
                    "netIncome": payment["netIncome"],
                }
            )
    income_history.sort(key=lambda item: item["date"])

    cash_events = read_cash_history(workbook["Pinigai"])
    platform_history = []
    cumulative_invested = 0.0

    for event in cash_events:
        cumulative_invested += event["deposited"] - event["withdrawn"]
        platform_history.append(
            {
                "date": event["date"].isoformat(),
                "invested": round_money(cumulative_invested),
                "value": round_money(cumulative_invested),
            }
        )

    last_date = max(
        [event["date"] for event in cash_events] + [today],
        default=today,
    )
    if not platform_history or platform_history[-1]["date"] != last_date.isoformat():
        platform_history.append(
            {
                "date": last_date.isoformat(),
                "invested": deposited,
                "value": current_value,
            }
        )
    else:
        platform_history[-1]["value"] = current_value

    unique_originators = len({item["originator"] for item in loans})
    diversification_score = min(20, 8 + len(loans) * 2 + unique_originators)
    quality_score = max(0, round(20 - late_share * 0.8))
    cash_score = max(0, round(20 - cash_share * 0.8))
    repayment_score = 18 if any(
        number(payment["delayDays"]) > 0
        for loan in loans
        for payment in loan["payments"]
    ) else 20
    concentration_score = max(0, round(20 - max(0, largest_share - 20) * 0.45))
    health_score = int(
        diversification_score
        + quality_score
        + cash_score
        + repayment_score
        + concentration_score
    )

    details = {
        "type": "p2p",
        "modules": {"loans": True},
        "p2pSummary": {
            "totalLoans": len(loans),
            "activeLoans": len(active_loans),
            "completedLoans": len(completed_loans),
            "lateLoans": len(late_loans),
            "averageInterestRate": round(weighted_rate, 2),
            "cash": cash,
            "interestReceived": round_money(
                sum(item["interestReceived"] for item in loans)
            ),
            "principalRepaid": round_money(
                sum(item["principalRepaid"] for item in loans)
            ),
            "outstandingPrincipal": outstanding,
        },
        "cashflow": {
            "deposited": deposited,
            "withdrawn": withdrawn,
            "bonuses": bonuses,
            "cash": cash,
            "invested": current_value,
            "principalRepaid": round_money(
                sum(item["principalRepaid"] for item in loans)
            ),
            "interestReceived": round_money(
                sum(item["interestReceived"] for item in loans)
            ),
            "fees": round_money(sum(item["fees"] for item in loans)),
            "netIncome": round_money(
                sum(item["netIncome"] for item in loans) + bonuses
            ),
        },
        "health": {
            "score": health_score,
            "label": (
                "Puiki būklė"
                if health_score >= 85
                else "Gera būklė"
                if health_score >= 70
                else "Reikia stebėti"
            ),
            "summary": (
                "Aktyvių vėluojančių paskolų nėra, tačiau portfelis dar nedidelis "
                "ir gana koncentruotas."
                if not late_loans
                else "Portfelyje yra vėluojančių paskolų, todėl reikalinga papildoma priežiūra."
            ),
            "largestLoanShare": round(largest_share, 2),
            "lateShare": round(late_share, 2),
            "cashShare": round(cash_share, 2),
            "metrics": [
                {
                    "key": "diversification",
                    "label": "Diversifikacija",
                    "score": diversification_score,
                    "max": 20,
                    "description": f"{len(loans)} paskolos ir {unique_originators} skirtingi skolintojai.",
                },
                {
                    "key": "loanQuality",
                    "label": "Paskolų kokybė",
                    "score": quality_score,
                    "max": 20,
                    "description": f"{len(late_loans)} aktyvių vėluojančių paskolų.",
                },
                {
                    "key": "cashEfficiency",
                    "label": "Pinigų efektyvumas",
                    "score": cash_score,
                    "max": 20,
                    "description": f"Laisvų pinigų likutis – {cash:.2f} €.",
                },
                {
                    "key": "repaymentStability",
                    "label": "Grąžinimų stabilumas",
                    "score": repayment_score,
                    "max": 20,
                    "description": "Vertinama pagal faktinių mokėjimų vėlavimus.",
                },
                {
                    "key": "concentration",
                    "label": "Koncentracijos rizika",
                    "score": concentration_score,
                    "max": 20,
                    "description": f"Didžiausia aktyvi paskola sudaro {largest_share:.1f} % portfelio.",
                },
            ],
        },
        "history": income_history,
        "incomeForecast": income_forecast,
        "loans": loans,
    }

    return {
        "name": "Income",
        "slug": "income",
        "category": "P2P paskolos",
        "assetClass": "p2p",
        "currency": "EUR",
        "active": True,
        "website": "https://getincome.com",
        "invested": deposited,
        "value": current_value,
        "profit": profit,
        "returnRate": return_rate,
        "xirr": 0,
        "history": platform_history,
        "analytics": {
            "startDate": platform_history[0]["date"] if platform_history else "",
            "months": max(1, len({item["date"][:7] for item in platform_history})),
            "highestValue": max(
                [number(item["value"]) for item in platform_history],
                default=current_value,
            ),
            "averageMonthlyReturn": return_rate,
            "maxDrawdown": 0,
            "winningRate": 100 if profit >= 0 else 0,
            "volatility": 0,
            "monthlyPerformance": [
                {
                    "date": today.strftime("%Y-%m"),
                    "previousValue": deposited,
                    "value": current_value,
                    "monthlyProfit": profit,
                    "monthlyReturn": return_rate,
                }
            ],
        },
        "details": details,
    }


def merge_into_portfolio(platform: dict[str, Any]) -> None:
    if not PORTFOLIO_PATH.exists():
        raise FileNotFoundError(
            f"Nerastas {PORTFOLIO_PATH}. Patikrink, ar scenarijus įkeltas į projekto scripts aplanką."
        )

    portfolio = json.loads(PORTFOLIO_PATH.read_text(encoding="utf-8"))
    platforms = portfolio.setdefault("platforms", [])

    existing_index = next(
        (
            index
            for index, item in enumerate(platforms)
            if item.get("slug") == "income"
            or str(item.get("name", "")).strip().lower() == "income"
        ),
        None,
    )

    if existing_index is None:
        platforms.append(platform)
        action = "pridėta"
    else:
        platforms[existing_index] = platform
        action = "atnaujinta"

    backup_path = PORTFOLIO_PATH.with_suffix(".before-income.json")
    shutil.copy2(PORTFOLIO_PATH, backup_path)

    PORTFOLIO_PATH.write_text(
        json.dumps(portfolio, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    INCOME_PREVIEW_PATH.write_text(
        json.dumps(platform, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Income platforma {action}.")
    print(f"Atnaujintas failas: {PORTFOLIO_PATH}")
    print(f"Atsarginė kopija: {backup_path}")
    print(f"Income peržiūros failas: {INCOME_PREVIEW_PATH}")
    print("Atidaryk: http://localhost:5173/platforms/income")


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Nerastas {EXCEL_PATH}. Įkelk Income.xlsx į projekto excel aplanką."
        )

    platform = build_platform(EXCEL_PATH)
    merge_into_portfolio(platform)


if __name__ == "__main__":
    main()
