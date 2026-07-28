import json
import re
import shutil
import unicodedata
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / "excel"
PORTFOLIO = ROOT / "public" / "data" / "portfolio.json"
MAIN = EXCEL / "Investavimas.xlsx"

CONFIG = {
    "seb-robo": {"name":"SEB Robo","file":"SEB Robo.xlsx","type":"robo","category":"Robo Advisor","domain":"seb.lt","header_row":2,"positions_from":5,"positions_until":"Viso","invested":"Inešta","value":"Vertė","labels":{"SGAS":"iShares MSCI USA ESG Screened","SLMC":"iShares MSCI Europe ESG Screened","AYEM":"iShares MSCI EM IMI ESG Screened","SGAJ":"iShares MSCI Japan ESG Screened","EDMU":"iShares MSCI USA ESG Enhanced"}},
    "seb-fondai": {"name":"SEB Fondai","file":"SEB Fondai.xlsx","type":"fund","category":"Investiciniai fondai","domain":"seb.lt","header_row":2,"positions_from":5,"positions_until":"Vertė","invested":"Investuota","value":"Vertė","labels":{"Active 55":"SEB Active 55","Active 80":"SEB Active 80","Artificial":"SEB Artificial Intelligence Fund","Glob. Yield":"SEB Global High Yield Fund","Glob.Ex.UD":"SEB Global Exposure Fund UD","Eur.Ex.UC":"SEB Europe Exposure Fund UC","US.Ex.UD":"SEB US Exposure Fund UD","Eur.Ex.UD":"SEB Europe Exposure Fund UD"}},
    "revolut-robo": {"name":"Revolut Robo","file":"Revolut Robo.xlsx","type":"robo","category":"Robo Advisor","domain":"revolut.com","header_row":2,"positions_from":5,"positions_until":"Viso","invested":"Inešta","value":"Vertė","closed":True,"labels":{}},
    "synergy": {"name":"Synergy","file":"Synergy.xlsx","type":"fund","category":"Investiciniai fondai","domain":"synergy-finance.com","header_row":1,"positions_from":4,"positions_until":"Vertė","invested":"Inešta","value":"Vertė","labels":{"NTF":"Nextury Technology Fund","NAF":"Nextury Asia Technology Fund"}},
}

def num(v):
    if v is None or isinstance(v,bool): return 0.0
    try: return float(str(v).replace(" ","").replace(",","."))
    except (TypeError,ValueError): return 0.0

def money(v): return round(num(v),2)

def norm(v):
    t=unicodedata.normalize("NFKD",str(v or "")).encode("ascii","ignore").decode("ascii").lower()
    return re.sub(r"\s+"," ",t).strip()

def slug(v): return re.sub(r"[^a-z0-9]+","-",norm(v)).strip("-")

def month_date(v):
    if isinstance(v,(datetime,date)): return v.strftime("%Y-%m-%d")
    m=re.search(r"(\d{4})[.\-/](\d{1,2})",str(v or "").replace(",","."))
    if not m: return ""
    y,mo=int(m.group(1)),int(m.group(2))
    if not 1<=mo<=12: return ""
    nxt=date(y+(mo==12),1 if mo==12 else mo+1,1)
    return date.fromordinal(nxt.toordinal()-1).isoformat()

def headers(ws,row):
    return {norm(ws.cell(row,c).value):c for c in range(1,ws.max_column+1) if ws.cell(row,c).value is not None}

def header_col(mapping,name):
    key=norm(name)
    if key not in mapping: raise ValueError(f"Nerasta antraštė „{name}“.")
    return mapping[key]

def latest_month_row(ws,start):
    for r in range(ws.max_row,start,-1):
        if month_date(ws.cell(r,1).value): return r
    raise ValueError(f"Lape „{ws.title}“ nerasta mėnesinių duomenų.")

def parse_workbook(cfg):
    path=EXCEL/cfg["file"]
    if not path.exists(): raise FileNotFoundError(f"Nerastas excel/{cfg['file']}.")
    wb=load_workbook(path,data_only=True); ws=wb["Apžvalga"]
    hr=cfg["header_row"]; hm=headers(ws,hr); row=latest_month_row(ws,hr)
    invested=money(ws.cell(row,header_col(hm,cfg["invested"])).value)
    value=money(ws.cell(row,header_col(hm,cfg["value"])).value)
    end=header_col(hm,cfg["positions_until"]); positions=[]
    for c in range(cfg["positions_from"],end):
        ticker=str(ws.cell(hr,c).value or "").strip(); val=money(ws.cell(row,c).value)
        if not ticker or val<=0: continue
        positions.append({"id":slug(ticker),"ticker":ticker,"name":cfg["labels"].get(ticker,ticker),"assetType":"Fondas" if cfg["type"]=="fund" else "ETF","currency":"EUR","status":"active","active":True,"value":val})
    positions.sort(key=lambda x:x["value"],reverse=True)
    note=""
    for r in range(hr+1,ws.max_row+1):
        for c in range(1,min(4,ws.max_column)+1):
            cell=ws.cell(r,c).value
            if isinstance(cell,str) and "uzdaryta" in norm(cell): note=cell.strip(); break
        if note: break
    return {"invested":invested,"value":value,"latestDate":month_date(ws.cell(row,1).value),"positions":positions,"closureNote":note}

def main_history(name):
    if not MAIN.exists(): return []
    wb=load_workbook(MAIN,data_only=True); ws=wb["Investavimas"]; start=None
    for c in range(1,ws.max_column):
        if norm(ws.cell(1,c).value)==norm(name): start=c; break
    if start is None: return []
    out=[]; begun=False
    for r in range(3,ws.max_row+1):
        raw=ws.cell(r,1).value
        dt=raw.strftime("%Y-%m-%d") if isinstance(raw,(datetime,date)) else month_date(raw)
        if not dt: continue
        inv=money(ws.cell(r,start).value); val=money(ws.cell(r,start+1).value)
        if inv or val: begun=True
        if not begun: continue
        out.append({"date":dt,"invested":inv,"value":val,"profit":round(val-inv,2),"returnRate":round(((val-inv)/inv*100) if inv else 0,2)})
    return out

def monthly_performance(history):
    out=[]
    for i in range(1,len(history)):
        p,c=history[i-1],history[i]
        pv,cv,pi,ci=num(p["value"]),num(c["value"]),num(p["invested"]),num(c["invested"])
        flow=ci-pi; profit=cv-pv-flow; ret=profit/pv*100 if pv else 0
        out.append({"date":c["date"],"previousValue":round(pv,2),"currentValue":round(cv,2),"previousInvested":round(pi,2),"currentInvested":round(ci,2),"cashFlow":round(flow,2),"monthlyProfit":round(profit,2),"monthlyReturn":round(ret,2)})
    return out

def build_analytics(history):
    perf=monthly_performance(history); returns=[num(x["monthlyReturn"]) for x in perf]; values=[num(x["value"]) for x in history]
    peak=0; drawdown=0
    for value in values:
        peak=max(peak,value)
        if peak: drawdown=min(drawdown,(value-peak)/peak*100)
    return {"startDate":history[0]["date"] if history else "","months":len(history),"highestValue":round(max(values),2) if values else 0,"averageMonthlyReturn":round(sum(returns)/len(returns),2) if returns else 0,"bestMonth":round(max(returns),2) if returns else 0,"worstMonth":round(min(returns),2) if returns else 0,"winningRate":round(len([x for x in returns if x>0])/len(returns)*100,2) if returns else 0,"maxDrawdown":round(drawdown,2),"monthlyPerformance":perf}

def find_platform(portfolio,p_slug,name):
    for p in portfolio.setdefault("platforms",[]):
        if (p.get("slug") or slug(p.get("name")))==p_slug: return p
    p={"name":name,"slug":p_slug}; portfolio["platforms"].append(p); return p

def update_platform(portfolio,p_slug,cfg):
    parsed=parse_workbook(cfg); history=main_history(cfg["name"])
    if not cfg.get("closed"):
        while history and history[-1]["invested"]==0 and history[-1]["value"]==0: history.pop()
    point={"date":parsed["latestDate"],"invested":parsed["invested"],"value":parsed["value"],"profit":round(parsed["value"]-parsed["invested"],2),"returnRate":round(((parsed["value"]-parsed["invested"])/parsed["invested"]*100) if parsed["invested"] else 0,2)}
    if point["date"]:
        if history and history[-1]["date"]==point["date"]: history[-1]=point
        elif not history or history[-1]["date"]<point["date"]: history.append(point)
    invested=parsed["invested"]; value=parsed["value"]; active=not cfg.get("closed",False)
    if cfg.get("closed"): active=False; value=0.0
    total=sum(x["value"] for x in parsed["positions"])
    for x in parsed["positions"]:
        share=x["value"]/total if total else 0
        x["share"]=round(share*100,2); x["invested"]=round(invested*share,2); x["profit"]=round(x["value"]-x["invested"],2); x["returnRate"]=round(x["profit"]/x["invested"]*100 if x["invested"] else 0,2); x["xirr"]=0
    profit=round(value-invested,2); ret=round(profit/invested*100 if invested else 0,2)
    p=find_platform(portfolio,p_slug,cfg["name"])
    p.update({"name":cfg["name"],"slug":p_slug,"assetClass":cfg["type"],"category":cfg["category"],"logoUrl":f"https://www.google.com/s2/favicons?domain={cfg['domain']}&sz=128","website":f"https://{cfg['domain']}","currency":"EUR","active":active,"invested":invested,"value":value,"profit":profit,"returnRate":ret,"history":history,"analytics":build_analytics(history),"details":{"type":cfg["type"],"sourceFile":cfg["file"],"updatedAt":datetime.now().isoformat(timespec="seconds"),"modules":{"broker":True,"positions":True,"allocation":True},"summary":{"invested":invested,"currentValue":value,"profit":profit,"returnRate":ret,"activePositions":len(parsed["positions"]) if active else 0,"closed":not active,"closureNote":parsed["closureNote"]},"positions":{"active":parsed["positions"] if active else [],"sold":parsed["positions"] if not active else []},"allocation":parsed["positions"]}})
    print(f"{cfg['name']}: {len(parsed['positions'])} pozicijos, investuota {invested:.2f} €, vertė {value:.2f} €.")

def main():
    if not PORTFOLIO.exists(): raise FileNotFoundError("Nerastas public/data/portfolio.json.")
    portfolio=json.loads(PORTFOLIO.read_text(encoding="utf-8")); backup=PORTFOLIO.with_name("portfolio.before-brokers.json"); shutil.copy2(PORTFOLIO,backup)
    errors=[]
    for p_slug,cfg in CONFIG.items():
        try: update_platform(portfolio,p_slug,cfg)
        except Exception as e: errors.append(f"{cfg['name']}: {e}")
    PORTFOLIO.write_text(json.dumps(portfolio,ensure_ascii=False,indent=2),encoding="utf-8")
    print(f"Atnaujintas: {PORTFOLIO}\nAtsarginė kopija: {backup}")
    if errors:
        print("\nKlaidos:")
        for e in errors: print(f"- {e}")
        raise SystemExit(1)
    print("\nVisi keturi brokeriai atnaujinti.")

if __name__=="__main__": main()
