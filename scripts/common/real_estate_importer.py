from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

SCHEMA_VERSION = 1
PROJECT_ROOT = Path(__file__).resolve().parents[2]

@dataclass(frozen=True)
class PlatformConfig:
    slug: str
    name: str
    website: str
    excel_name: str
    overview_sheet: str = "Overview"
    payment_style: str = "standard"  # standard | rontgen | nordstreet
    completion_requires_date: bool = False
    nordstreet_codes: bool = False
    maturity_labels: tuple[str, ...] = ("Paskolos pabaiga",)
    paid_row_is_settled: bool = False

CONFIG: PlatformConfig | None = None
PLATFORM_SLUG = ""
DEFAULT_OUTPUT = Path()
EXCEL_CANDIDATES: tuple[Path, ...] = ()

def configure(config: PlatformConfig) -> None:
    global CONFIG, PLATFORM_SLUG, DEFAULT_OUTPUT, EXCEL_CANDIDATES
    CONFIG = config
    PLATFORM_SLUG = config.slug
    DEFAULT_OUTPUT = PROJECT_ROOT / "public" / "data" / "platforms" / f"{config.slug}.json"
    EXCEL_CANDIDATES = (
        PROJECT_ROOT / "excel" / f"{config.excel_name}.xlsx",
        PROJECT_ROOT / "excel" / f"{config.excel_name.lower()}.xlsx",
        PROJECT_ROOT / f"{config.excel_name}.xlsx",
        PROJECT_ROOT / f"{config.excel_name.lower()}.xlsx",
    )


def finite_number(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        cleaned = (
            value.replace("\u00a0", "")
            .replace("€", "")
            .replace("%", "")
            .replace(",", ".")
            .strip()
        )
        if cleaned in {"", "-", "—"}:
            return default
        value = cleaned
    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    return result if math.isfinite(result) else default


def rounded(value: Any, digits: int = 2) -> float:
    return round(finite_number(value), digits)


def nullable_number(value: Any, digits: int = 2) -> float | None:
    if value is None or value == "" or (isinstance(value, str) and value.strip() in {"-", "—"}):
        return None
    result = finite_number(value, float("nan"))
    if not math.isfinite(result):
        return None
    return round(result, digits)


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def nullable_text(value: Any) -> str | None:
    value = clean_text(value)
    return value or None




def normalize_label(value: Any) -> str:
    text = clean_text(value).lower().replace("\u00a0", " ")
    text = re.sub(r"[^0-9a-ząčęėįšųūž]+", " ", text)
    return " ".join(text.split())


def find_sheet_name(workbook: Any, expected: str) -> str | None:
    target = normalize_label(expected)
    for name in workbook.sheetnames:
        if normalize_label(name) == target:
            return name
    return None


def find_header_column(
    ws: Any,
    header: str,
    *,
    row: int = 1,
    occurrence: int = 1,
    aliases: tuple[str, ...] = (),
) -> int:
    accepted = {normalize_label(header), *(normalize_label(a) for a in aliases)}
    found: list[int] = []
    for column in range(1, ws.max_column + 1):
        if normalize_label(ws.cell(row, column).value) in accepted:
            found.append(column)
    if len(found) < occurrence:
        raise ValueError(
            f"Lape '{ws.title}' nerasta antraštė '{header}' "
            f"(eilutė {row}, pasikartojimas {occurrence})."
        )
    return found[occurrence - 1]


def find_label_value(
    ws: Any,
    labels: tuple[str, ...],
    *,
    label_columns: tuple[int, ...] = (1,),
    value_offset: int = 1,
    min_row: int = 1,
    max_row: int | None = None,
) -> Any:
    accepted = {normalize_label(label) for label in labels}
    end_row = min(max_row or ws.max_row, ws.max_row)
    for row in range(min_row, end_row + 1):
        for column in label_columns:
            if normalize_label(ws.cell(row, column).value) in accepted:
                return ws.cell(row, column + value_offset).value
    raise ValueError(
        f"Lape '{ws.title}' nerastas laukas: " + " / ".join(labels)
    )


def overview_value(ws: Any, group: str, header: str, aliases: tuple[str, ...] = ()) -> Any:
    group_normalized = normalize_label(group)
    accepted = {normalize_label(header), *(normalize_label(a) for a in aliases)}
    active_group = ""
    for column in range(1, ws.max_column + 1):
        group_cell = normalize_label(ws.cell(1, column).value)
        if group_cell:
            active_group = group_cell
        if active_group == group_normalized and normalize_label(ws.cell(2, column).value) in accepted:
            return ws.cell(3, column).value
    raise ValueError(
        f"Lape '{ws.title}' nerasta suvestinė '{group} / {header}'."
    )


def excel_date(value: Any, epoch: datetime) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 0:
        try:
            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                continue
    return None


def iso_date(value: Any, epoch: datetime) -> str | None:
    parsed = excel_date(value, epoch)
    return parsed.isoformat() if parsed else None


def find_input(explicit: str | None) -> Path:
    if explicit:
        path = Path(explicit).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"{CONFIG.name} Excel failas nerastas: {path}")
        return path

    for candidate in EXCEL_CANDIDATES:
        if candidate.is_file():
            return candidate

    # Atsarginė paieška: priimame ir Windows automatiškai pridėtus numerius,
    # pvz. Crowdpear(10).xlsx, bet tik projekto excel aplanke / šaknyje.
    search_roots = [PROJECT_ROOT / "excel", PROJECT_ROOT]
    matches: list[Path] = []
    for root in search_roots:
        if root.is_dir():
            matches.extend(root.glob(f"{CONFIG.excel_name}*.xlsx"))
            matches.extend(root.glob(f"{CONFIG.excel_name.lower()}*.xlsx"))

    matches = sorted({path.resolve() for path in matches if path.is_file()})
    if matches:
        return matches[-1]

    checked = "\n".join(f"  - {path}" for path in EXCEL_CANDIDATES)
    raise FileNotFoundError(
        f"{CONFIG.name} Excel failas nerastas.\n"
        "Įkelkite jį į vieną iš šių vietų:\n"
        f"{checked}\n"
        "arba nurodykite kelią su --input."
    )


def read_summary(ws: Any) -> dict[str, Any]:
    deposited = rounded(overview_value(ws, "Pinigai", "Įnešta"))
    cash = rounded(overview_value(ws, "Pinigai", "Gryni"))
    current_value = rounded(overview_value(ws, "Vertė", "Viso"))
    profit = rounded(overview_value(ws, "Vertė", "Eur.", aliases=("Eur",)))
    income_received = rounded(overview_value(ws, "Grąžinta", "Palūkanos"))

    roi = nullable_number(overview_value(ws, "Vertė", "%"), 6)
    xirr = nullable_number(overview_value(ws, "Vertė", "XIRR"), 6)

    return {
        "invested": deposited,
        "currentValue": current_value,
        "profit": profit,
        "returnRate": round(roi * 100, 4) if roi is not None else None,
        "xirr": round(xirr * 100, 4) if xirr is not None else None,
        "cash": cash,
        "incomeReceived": income_received,
    }

def read_cash_flows(ws: Any, epoch: datetime) -> list[dict[str, Any]]:
    date_col = find_header_column(ws, "Data", row=1)
    deposited_col = find_header_column(ws, "Įnešta", row=1)
    withdrawn_col = find_header_column(ws, "Išimta", row=1)
    bonus_col = find_header_column(ws, "Premija", row=1, aliases=("Premijos",))

    flows: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        flow_date = excel_date(ws.cell(row, date_col).value, epoch)
        if not flow_date:
            continue
        deposited = rounded(ws.cell(row, deposited_col).value)
        withdrawn = rounded(ws.cell(row, withdrawn_col).value)
        bonus = rounded(ws.cell(row, bonus_col).value)
        if not any((deposited, withdrawn, bonus)):
            continue
        flows.append({
            "date": flow_date.isoformat(),
            "deposited": deposited,
            "withdrawn": withdrawn,
            "bonus": bonus,
        })
    return flows

def read_payments(ws: Any, epoch: datetime) -> list[dict[str, Any]]:
    planned_date_col = find_header_column(ws, "Data", row=1, occurrence=1)
    if CONFIG.payment_style == "nordstreet":
        def columns_after(header: str, start_column: int) -> list[int]:
            accepted = {normalize_label(header)}
            return [c for c in range(start_column + 1, ws.max_column + 1) if normalize_label(ws.cell(1, c).value) in accepted]
        principal_columns = columns_after("Paskola", planned_date_col)
        interest_columns = columns_after("Palūkanos", planned_date_col)
        if len(principal_columns) < 2 or len(interest_columns) < 2:
            raise ValueError(f"Lape '{ws.title}' nerasti planuoti ir faktiniai mokėjimų stulpeliai.")
        planned_principal_col, actual_principal_col = principal_columns[:2]
        planned_interest_col, actual_interest_col = interest_columns[:2]
        actual_extra_col = find_header_column(ws, "Kompens.", row=1, aliases=("Papildomos", "Kompensacija", "Kompensacijos"))
    elif CONFIG.payment_style == "rontgen":
        planned_principal_col = find_header_column(ws, "Paskola", row=1, occurrence=1)
        actual_principal_col = find_header_column(ws, "Paskola", row=1, occurrence=2)
        planned_interest_col = find_header_column(ws, "Palūkanos", row=1, occurrence=1)
        actual_interest_col = find_header_column(ws, "Palūkanos", row=1, occurrence=2)
        actual_extra_col = find_header_column(ws, "Papildomos", row=1)
    else:
        planned_principal_col = find_header_column(ws, "Paskola", row=1, occurrence=2)
        actual_principal_col = find_header_column(ws, "Paskola", row=1, occurrence=3)
        planned_interest_col = find_header_column(ws, "Palūkanos", row=1, occurrence=1)
        actual_interest_col = find_header_column(ws, "Palūkanos", row=1, occurrence=2)
        actual_extra_col = find_header_column(ws, "Papildomos", row=1)
    actual_date_col = find_header_column(ws, "Data", row=1, occurrence=2)
    delay_col = find_header_column(ws, "Vėlavimas", row=1)
    payments=[]
    for row in range(2, ws.max_row + 1):
        planned_date = excel_date(ws.cell(row, planned_date_col).value, epoch)
        actual_date = excel_date(ws.cell(row, actual_date_col).value, epoch)
        planned_principal = rounded(ws.cell(row, planned_principal_col).value)
        planned_interest = rounded(ws.cell(row, planned_interest_col).value)
        actual_principal = rounded(ws.cell(row, actual_principal_col).value)
        actual_interest = rounded(ws.cell(row, actual_interest_col).value)
        actual_extra = rounded(ws.cell(row, actual_extra_col).value)
        delay_days = int(round(finite_number(ws.cell(row, delay_col).value)))
        if not planned_date and not actual_date and not any((planned_principal, planned_interest, actual_principal, actual_interest, actual_extra)):
            continue
        payments.append({"plannedDate": planned_date.isoformat() if planned_date else None,"plannedPrincipal": planned_principal,"plannedInterest": planned_interest,"actualDate": actual_date.isoformat() if actual_date else None,"actualPrincipal": actual_principal,"actualInterest": actual_interest,"actualExtra": actual_extra,"historicalDelayDays": max(delay_days, 0)})
    return payments

def current_delay_days(payments: list[dict[str, Any]], today: date | None = None) -> int:
    """Grąžina tik šiuo metu neapmokėtų, jau suėjusių įmokų vėlavimą.

    Istorinis vėlavimas K stulpelyje statusui įtakos nebeturi: jei suplanuota
    įmoka jau sumokėta, projektas laikomas nevėluojančiu.
    """
    today = today or date.today()
    overdue_days = 0

    for payment in payments:
        planned_date_text = payment.get("plannedDate")
        if not planned_date_text:
            continue

        planned_date = date.fromisoformat(planned_date_text)
        planned_amount = rounded(
            finite_number(payment.get("plannedPrincipal"))
            + finite_number(payment.get("plannedInterest"))
        )
        actual_amount = rounded(
            finite_number(payment.get("actualPrincipal"))
            + finite_number(payment.get("actualInterest"))
        )

        # Dar neatėjusi įmoka arba eilutė be planuotos sumos nėra vėlavimas.
        if planned_date >= today or planned_amount <= 0.005:
            continue

        # Jei eilutėje yra faktinė mokėjimo data ir gauta suma, įmoka jau
        # sumokėta. Istorinis vėlavimas ar centų apvalinimo skirtumas statusui
        # įtakos nebeturi.
        if CONFIG.paid_row_is_settled and payment.get("actualDate") and actual_amount > 0.005:
            continue

        tolerance = 0.02 if CONFIG.paid_row_is_settled else 0.005
        if actual_amount + tolerance < planned_amount:
            overdue_days = max(overdue_days, (today - planned_date).days)

    return overdue_days


def project_status(
    completion_date: date | None,
    outstanding: float,
    delay_days: int,
) -> str:
    if completion_date or (not CONFIG.completion_requires_date and outstanding <= 0.005):
        return "completed"
    if delay_days > 0:
        return "delayed"
    return "active"


def project_code(raw_value: Any, sheet_name: str) -> str:
    raw = clean_text(raw_value)
    match = re.search(r"(?:Project\s*ID\s*[-:]?\s*)?(\d+)\s*\(LT\)", raw, re.IGNORECASE)
    if match:
        return f"LT-{match.group(1)}"
    match = re.search(r"\bLT[-\s]?(\d+)\b", raw, re.IGNORECASE)
    if match:
        return f"LT-{match.group(1)}"
    slug = re.sub(r"[^0-9a-ząčęėįšųūž]+", "-", sheet_name.lower()).strip("-")
    return f"NORDSTREET-{slug.upper()}"


def read_project(ws: Any, epoch: datetime) -> dict[str, Any]:
    raw_code = find_label_value(ws, ("Paskola", "Sutartis"), min_row=1, max_row=3)
    code = project_code(raw_code, ws.title) if CONFIG.nordstreet_codes else clean_text(raw_code)
    name = clean_text(find_label_value(ws, ("Objektas",), min_row=1, max_row=4)) or ws.title
    if not code:
        raise ValueError(f"Lape '{ws.title}' nerastas paskolos kodas.")

    investment_date = excel_date(find_label_value(ws, ("Investicijos data",), max_row=20), epoch)
    interest_start = excel_date(find_label_value(ws, ("Palūkanų pradžia",), max_row=20), epoch)
    rating = nullable_text(find_label_value(ws, ("Projekto reitingas", "Reitingas"), max_row=20))
    ltv = nullable_number(find_label_value(ws, ("LTV, %", "LTV", "LTC, %", "LTC"), max_row=20), 4)
    max_ltv = nullable_number(find_label_value(ws, ("LTV max, %", "LTV max", "LTC max, %", "LTC max"), max_row=20), 4)
    duration_months = int(round(finite_number(find_label_value(
        ws, ("Investicijos laikotarpis, mėn.", "Investicijos laikotarpis", "Laikotarpis, mėn."), max_row=20
    ))))
    payment_frequency = nullable_text(find_label_value(ws, ("Palūkanų mokėjimas",), max_row=20))
    interest_rate = nullable_number(find_label_value(ws, ("Palūkanos, %", "Palūkanos"), max_row=20), 6)
    invested = rounded(find_label_value(
        ws, ("Investuota,Eur", "Investuota, Eur", "Investuota viso", "Investuota"), max_row=20
    ))
    maturity_date = excel_date(find_label_value(ws, CONFIG.maturity_labels, max_row=20), epoch)
    completion_raw = find_label_value(ws, ("Grąžinta",), max_row=20)
    completion_date = excel_date(completion_raw, epoch)

    payments = read_payments(ws, epoch)
    repaid_principal = round(sum(item["actualPrincipal"] for item in payments), 2)
    income_received = round(
        sum(item["actualInterest"] + item["actualExtra"] for item in payments), 2
    )
    planned_interest = round(sum(item["plannedInterest"] for item in payments), 2)
    outstanding = max(round(invested - repaid_principal, 2), 0.0)
    delay_days = current_delay_days(payments)
    status = project_status(completion_date, outstanding, delay_days)

    return {
        "id": code,
        "code": code,
        "name": name,
        "investmentType": "real_estate_project",
        "status": status,
        "currency": "EUR",
        "invested": invested,
        "currentValue": outstanding,
        "profit": income_received,
        "investmentDate": investment_date.isoformat() if investment_date else None,
        "completionDate": completion_date.isoformat() if completion_date else None,
        "country": None,
        "interestRate": round(interest_rate * 100, 4) if interest_rate is not None else None,
        "incomeReceived": income_received,
        "outstanding": outstanding,
        "delayDays": delay_days,
        "ltv": round(ltv * 100, 4) if ltv is not None else None,
        "quantity": None,
        "averagePrice": None,
        "marketPrice": None,
        "ticker": None,
        "isin": None,
        "rating": rating,
        "maturityDate": maturity_date.isoformat() if maturity_date else None,
        "payments": [
            payment
            for row in payments
            for payment in (
                ([{
                    "date": row["actualDate"],
                    "type": "principal",
                    "amount": row["actualPrincipal"],
                    "note": None,
                }] if row["actualDate"] and row["actualPrincipal"] else [])
                + ([{
                    "date": row["actualDate"],
                    "type": "interest",
                    "amount": row["actualInterest"],
                    "note": None,
                }] if row["actualDate"] and row["actualInterest"] else [])
                + ([{
                    "date": row["actualDate"],
                    "type": "other",
                    "amount": row["actualExtra"],
                    "note": "Papildoma išmoka",
                }] if row["actualDate"] and row["actualExtra"] else [])
            )
        ],
        "paymentSchedule": payments,
        "interestStartDate": interest_start.isoformat() if interest_start else None,
        "durationMonths": duration_months,
        "paymentFrequency": payment_frequency,
        "maxLtv": round(max_ltv * 100, 4) if max_ltv is not None else None,
        "repaidPrincipal": repaid_principal,
        "plannedInterest": planned_interest,
        "sourceSheets": [ws.title],
    }

def merge_projects(projects: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for project in projects:
        grouped[project["code"]].append(project)

    merged_projects: list[dict[str, Any]] = []
    merged_codes: list[str] = []

    for code, items in grouped.items():
        if len(items) == 1:
            merged_projects.append(items[0])
            continue

        merged_codes.append(code)
        first = items[0].copy()
        first["invested"] = round(sum(item["invested"] for item in items), 2)
        first["currentValue"] = round(sum(item["currentValue"] for item in items), 2)
        first["profit"] = round(sum(item["profit"] for item in items), 2)
        first["incomeReceived"] = round(sum(item["incomeReceived"] for item in items), 2)
        first["outstanding"] = round(sum(item["outstanding"] for item in items), 2)
        first["repaidPrincipal"] = round(sum(item["repaidPrincipal"] for item in items), 2)
        first["plannedInterest"] = round(sum(item["plannedInterest"] for item in items), 2)
        first["delayDays"] = max(item["delayDays"] for item in items)
        first["sourceSheets"] = [sheet for item in items for sheet in item["sourceSheets"]]
        first["payments"] = sorted(
            [payment for item in items for payment in item["payments"]],
            key=lambda item: (item["date"], item["type"]),
        )
        first["paymentSchedule"] = sorted(
            [row for item in items for row in item["paymentSchedule"]],
            key=lambda item: (
                item["plannedDate"] or item["actualDate"] or "9999-12-31",
                item["actualDate"] or "9999-12-31",
            ),
        )

        dates = [item["investmentDate"] for item in items if item["investmentDate"]]
        first["investmentDate"] = min(dates) if dates else None
        completion_dates = [item["completionDate"] for item in items if item["completionDate"]]
        first["completionDate"] = max(completion_dates) if len(completion_dates) == len(items) else None
        maturity_dates = [item["maturityDate"] for item in items if item["maturityDate"]]
        first["maturityDate"] = max(maturity_dates) if maturity_dates else None

        statuses = {item["status"] for item in items}
        if "delayed" in statuses:
            first["status"] = "delayed"
        elif statuses == {"completed"}:
            first["status"] = "completed"
        else:
            first["status"] = "active"

        # Svertinis palūkanų ir LTV vidurkis pagal investuotą sumą.
        total_invested = first["invested"]
        for field in ("interestRate", "ltv", "maxLtv"):
            weighted = sum(
                (item[field] or 0) * item["invested"]
                for item in items
                if item[field] is not None
            )
            weight = sum(item["invested"] for item in items if item[field] is not None)
            first[field] = round(weighted / weight, 4) if weight else None

        merged_projects.append(first)

    merged_projects.sort(key=lambda item: (item["investmentDate"] or "9999-12-31", item["code"]))
    return merged_projects, sorted(merged_codes)


def weighted_average(projects: list[dict[str, Any]], field: str) -> float | None:
    numerator = 0.0
    denominator = 0.0
    for project in projects:
        value = project.get(field)
        invested = finite_number(project.get("invested"))
        if value is None or invested <= 0:
            continue
        numerator += finite_number(value) * invested
        denominator += invested
    return round(numerator / denominator, 4) if denominator else None


def build_history(summary: dict[str, Any], updated_at: date) -> list[dict[str, Any]]:
    # Nordstreet faile nėra pilnos mėnesinės platformos vertės istorijos.
    # Įrašome tik aktualų mėnesio momentinį vaizdą; senesnę istoriją vėliau
    # papildys bendras portfolio istorijos importeris.
    month_start = updated_at.replace(day=1)
    return [{
        "date": month_start.isoformat(),
        "invested": summary["invested"],
        "value": summary["currentValue"],
        "profit": summary["profit"],
        "cash": summary["cash"],
        "income": 0.0,
    }]


def validate_output(data: dict[str, Any]) -> None:
    required_root = {"schemaVersion", "generatedAt", "platform", "summary", "history", "investments"}
    missing = required_root - data.keys()
    if missing:
        raise ValueError("Trūksta JSON laukų: " + ", ".join(sorted(missing)))

    codes = [item["code"] for item in data["investments"]]
    duplicates = sorted({code for code in codes if codes.count(code) > 1})
    if duplicates:
        raise ValueError("Po sujungimo liko dubliuoti kodai: " + ", ".join(duplicates))

    for item in data["investments"]:
        if item["id"] != item["code"]:
            raise ValueError(f"Investicijos id/code nesutampa: {item['code']}")
        if item["status"] not in {"active", "delayed", "completed", "sold", "written_off"}:
            raise ValueError(f"Neleistinas statusas: {item['status']}")


def run_import(config: PlatformConfig) -> int:
    configure(config)
    parser = argparse.ArgumentParser(description=f"Importuoja {CONFIG.name} Excel į Portfolio V2 JSON.")
    parser.add_argument("--input", help=f"{CONFIG.name} .xlsx failo kelias")
    parser.add_argument("--output", help="Rezultato .json failo kelias")
    args = parser.parse_args()

    input_path = find_input(args.input)
    output_path = Path(args.output).expanduser().resolve() if args.output else DEFAULT_OUTPUT

    workbook = load_workbook(input_path, data_only=True, read_only=True)
    try:
        overview_name = find_sheet_name(workbook, CONFIG.overview_sheet)
        cash_name = find_sheet_name(workbook, "Pinigai")
        if not overview_name:
            raise ValueError(f"Excel faile nerastas lapas '{CONFIG.overview_sheet}'.")
        if not cash_name:
            raise ValueError("Excel faile nerastas lapas 'Pinigai'.")

        epoch = workbook.epoch
        overview_ws = workbook[overview_name]
        cash_ws = workbook[cash_name]
        summary = read_summary(overview_ws)
        cash_flows = read_cash_flows(cash_ws, epoch)

        excluded_sheets = {overview_name, cash_name}
        raw_projects = [
            read_project(workbook[sheet_name], epoch)
            for sheet_name in workbook.sheetnames
            if sheet_name not in excluded_sheets
        ]
        investments, merged_codes = merge_projects(raw_projects)

        active = sum(item["status"] == "active" for item in investments)
        delayed = sum(item["status"] == "delayed" for item in investments)
        completed = sum(item["status"] == "completed" for item in investments)

        summary.update({
            "activeInvestments": active,
            "delayedInvestments": delayed,
            "completedInvestments": completed,
            "averageRate": weighted_average(investments, "interestRate"),
            "averageLtv": weighted_average(investments, "ltv"),
        })

        overview_date_col = find_header_column(overview_ws, "Data", row=2)
        all_dates = [
            excel_date(overview_ws.cell(row, overview_date_col).value, epoch)
            for row in range(3, overview_ws.max_row + 1)
        ]
        start_dates = [item for item in all_dates if item]
        start_date = min(start_dates) if start_dates else None
        updated_at = date.today()

        result = {
            "schemaVersion": SCHEMA_VERSION,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "platform": {
                "id": PLATFORM_SLUG,
                "slug": PLATFORM_SLUG,
                "name": CONFIG.name,
                "group": "real_estate",
                "type": "real_estate",
                "category": "NT sutelktinis finansavimas",
                "currency": "EUR",
                "active": True,
                "startDate": start_date.isoformat() if start_date else None,
                "updatedAt": updated_at.isoformat(),
                "website": CONFIG.website,
            },
            "summary": summary,
            "history": build_history(summary, updated_at),
            "investments": investments,
        }

        validate_output(result)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary = output_path.with_suffix(output_path.suffix + ".tmp")
        with temporary.open("w", encoding="utf-8") as file:
            json.dump(result, file, ensure_ascii=False, indent=2)
            file.write("\n")
        temporary.replace(output_path)

        print(f"\n{CONFIG.name.upper()} IMPORTAS BAIGTAS")
        print("==========================")
        print(f"Excel: {input_path}")
        print(f"JSON:  {output_path}")
        print(f"Projektų lapų: {len(raw_projects)}")
        print(f"Unikalių projektų: {len(investments)}")
        print(f"Aktyvių: {active}")
        print(f"Vėluojančių: {delayed}")
        print(f"Užbaigtų: {completed}")
        print(f"Vertė: {summary['currentValue']:.2f} EUR")
        if merged_codes:
            print("Sujungti kodai: " + ", ".join(merged_codes))
        print()
        return 0
    finally:
        workbook.close()

