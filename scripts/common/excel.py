from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .utils import is_month_sheet


def load_excel(path: Path):
    if not path.is_file():
        raise FileNotFoundError(f"Excel failas nerastas: {path}")

    return load_workbook(
        path,
        data_only=True,
        read_only=False,
    )


def require_sheet(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Nerastas privalomas lapas „{sheet_name}“."
        )

    return workbook[sheet_name]


def get_month_sheet_names(workbook) -> list[str]:
    return [
        name
        for name in workbook.sheetnames
        if is_month_sheet(name)
    ]


def get_instrument_sheet_names(
    workbook,
    *,
    excluded: set[str] | None = None,
) -> list[str]:
    excluded = excluded or set()

    return [
        name
        for name in workbook.sheetnames
        if name not in excluded and not is_month_sheet(name)
    ]
