from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "excel" / "Investavimas.xlsx"
DEFAULT_OUTPUT = ROOT / "public" / "data" / "platform_history.json"


def slugify(value: Any) -> str:
    text = str(value or "").strip().lower()
    replacements = str.maketrans({"ą":"a","č":"c","ę":"e","ė":"e","į":"i","š":"s","ų":"u","ū":"u","ž":"z"})
    text = text.translate(replacements)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    text = str(value).strip().replace("€", "").replace("%", "").replace(" ", "").replace(",", ".")
    try:
        return round(float(text), 4)
    except ValueError:
        return 0.0


def iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in (None, ""):
        return None
    text = str(value).strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def merged_value(ws, row: int, col: int) -> Any:
    cell = ws.cell(row=row, column=col)
    if cell.value not in (None, ""):
        return cell.value
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col).value
    return None


def find_sheet(wb):
    for preferred in ("Investavimas", "investavimas", "Istorija", "History"):
        if preferred in wb.sheetnames:
            return wb[preferred]
    return wb[wb.sheetnames[0]]


def find_date_column(ws) -> int:
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, min(ws.max_column, 12) + 1):
            if slugify(ws.cell(row, col).value) in {"data", "date", "menuo"}:
                return col
    return 1


def find_header_rows(ws) -> tuple[int, int]:
    metric_names = {"inesta", "investuota", "verte", "value", "graza", "return", "roi"}
    for metric_row in range(1, min(ws.max_row, 10) + 1):
        matches = 0
        for col in range(1, ws.max_column + 1):
            if slugify(merged_value(ws, metric_row, col)) in metric_names:
                matches += 1
        if matches >= 2:
            return max(1, metric_row - 1), metric_row
    raise RuntimeError("Nepavyko rasti platformų antraščių eilučių.")


def metric_kind(value: Any) -> str | None:
    key = slugify(value)
    if key in {"inesta", "investuota", "invested", "kapitalas"}:
        return "invested"
    if key in {"verte", "value", "dabartine-verte"}:
        return "value"
    if key in {"graza", "return", "roi", "proc"} or str(value or "").strip() == "%":
        return "returnRate"
    return None


def build_platform_history(excel_path: Path) -> dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    ws = find_sheet(wb)
    platform_row, metric_row = find_header_rows(ws)
    date_col = find_date_column(ws)

    columns: dict[str, dict[str, Any]] = {}
    last_platform = None
    for col in range(1, ws.max_column + 1):
        platform_name = merged_value(ws, platform_row, col)
        if platform_name not in (None, ""):
            last_platform = str(platform_name).strip()
        metric = metric_kind(merged_value(ws, metric_row, col))
        if not last_platform or not metric:
            continue
        slug = slugify(last_platform)
        if slug in {"data", "date"}:
            continue
        columns.setdefault(slug, {"name": last_platform, "columns": {}})["columns"][metric] = col

    platforms: dict[str, Any] = {}
    first_data_row = metric_row + 1
    for slug, meta in columns.items():
        rows = []
        for row in range(first_data_row, ws.max_row + 1):
            row_date = iso_date(ws.cell(row, date_col).value)
            if not row_date:
                continue
            invested = number(ws.cell(row, meta["columns"].get("invested", 0)).value) if meta["columns"].get("invested") else 0.0
            value = number(ws.cell(row, meta["columns"].get("value", 0)).value) if meta["columns"].get("value") else 0.0
            return_rate = number(ws.cell(row, meta["columns"].get("returnRate", 0)).value) if meta["columns"].get("returnRate") else (round((value-invested)/invested*100,4) if invested else 0.0)
            if invested == 0 and value == 0 and return_rate == 0:
                continue
            rows.append({
                "date": row_date,
                "invested": invested,
                "value": value,
                "profit": round(value - invested, 4),
                "returnRate": return_rate,
                "source": "Investavimas.xlsx",
            })
        if rows:
            platforms[slug] = {"name": meta["name"], "history": rows}

    return {
        "schemaVersion": 1,
        "type": "platform_history",
        "currency": "EUR",
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": {"file": excel_path.name, "sheet": ws.title},
        "platforms": platforms,
    }


def main() -> int:
    excel_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_EXCEL
    output_path = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else DEFAULT_OUTPUT
    if not excel_path.exists():
        print(f"KLAIDA: nerastas Excel failas: {excel_path}")
        return 1
    payload = build_platform_history(excel_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Sukurta: {output_path}")
    print(f"Platformų: {len(payload['platforms'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
