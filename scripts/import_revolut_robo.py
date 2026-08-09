from __future__ import annotations

import argparse
import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHEMA_VERSION = 1
EPSILON = 0.000001

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = PROJECT_ROOT / "excel" / "Revolut Robo.xlsx"
DEFAULT_OUTPUT = (
    PROJECT_ROOT / "public" / "data" / "platforms" / "revolut-robo.json"
)

OVERVIEW_SHEET = "Apžvalga"
CASH_SHEET = "Cash"
MONTH_SHEET_RE = re.compile(r"^\d{4}\.\d{2}$")
CLOSURE_RE = re.compile(
    r"uždaryta\s+(\d{4})-(\d{2})-(\d{2}).*?išmokėta\s+([\d.,]+)",
    re.IGNORECASE,
)


def finite_number(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if math.isfinite(number) else default


def rounded(value: Any, digits: int = 2) -> float:
    return round(finite_number(value), digits)


def month_key(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}-{value.month:02d}"

    if isinstance(value, (int, float)):
        text = f"{float(value):.2f}"
    else:
        text = str(value or "").strip().replace(",", ".")

    match = re.fullmatch(r"(\d{4})\.(\d{1,2})", text)
    if not match:
        return None

    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        return None

    return f"{year:04d}-{month:02d}"


def month_start(value: Any) -> str | None:
    key = month_key(value)
    return f"{key}-01" if key else None


def month_end(value: Any) -> str | None:
    key = month_key(value)
    if not key:
        return None

    year, month = map(int, key.split("-"))
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)

    return date.fromordinal(next_month.toordinal() - 1).isoformat()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def find_last_data_row(ws, start_row: int = 3) -> int:
    last_row = 0
    for row in range(start_row, ws.max_row + 1):
        if month_key(ws.cell(row, 1).value):
            last_row = row

    if not last_row:
        raise ValueError(f"Lape „{ws.title}“ nerasta mėnesinių duomenų.")

    return last_row


def find_closure(ws) -> tuple[str | None, float | None]:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue

            match = CLOSURE_RE.search(value)
            if not match:
                continue

            year, month, day, amount = match.groups()
            closure_date = f"{year}-{month}-{day}"
            payout = finite_number(amount.replace(",", "."))
            return closure_date, payout

    return None, None


def read_overview(ws) -> tuple[list[dict[str, Any]], list[str], int]:
    tickers: list[str] = []
    ticker_columns: list[int] = []

    # E:AG – visi Revolut Robo instrumentai.
    for col in range(5, 34):
        ticker = str(ws.cell(2, col).value or "").strip()
        if ticker:
            tickers.append(ticker)
            ticker_columns.append(col)

    last_row = find_last_data_row(ws)
    history: list[dict[str, Any]] = []
    previous_contributed = 0.0

    for row in range(3, last_row + 1):
        period = month_start(ws.cell(row, 1).value)
        if not period:
            continue

        contributed = finite_number(ws.cell(row, 2).value)
        net_invested = finite_number(ws.cell(row, 3).value)
        cash = finite_number(ws.cell(row, 4).value)
        holdings_value = finite_number(ws.cell(row, 34).value)
        current_value = finite_number(ws.cell(row, 35).value)
        profit = finite_number(ws.cell(row, 36).value)
        return_raw = ws.cell(row, 37).value
        dividends = finite_number(ws.cell(row, 38).value)
        fees = finite_number(ws.cell(row, 39).value)

        active_count = sum(
            1
            for col in ticker_columns
            if finite_number(ws.cell(row, col).value) > EPSILON
        )

        history.append({
            "date": period,
            "month": period[:7],
            "invested": rounded(contributed),
            "netInvested": rounded(net_invested),
            "holdingsValue": rounded(holdings_value),
            "currentValue": rounded(current_value),
            "profit": rounded(profit),
            "returnRate": (
                round(finite_number(return_raw) * 100, 4)
                if return_raw is not None
                else None
            ),
            "cash": rounded(cash),
            "income": rounded(dividends),
            "fees": rounded(fees),
            "contributions": rounded(contributed - previous_contributed),
            "withdrawals": 0.0,
            "activeInvestments": active_count,
            "delayedInvestments": 0,
            "completedInvestments": 0,
        })

        previous_contributed = contributed

    return history, tickers, last_row


def read_investment(
    ws,
    ticker: str,
    overview_value: float,
    force_completed: bool,
    closure_date: str | None,
) -> dict[str, Any]:
    rows = [
        row
        for row in range(5, ws.max_row + 1)
        if month_key(ws.cell(row, 1).value)
    ]

    if not rows:
        raise ValueError(
            f"{ticker}: investicijos lape nėra mėnesinių duomenų."
        )

    first_row = rows[0]
    last_row = rows[-1]

    full_name = str(ws["A1"].value or ticker).strip()
    name = full_name
    if full_name.upper().startswith(ticker.upper() + " "):
        name = full_name[len(ticker):].strip()

    historical_contributed = finite_number(ws.cell(last_row, 3).value)
    sheet_quantity = finite_number(ws.cell(last_row, 5).value)
    sheet_price = finite_number(ws.cell(last_row, 6).value)
    sheet_value = finite_number(ws.cell(last_row, 7).value)
    sheet_profit = finite_number(ws.cell(last_row, 8).value)
    return_raw = ws.cell(last_row, 9).value
    dividends = finite_number(ws.cell(last_row, 10).value)

    is_active = (
        not force_completed
        and sheet_quantity > 0.01
        and overview_value > EPSILON
    )

    return {
        "id": ticker.lower(),
        "slug": slugify(ticker),
        "ticker": ticker,
        "name": name,
        "fullName": full_name,
        "type": "ETF",
        "status": "active" if is_active else "completed",
        "currency": "EUR",
        "startDate": month_start(ws.cell(first_row, 1).value),
        "endDate": (
            None
            if is_active
            else closure_date or month_end(ws.cell(last_row, 1).value)
        ),
        "invested": rounded(historical_contributed),
        "netInvested": rounded(
            historical_contributed if is_active else 0.0
        ),
        "currentValue": rounded(overview_value if is_active else 0.0),
        "sheetValue": rounded(sheet_value),
        "profit": rounded(sheet_profit),
        "returnRate": (
            round(finite_number(return_raw) * 100, 4)
            if return_raw is not None
            else None
        ),
        "quantity": round(sheet_quantity if is_active else 0.0, 8),
        "price": rounded(sheet_price if is_active else 0.0, 6),
        "realizedProceeds": 0.0,
        "dividends": rounded(dividends),
        "fees": 0.0,
    }


def build_document(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        input_path,
        data_only=True,
        read_only=False,
    )

    if OVERVIEW_SHEET not in workbook.sheetnames:
        raise ValueError(f"Nerastas privalomas lapas „{OVERVIEW_SHEET}“.")

    overview = workbook[OVERVIEW_SHEET]
    history, tickers, last_row = read_overview(overview)
    closure_date, payout = find_closure(overview)
    force_completed = closure_date is not None and payout is not None

    missing_sheets = [
        ticker for ticker in tickers if ticker not in workbook.sheetnames
    ]
    if missing_sheets:
        raise ValueError(
            "Nerasti instrumentų lapai: " + ", ".join(missing_sheets)
        )

    investments: list[dict[str, Any]] = []

    for col, ticker in enumerate(tickers, start=5):
        overview_value = finite_number(overview.cell(last_row, col).value)
        investments.append(
            read_investment(
                workbook[ticker],
                ticker,
                overview_value,
                force_completed,
                closure_date,
            )
        )

    investments.sort(
        key=lambda item: (-item["currentValue"], item["ticker"])
    )

    latest = history[-1]
    active = [
        item for item in investments if item["status"] == "active"
    ]
    completed = [
        item for item in investments if item["status"] == "completed"
    ]

    total_contributed = rounded(latest["invested"])
    total_dividends = rounded(
        sum(finite_number(point["income"]) for point in history)
    )
    total_fees = rounded(
        sum(finite_number(point["fees"]) for point in history)
    )

    if force_completed:
        invested = 0.0
        holdings_value = 0.0
        cash = 0.0
        current_value = 0.0
        realized_proceeds = rounded(payout)
        realized_profit = rounded(realized_proceeds - total_contributed)
        profit = realized_profit
        return_rate = (
            round(realized_profit / total_contributed * 100, 4)
            if total_contributed > EPSILON
            else None
        )
    else:
        invested = rounded(
            sum(
                finite_number(item["netInvested"])
                for item in active
            )
        )
        holdings_value = rounded(
            sum(item["currentValue"] for item in active)
        )
        cash = rounded(latest["cash"])
        current_value = rounded(holdings_value + cash)
        realized_proceeds = 0.0
        profit = rounded(latest["profit"])
        realized_profit = rounded(
            sum(
                item["profit"]
                for item in completed
            )
        )
        return_rate = latest["returnRate"]

    latest["totalContributed"] = total_contributed
    latest["invested"] = invested
    latest["netInvested"] = invested
    latest["holdingsValue"] = holdings_value
    latest["cash"] = cash
    latest["currentValue"] = current_value
    latest["profit"] = profit
    latest["returnRate"] = return_rate
    latest["activeInvestments"] = len(active)
    latest["completedInvestments"] = len(completed)

    overall_xirr_raw = overview.cell(2, 29).value
    overall_xirr = (
        round(finite_number(overall_xirr_raw) * 100, 4)
        if overall_xirr_raw not in (None, "")
        else None
    )

    document = {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAt": datetime.now().astimezone().isoformat(
            timespec="seconds"
        ),
        "platform": {
            "id": "revolut-robo",
            "slug": "revolut-robo",
            "name": "Revolut Robo",
            "group": "robo",
            "type": "robo-advisor",
            "category": "Robo Advisor ETF portfelis",
            "currency": "EUR",
            "active": bool(active),
            "startDate": history[0]["date"],
            "updatedAt": closure_date or month_end(
                overview.cell(last_row, 1).value
            ),
            "website": "https://www.revolut.com",
        },
        "summary": {
            "invested": invested,
            "netInvested": invested,
            "totalContributed": total_contributed,
            "holdingsValue": holdings_value,
            "currentValue": current_value,
            "profit": profit,
            "realizedProfit": realized_profit,
            "realizedProceeds": realized_proceeds,
            "returnRate": return_rate,
            "xirr": overall_xirr,
            "cash": cash,
            "incomeReceived": total_dividends,
            "fees": total_fees,
            "activeInvestments": len(active),
            "delayedInvestments": 0,
            "completedInvestments": len(completed),
            "averageRate": None,
            "averageLtv": None,
            "totalInvestments": len(investments),
        },
        "history": history,
        "investments": investments,
        "distributions": {
            "status": [
                {
                    "key": "active",
                    "label": "Aktyvios",
                    "count": len(active),
                    "value": rounded(
                        sum(item["currentValue"] for item in active)
                    ),
                },
                {
                    "key": "completed",
                    "label": "Parduotos",
                    "count": len(completed),
                    "value": realized_proceeds,
                },
            ],
            "holdings": [
                {
                    "key": item["ticker"],
                    "label": item["ticker"],
                    "name": item["name"],
                    "value": item["currentValue"],
                    "weight": (
                        round(
                            item["currentValue"]
                            / holdings_value
                            * 100,
                            4,
                        )
                        if holdings_value > EPSILON
                        else 0.0
                    ),
                }
                for item in active
            ],
        },
        "latestMonth": latest,
        "largestInvestment": (
            {
                "id": active[0]["id"],
                "ticker": active[0]["ticker"],
                "name": active[0]["name"],
                "currentValue": active[0]["currentValue"],
            }
            if active
            else None
        ),
        "source": {
            "file": input_path.name,
            "overviewSheet": OVERVIEW_SHEET,
            "instrumentSheets": tickers,
            "monthlySheets": [
                name
                for name in workbook.sheetnames
                if MONTH_SHEET_RE.fullmatch(name)
            ],
            "closureDate": closure_date,
            "closurePayout": rounded(payout) if payout is not None else None,
        },
    }

    validate_document(document)
    return document


def validate_document(document: dict[str, Any]) -> None:
    summary = document["summary"]
    investments = document["investments"]
    errors: list[str] = []

    if not document["history"]:
        errors.append("History yra tuščias.")

    if not investments:
        errors.append("Investicijų sąrašas yra tuščias.")

    active_count = sum(
        item["status"] == "active" for item in investments
    )
    completed_count = sum(
        item["status"] == "completed" for item in investments
    )

    if active_count != summary["activeInvestments"]:
        errors.append("Nesutampa aktyvių investicijų skaičius.")

    if completed_count != summary["completedInvestments"]:
        errors.append("Nesutampa parduotų investicijų skaičius.")

    if not document["platform"]["active"]:
        if abs(summary["invested"]) > 0.01:
            errors.append("Uždaryto portfelio invested turi būti 0.")
        if abs(summary["currentValue"]) > 0.01:
            errors.append("Uždaryto portfelio currentValue turi būti 0.")
        if active_count != 0:
            errors.append("Uždarytas portfelis negali turėti aktyvių pozicijų.")

    holdings_value = rounded(
        sum(
            item["currentValue"]
            for item in investments
            if item["status"] == "active"
        )
    )

    if abs(holdings_value - summary["holdingsValue"]) > 0.05:
        errors.append("Nesutampa aktyvių pozicijų vertė.")

    ids = [item["id"] for item in investments]
    if len(ids) != len(set(ids)):
        errors.append("Rasti dubliuoti investicijų ID.")

    if errors:
        raise ValueError("\n".join(errors))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Revolut Robo Excel importeris į Portfolio V2 JSON."
    )

    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Excel failas. Numatyta: {DEFAULT_INPUT}",
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"JSON failas. Numatyta: {DEFAULT_OUTPUT}",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    print("=" * 64)
    print("REVOLUT ROBO IMPORTER V1")
    print("=" * 64)
    print(f"Excel failas: {input_path}")

    if not input_path.is_file():
        raise FileNotFoundError(f"Excel failas nerastas: {input_path}")

    document = build_document(input_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8") as file:
        json.dump(
            document,
            file,
            ensure_ascii=False,
            indent=2,
        )

    summary = document["summary"]

    print(f"✅ Nuskaityta pozicijų: {summary['totalInvestments']}")
    print(f"✅ Aktyvių: {summary['activeInvestments']}")
    print(f"✅ Parduotų: {summary['completedInvestments']}")
    print(f"✅ Investuota dabar: {summary['invested']:.2f} EUR")
    print(
        "✅ Istoriškai įnešta: "
        f"{summary['totalContributed']:.2f} EUR"
    )
    print(f"✅ Dabartinė vertė: {summary['currentValue']:.2f} EUR")
    print(f"✅ Realizuotas pelnas: {summary['realizedProfit']:.2f} EUR")
    print(f"✅ Išmokėta: {summary['realizedProceeds']:.2f} EUR")
    print(f"✅ JSON sukurtas: {output_path}")


if __name__ == "__main__":
    main()
