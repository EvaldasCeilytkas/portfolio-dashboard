from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base_importer import (
    build_standard_result,
    format_period,
    safe_number,
)


OVERVIEW_SHEET = "Apžvalga"
OVERVIEW_FIRST_DATA_ROW = 3
ETF_SUMMARY_ROW = 5

MONTH_SHEET_PATTERN = re.compile(r"^\d{4}\.\d{2}$")



def is_month_sheet(sheet_name: str) -> bool:
    return bool(MONTH_SHEET_PATTERN.fullmatch(str(sheet_name).strip()))


def find_last_overview_row(ws) -> int:
    for row_number in range(ws.max_row, OVERVIEW_FIRST_DATA_ROW - 1, -1):
        if ws.cell(row=row_number, column=1).value not in (None, ""):
            return row_number

    raise ValueError("Revolut Brokerage faile nerasta Apžvalgos istorija.")


def get_overview_symbols(ws) -> dict[str, int]:
    symbols: dict[str, int] = {}

    # Apžvalgos lape ETF simboliai yra 2 eilutėje nuo F stulpelio
    # iki stulpelio prieš bendrą „Vertė“.
    for column_number in range(6, ws.max_column + 1):
        header = str(ws.cell(row=2, column=column_number).value or "").strip()

        if header == "Vertė":
            break

        if header:
            symbols[header.upper()] = column_number

    return symbols


def build_monthly_history(ws) -> list[dict[str, Any]]:
    history: list[dict[str, Any]] = []

    for row_number in range(OVERVIEW_FIRST_DATA_ROW, ws.max_row + 1):
        period = format_period(ws.cell(row=row_number, column=1).value)

        if not period:
            continue

        deposited = safe_number(ws.cell(row=row_number, column=2).value)
        withdrawn = safe_number(ws.cell(row=row_number, column=3).value)
        invested = safe_number(ws.cell(row=row_number, column=4).value)
        cash = safe_number(ws.cell(row=row_number, column=5).value)
        value = safe_number(ws.cell(row=row_number, column=18).value)
        profit = safe_number(ws.cell(row=row_number, column=19).value)
        return_rate = safe_number(ws.cell(row=row_number, column=20).value)
        dividends = safe_number(ws.cell(row=row_number, column=21).value)
        buy_fees = safe_number(ws.cell(row=row_number, column=22).value)
        sell_fees = safe_number(ws.cell(row=row_number, column=23).value)

        previous_value = history[-1]["value"] if history else 0.0
        previous_invested = history[-1]["invested"] if history else 0.0
        cash_flow = invested - previous_invested

        history.append(
            {
                "date": f"{period}-01",
                "period": period,
                "deposited": round(deposited, 2),
                "withdrawn": round(withdrawn, 2),
                "invested": round(invested, 2),
                "cash": round(cash, 2),
                "value": round(value, 2),
                "profit": round(profit, 2),
                "returnRate": round(return_rate * 100, 2),
                "dividends": round(dividends, 2),
                "buyFees": round(buy_fees, 2),
                "sellFees": round(sell_fees, 2),
                "previousValue": round(previous_value, 2),
                "cashFlow": round(cash_flow, 2),
                "monthlyProfit": round(
                    value - previous_value - cash_flow,
                    2,
                ),
            }
        )

    return history


def build_position_history(ws) -> list[dict[str, Any]]:
    history: list[dict[str, Any]] = []

    for row_number in range(ETF_SUMMARY_ROW, ws.max_row + 1):
        period = format_period(ws.cell(row=row_number, column=1).value)

        if not period:
            continue

        invested = safe_number(ws.cell(row=row_number, column=4).value)
        value = safe_number(ws.cell(row=row_number, column=8).value)
        profit = safe_number(ws.cell(row=row_number, column=9).value)
        return_rate = safe_number(ws.cell(row=row_number, column=10).value)

        history.append(
            {
                "date": f"{period}-01",
                "invested": round(invested, 2),
                "value": round(value, 2),
                "profit": round(profit, 2),
                "returnRate": round(return_rate * 100, 2),
            }
        )

    return history



def format_activity_date(value: Any) -> str:
    """Paverčia sandorio datą į YYYY-MM-DD formatą."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    text = str(value or "").strip()

    if not text:
        return ""

    # Leidžia naudoti ir tekstines YYYY-MM-DD reikšmes.
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text

    return ""


def get_block_symbol(value: Any, known_symbols: set[str]) -> str:
    """
    Iš mėnesio lapo pozicijos antraštės nustato tickerį.

    Pvz. „SPYL SPDR S&P 500 UCITS (Acc)“ -> „SPYL“.
    """
    title = str(value or "").strip()

    if not title:
        return ""

    first_word = title.split()[0].upper()
    return first_word if first_word in known_symbols else ""


def build_position_activity(
    workbook,
    known_symbols: set[str],
) -> dict[str, dict[str, list[dict[str, Any]]]]:
    """
    Iš YYYY.MM lapų surenka tikrus pirkimus, pardavimus ir dividendus.

    Mėnesio lapuose kiekvienos pozicijos blokas turi:
    A:F  – pirkimus,
    H:L  – pardavimus,
    N:O  – dividendus.
    """
    activity = {
        symbol: {
            "transactions": [],
            "dividends": [],
        }
        for symbol in known_symbols
    }

    for sheet_name in workbook.sheetnames:
        if not is_month_sheet(sheet_name):
            continue

        ws = workbook[sheet_name]
        row_number = 1

        while row_number <= ws.max_row:
            symbol = get_block_symbol(
                ws.cell(row=row_number, column=1).value,
                known_symbols,
            )

            # Tikras pozicijos blokas: antraštė, tada „Pirkimas“.
            if (
                not symbol
                or str(
                    ws.cell(row=row_number + 1, column=1).value or ""
                ).strip()
                != "Pirkimas"
            ):
                row_number += 1
                continue

            data_row = row_number + 3

            while data_row <= ws.max_row:
                first_value = str(
                    ws.cell(row=data_row, column=1).value or ""
                ).strip()

                if first_value == "Viso:":
                    break

                buy_date = format_activity_date(
                    ws.cell(row=data_row, column=1).value
                )
                sell_date = format_activity_date(
                    ws.cell(row=data_row, column=8).value
                )
                dividend_date = format_activity_date(
                    ws.cell(row=data_row, column=14).value
                )

                if buy_date:
                    amount = safe_number(
                        ws.cell(row=data_row, column=2).value
                    )
                    invested = safe_number(
                        ws.cell(row=data_row, column=3).value
                    )
                    fee = safe_number(
                        ws.cell(row=data_row, column=4).value
                    )
                    price = safe_number(
                        ws.cell(row=data_row, column=5).value
                    )
                    quantity = safe_number(
                        ws.cell(row=data_row, column=6).value
                    )

                    activity[symbol]["transactions"].append(
                        {
                            "id": (
                                f"{symbol}-buy-{buy_date}-"
                                f"{data_row}-{sheet_name}"
                            ),
                            "date": buy_date,
                            "type": "buy",
                            "amount": round(amount, 2),
                            "netAmount": round(invested, 2),
                            "fee": round(fee, 2),
                            "price": round(price, 6),
                            "quantity": round(quantity, 8),
                            "currency": "EUR",
                            "sourcePeriod": sheet_name,
                        }
                    )

                if sell_date:
                    amount = safe_number(
                        ws.cell(row=data_row, column=9).value
                    )
                    fee = safe_number(
                        ws.cell(row=data_row, column=10).value
                    )
                    price = safe_number(
                        ws.cell(row=data_row, column=11).value
                    )
                    quantity = safe_number(
                        ws.cell(row=data_row, column=12).value
                    )

                    activity[symbol]["transactions"].append(
                        {
                            "id": (
                                f"{symbol}-sell-{sell_date}-"
                                f"{data_row}-{sheet_name}"
                            ),
                            "date": sell_date,
                            "type": "sell",
                            "amount": round(amount, 2),
                            "netAmount": round(amount - fee, 2),
                            "fee": round(fee, 2),
                            "price": round(price, 6),
                            "quantity": round(quantity, 8),
                            "currency": "EUR",
                            "sourcePeriod": sheet_name,
                        }
                    )

                if dividend_date:
                    gross = safe_number(
                        ws.cell(row=data_row, column=15).value
                    )

                    activity[symbol]["dividends"].append(
                        {
                            "id": (
                                f"{symbol}-dividend-{dividend_date}-"
                                f"{data_row}-{sheet_name}"
                            ),
                            "date": dividend_date,
                            "gross": round(gross, 2),
                            "tax": 0.0,
                            "net": round(gross, 2),
                            "currency": "EUR",
                            "sourcePeriod": sheet_name,
                        }
                    )

                data_row += 1

            row_number = max(row_number + 1, data_row + 1)

    for symbol_activity in activity.values():
        symbol_activity["transactions"].sort(
            key=lambda item: (item["date"], item["type"], item["id"])
        )
        symbol_activity["dividends"].sort(
            key=lambda item: (item["date"], item["id"])
        )

    return activity


def build_position_events(
    symbol: str,
    transactions: list[dict[str, Any]],
    dividends: list[dict[str, Any]],
    history: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Sukuria universalią Investment Profile laiko juostą."""
    events: list[dict[str, Any]] = []
    buy_transactions = [
        item for item in transactions if item.get("type") == "buy"
    ]

    for index, transaction in enumerate(transactions):
        transaction_type = transaction.get("type")

        events.append(
            {
                "id": f"event-{transaction['id']}",
                "date": transaction.get("date", ""),
                "type": (
                    "opened"
                    if transaction_type == "buy" and index == 0
                    else transaction_type
                ),
                "title": (
                    "Pirmasis pirkimas"
                    if transaction_type == "buy" and index == 0
                    else (
                        "Pozicijos papildymas"
                        if transaction_type == "buy"
                        else "Pozicijos pardavimas"
                    )
                ),
                "amount": transaction.get("netAmount", 0),
                "quantity": transaction.get("quantity", 0),
                "price": transaction.get("price", 0),
                "currency": transaction.get("currency", "EUR"),
            }
        )

    for dividend in dividends:
        events.append(
            {
                "id": f"event-{dividend['id']}",
                "date": dividend.get("date", ""),
                "type": "dividend",
                "title": "Gauti dividendai",
                "amount": dividend.get("net", 0),
                "currency": dividend.get("currency", "EUR"),
            }
        )

    valid_history = [
        item
        for item in history
        if item.get("date") and safe_number(item.get("value")) > 0
    ]

    if valid_history:
        highest_point = max(
            valid_history,
            key=lambda item: safe_number(item.get("value")),
        )

        events.append(
            {
                "id": f"{symbol}-all-time-high",
                "date": highest_point.get("date", ""),
                "type": "allTimeHigh",
                "title": "Didžiausia istorinė vertė",
                "amount": round(
                    safe_number(highest_point.get("value")),
                    2,
                ),
                "currency": "EUR",
            }
        )

    # Pašaliname visiškai tuščius įvykius ir išrikiuojame chronologiškai.
    events = [event for event in events if event.get("date")]
    events.sort(key=lambda item: (item["date"], item["id"]))

    return events


def load_position(
    ws,
    symbol: str,
    current_value: float,
    position_activity: dict[str, list[dict[str, Any]]] | None = None,
) -> dict[str, Any]:
    name = str(ws.cell(row=1, column=1).value or symbol).strip()

    invested = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=16).value)
    quantity = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=17).value)
    sold_value = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=18).value)
    summary_profit = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=19).value)
    summary_return = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=20).value)
    xirr = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=21).value)
    buy_fees = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=22).value)
    sell_fees = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=23).value)
    dividends = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=24).value)
    average_price = safe_number(ws.cell(row=ETF_SUMMARY_ROW, column=25).value)

    # Aktyvumą nustatome pagal paskutinę vertę Apžvalgos lape.
    # Q stulpelyje esantis kiekis yra istorinis nupirktas kiekis,
    # todėl parduotoms pozicijoms jis taip pat gali būti didesnis už nulį.
    active = current_value > 0.005

    if active:
        profit = current_value - invested + dividends
        return_rate = (profit / invested * 100) if invested else 0.0
    else:
        profit = summary_profit
        return_rate = summary_return * 100

    history = build_position_history(ws)
    activity = position_activity or {}
    transactions = list(activity.get("transactions") or [])
    dividend_history = list(activity.get("dividends") or [])
    events = build_position_events(
        symbol,
        transactions,
        dividend_history,
        history,
    )

    total_dividend_income = sum(
        safe_number(item.get("net"))
        for item in dividend_history
    )
    holding_since = (
        transactions[0].get("date", "")
        if transactions
        else (history[0].get("date", "") if history else "")
    )
    current_price = (
        current_value / quantity
        if active and quantity > 0
        else 0.0
    )

    summary = {
        "invested": round(invested, 2),
        "value": round(current_value, 2),
        "profit": round(profit, 2),
        "returnRate": round(return_rate, 2),
        "xirr": round(xirr * 100, 2),
        "quantity": round(quantity, 8),
        "averagePrice": round(average_price, 4),
        "currentPrice": round(current_price, 6),
        "dividendIncome": round(total_dividend_income, 2),
        "buyFees": round(buy_fees, 2),
        "sellFees": round(sell_fees, 2),
        "fees": round(buy_fees + sell_fees, 2),
    }

    return {
        "symbol": symbol,
        "ticker": symbol,
        "name": name,
        "active": active,
        "status": "active" if active else "sold",
        "currency": "EUR",
        "invested": round(invested, 2),
        "value": round(current_value, 2),
        "currentValue": round(current_value, 2),
        "soldValue": round(sold_value, 2),
        "sold": round(sold_value, 2),
        "quantity": round(quantity, 8),
        "averagePrice": round(average_price, 4),
        "averageBuyPrice": round(average_price, 4),
        "currentPrice": round(current_price, 6),
        "holdingSince": holding_since,
        "profit": round(profit, 2),
        "returnRate": round(return_rate, 2),
        "xirr": round(xirr * 100, 2),
        "dividendIncome": round(total_dividend_income, 2),
        "legacyDividendTotal": round(dividends, 2),
        "buyFees": round(buy_fees, 2),
        "sellFees": round(sell_fees, 2),
        "fees": round(buy_fees + sell_fees, 2),
        "summary": summary,
        "history": history,
        "transactions": transactions,
        "dividends": dividend_history,
        "events": events,
    }


def load_revolut_brokerage(file_path: str | Path) -> dict[str, Any]:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Nerastas Revolut Brokerage failas: {path}")

    workbook = load_workbook(path, data_only=True)

    if OVERVIEW_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Revolut Brokerage faile nerastas lapas „{OVERVIEW_SHEET}“."
        )

    overview = workbook[OVERVIEW_SHEET]
    last_row = find_last_overview_row(overview)
    symbol_columns = get_overview_symbols(overview)
    monthly_history = build_monthly_history(overview)
    position_activity = build_position_activity(
        workbook,
        set(symbol_columns),
    )

    positions: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name == OVERVIEW_SHEET or is_month_sheet(sheet_name):
            continue

        symbol = sheet_name.strip().upper()
        current_value = 0.0

        if symbol in symbol_columns:
            current_value = safe_number(
                overview.cell(
                    row=last_row,
                    column=symbol_columns[symbol],
                ).value
            )

        positions.append(
            load_position(
                workbook[sheet_name],
                symbol,
                current_value,
                position_activity.get(symbol),
            )
        )

    active_positions = [
        position for position in positions if position["active"]
    ]
    sold_positions = [
        position for position in positions if not position["active"]
    ]

    active_value = sum(position["value"] for position in active_positions)

    for position in active_positions:
        position["share"] = round(
            position["value"] / active_value * 100
            if active_value > 0
            else 0,
            2,
        )

    latest = monthly_history[-1] if monthly_history else {}

    total_dividends = sum(
        safe_number(item.get("dividends"))
        for item in monthly_history
    )
    total_buy_fees = sum(
        safe_number(item.get("buyFees"))
        for item in monthly_history
    )
    total_sell_fees = sum(
        safe_number(item.get("sellFees"))
        for item in monthly_history
    )


    summary = {
        "deposited": round(
            safe_number(latest.get("deposited")),
            2,
        ),
        "withdrawn": round(
            safe_number(latest.get("withdrawn")),
            2,
        ),
        "invested": round(
            safe_number(latest.get("invested")),
            2,
        ),
        "cash": round(
            safe_number(latest.get("cash")),
            2,
        ),
        "value": round(
            safe_number(latest.get("value")),
            2,
        ),
        "profit": round(
            safe_number(latest.get("profit")),
            2,
        ),
        "returnRate": round(
            safe_number(latest.get("returnRate")),
            2,
        ),
        "dividends": round(total_dividends, 2),
        "buyFees": round(total_buy_fees, 2),
        "sellFees": round(total_sell_fees, 2),
        "fees": round(
            total_buy_fees + total_sell_fees,
            2,
        ),
    }

    return build_standard_result(
        platform_name="Revolut Brokerage",
        source_file=path,
        history=monthly_history,
        active_positions=active_positions,
        sold_positions=sold_positions,
        summary=summary,
        module_type="brokerage",
        cashflow={
            "deposited": summary["deposited"],
            "withdrawn": summary["withdrawn"],
            "dividends": summary["dividends"],
            "buyFees": summary["buyFees"],
            "sellFees": summary["sellFees"],
            "fees": summary["fees"],
            "cash": summary["cash"],
        },
        updated_at=latest.get("date", ""),
    )
