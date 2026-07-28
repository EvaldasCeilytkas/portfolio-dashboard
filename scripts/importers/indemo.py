"""Indemo NPL / mortgage investments Excel importer."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base_importer import (
    build_standard_result,
    create_slug,
    format_date,
    safe_number,
)


OVERVIEW_SHEET_NAMES = ("Overwiev", "Overview")
MONEY_SHEET_NAME = "Pinigai"
IGNORED_SHEETS = {*OVERVIEW_SHEET_NAMES, MONEY_SHEET_NAME}


def _first_existing_sheet(workbook, names):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]

    raise ValueError(
        "Indemo faile nerastas suvestinės lapas: "
        + " arba ".join(names)
    )


def _date_key(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    return format_date(value)


def _month_key(value: Any) -> str:
    text = _date_key(value)
    return text[:7] if len(text) >= 7 else ""


def _round(value: Any, digits: int = 2) -> float:
    return round(safe_number(value), digits)


def _is_project_sheet(ws) -> bool:
    return (
        str(ws["A1"].value or "").strip().lower() == "mortgage loan id"
        and bool(ws["B1"].value)
    )


def _read_overview(ws) -> dict[str, Any]:
    projects = []
    row_number = 3

    while row_number <= ws.max_row:
        project_id = str(ws.cell(row_number, 2).value or "").strip()

        if not project_id:
            row_number += 1
            continue

        invested = _round(ws.cell(row_number, 3).value)
        invested_date = _date_key(ws.cell(row_number, 4).value)
        returned_date = _date_key(ws.cell(row_number, 5).value)
        duration_days = _round(ws.cell(row_number, 6).value)
        returned_principal = _round(ws.cell(row_number, 7).value)
        interest = _round(ws.cell(row_number, 8).value)
        xirr_value = safe_number(ws.cell(row_number, 9).value)

        projects.append(
            {
                "id": project_id,
                "slug": create_slug(project_id),
                "invested": invested,
                "investedDate": invested_date,
                "returnedDate": returned_date,
                "durationDays": duration_days,
                "returnedPrincipal": returned_principal,
                "interest": interest,
                "xirr": round(xirr_value * 100, 2) if abs(xirr_value) <= 2 else round(xirr_value, 2),
            }
        )

        row_number += 1

    summary = {
        "deposited": _round(ws["K3"].value),
        "invested": _round(ws["L3"].value),
        "bonuses": _round(ws["M3"].value),
        "cash": _round(ws["N3"].value),
        "returnedPrincipal": _round(ws["O3"].value),
        "interest": _round(ws["P3"].value),
        "value": _round(ws["Q3"].value),
        "profit": _round(ws["R3"].value),
        "returnRate": round(safe_number(ws["S3"].value) * 100, 2),
        "xirr": round(safe_number(ws["T3"].value) * 100, 2),
    }

    return {
        "projects": projects,
        "summary": summary,
    }


def _read_money_events(ws) -> list[dict[str, Any]]:
    events = []

    for row_number in range(2, ws.max_row + 1):
        event_date = ws.cell(row_number, 1).value

        if not event_date:
            continue

        deposited = _round(ws.cell(row_number, 2).value)
        withdrawn = _round(ws.cell(row_number, 3).value)
        bonus = _round(ws.cell(row_number, 4).value)

        if deposited == 0 and withdrawn == 0 and bonus == 0:
            continue

        events.append(
            {
                "date": _date_key(event_date),
                "deposited": deposited,
                "withdrawn": withdrawn,
                "bonus": bonus,
                "netCashFlow": round(deposited - withdrawn, 2),
            }
        )

    events.sort(key=lambda item: item["date"])
    return events


def _read_project_sheet(ws) -> dict[str, Any]:
    project_id = str(ws["B1"].value or ws.title).strip()
    project_name = str(ws["B2"].value or project_id).strip()

    ptv = safe_number(ws["B3"].value)
    pdt = safe_number(ws["B4"].value)

    investment_date = _date_key(ws["B7"].value)
    loan_start_date = _date_key(ws["B8"].value)
    returned_date = _date_key(ws["B9"].value)

    invested = _round(ws["B10"].value)
    returned_principal = _round(ws["B11"].value)
    interest = _round(ws["B12"].value)
    xirr_value = safe_number(ws["B13"].value)
    duration_days = _round(ws["B14"].value)

    transactions = []

    for row_number in range(3, min(ws.max_row, 20) + 1):
        sequence = ws.cell(row_number, 4).value
        note_id = str(ws.cell(row_number, 5).value or "").strip()
        transaction_date = ws.cell(row_number, 6).value
        amount = _round(ws.cell(row_number, 7).value)
        repayment_date = ws.cell(row_number, 8).value
        principal = _round(ws.cell(row_number, 9).value)
        transaction_interest = _round(ws.cell(row_number, 10).value)
        profit_rate = safe_number(ws.cell(row_number, 11).value)
        transaction_duration = _round(ws.cell(row_number, 12).value)

        has_transaction = bool(
            note_id
            or transaction_date
            or amount
            or principal
            or transaction_interest
        )

        if not has_transaction:
            continue

        transactions.append(
            {
                "sequence": int(sequence) if isinstance(sequence, (int, float)) else sequence,
                "noteId": note_id,
                "date": _date_key(transaction_date),
                "amount": amount,
                "repaymentDate": _date_key(repayment_date),
                "returnedPrincipal": principal,
                "interest": transaction_interest,
                "profitRate": round(profit_rate * 100, 2) if abs(profit_rate) <= 2 else round(profit_rate, 2),
                "durationDays": transaction_duration,
                "status": "completed" if principal >= amount and amount > 0 else "active",
            }
        )

    if not invested:
        invested = round(sum(item["amount"] for item in transactions), 2)

    if not returned_principal:
        returned_principal = round(
            sum(item["returnedPrincipal"] for item in transactions),
            2,
        )

    if not interest:
        interest = round(sum(item["interest"] for item in transactions), 2)

    remaining = round(max(invested - returned_principal, 0), 2)
    profit = round(returned_principal + interest - invested, 2)
    status = "completed" if invested > 0 and remaining <= 0.01 else "active"

    return {
        "id": project_id,
        "slug": create_slug(project_id),
        "name": project_name,
        "investmentType": "mortgage",
        "ptv": round(ptv * 100, 2) if abs(ptv) <= 2 else round(ptv, 2),
        "pdt": round(pdt * 100, 2) if abs(pdt) <= 2 else round(pdt, 2),
        "investmentDate": investment_date,
        "loanStartDate": loan_start_date,
        "returnedDate": returned_date,
        "invested": invested,
        "returnedPrincipal": returned_principal,
        "interest": interest,
        "remaining": remaining,
        "profit": profit,
        "xirr": round(xirr_value * 100, 2) if abs(xirr_value) <= 2 else round(xirr_value, 2),
        "durationDays": duration_days,
        "status": status,
        "transactions": transactions,
        "transactionCount": len(transactions),
    }


def _build_history(
    projects: list[dict[str, Any]],
    money_events: list[dict[str, Any]],
    summary: dict[str, Any],
) -> list[dict[str, Any]]:
    monthly = defaultdict(
        lambda: {
            "deposited": 0.0,
            "withdrawn": 0.0,
            "bonuses": 0.0,
            "invested": 0.0,
            "returnedPrincipal": 0.0,
            "interest": 0.0,
        }
    )

    for event in money_events:
        month = _month_key(event["date"])

        if not month:
            continue

        monthly[month]["deposited"] += event["deposited"]
        monthly[month]["withdrawn"] += event["withdrawn"]
        monthly[month]["bonuses"] += event["bonus"]

    for project in projects:
        for transaction in project.get("transactions", []):
            investment_month = _month_key(transaction.get("date"))

            if investment_month:
                monthly[investment_month]["invested"] += transaction["amount"]

            repayment_month = _month_key(transaction.get("repaymentDate"))

            if repayment_month:
                monthly[repayment_month]["returnedPrincipal"] += transaction[
                    "returnedPrincipal"
                ]
                monthly[repayment_month]["interest"] += transaction["interest"]

    months = sorted(monthly)

    if not months:
        return []

    cumulative_deposited = 0.0
    cumulative_withdrawn = 0.0
    cumulative_bonuses = 0.0
    cumulative_invested = 0.0
    cumulative_returned = 0.0
    cumulative_interest = 0.0
    history = []

    for month in months:
        point = monthly[month]
        cumulative_deposited += point["deposited"]
        cumulative_withdrawn += point["withdrawn"]
        cumulative_bonuses += point["bonuses"]
        cumulative_invested += point["invested"]
        cumulative_returned += point["returnedPrincipal"]
        cumulative_interest += point["interest"]

        remaining_principal = max(cumulative_invested - cumulative_returned, 0)
        cash = (
            cumulative_deposited
            - cumulative_withdrawn
            + cumulative_bonuses
            - cumulative_invested
            + cumulative_returned
            + cumulative_interest
        )
        value = remaining_principal + cash
        invested_capital = cumulative_deposited - cumulative_withdrawn
        profit = value - invested_capital

        history.append(
            {
                "date": f"{month}-01",
                "invested": round(invested_capital, 2),
                "value": round(value, 2),
                "profit": round(profit, 2),
                "returnRate": round(
                    profit / invested_capital * 100
                    if invested_capital
                    else 0,
                    2,
                ),
                "cashFlow": round(
                    point["deposited"] - point["withdrawn"],
                    2,
                ),
                "monthlyProfit": round(
                    point["bonuses"] + point["interest"],
                    2,
                ),
                "deposited": round(cumulative_deposited, 2),
                "withdrawn": round(cumulative_withdrawn, 2),
                "bonuses": round(cumulative_bonuses, 2),
                "returnedPrincipal": round(cumulative_returned, 2),
                "interest": round(cumulative_interest, 2),
                "cash": round(cash, 2),
                "remainingPrincipal": round(remaining_principal, 2),
            }
        )

    # Galutinis taškas sulyginamas su Excel suvestine.
    latest = history[-1]
    latest["invested"] = round(summary.get("deposited", latest["invested"]), 2)
    latest["value"] = round(summary.get("value", latest["value"]), 2)
    latest["profit"] = round(summary.get("profit", latest["profit"]), 2)
    latest["returnRate"] = round(
        summary.get("returnRate", latest["returnRate"]),
        2,
    )
    latest["cash"] = round(summary.get("cash", latest["cash"]), 2)
    latest["returnedPrincipal"] = round(
        summary.get("returnedPrincipal", latest["returnedPrincipal"]),
        2,
    )
    latest["interest"] = round(
        summary.get("interest", latest["interest"]),
        2,
    )

    return history


def _build_distributions(projects: list[dict[str, Any]]) -> dict[str, Any]:
    status_distribution = defaultdict(float)
    type_distribution = defaultdict(float)
    ptv_buckets = defaultdict(float)

    for project in projects:
        value = project["remaining"]
        status_distribution[project["status"]] += value

        project_prefix = project["id"][:1].upper()
        type_distribution[project_prefix] += value

        ptv = project.get("ptv", 0)

        if ptv <= 40:
            bucket = "≤ 40 %"
        elif ptv <= 55:
            bucket = "41–55 %"
        elif ptv <= 70:
            bucket = "56–70 %"
        else:
            bucket = "> 70 %"

        ptv_buckets[bucket] += value

    def rows(values):
        return [
            {"name": key, "value": round(value, 2)}
            for key, value in sorted(
                values.items(),
                key=lambda item: item[1],
                reverse=True,
            )
        ]

    return {
        "status": rows(status_distribution),
        "projectSeries": rows(type_distribution),
        "ptv": rows(ptv_buckets),
    }


def load_indemo(source_file: str | Path) -> dict[str, Any]:
    """Nuskaito Indemo Excel failą į bendrą platformos rezultatą."""
    source_path = Path(source_file)

    if not source_path.exists():
        raise FileNotFoundError(f"Nerastas Indemo failas: {source_path}")

    workbook = load_workbook(source_path, data_only=True)
    overview_ws = _first_existing_sheet(workbook, OVERVIEW_SHEET_NAMES)

    if MONEY_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Indemo faile nerastas lapas '{MONEY_SHEET_NAME}'."
        )

    overview = _read_overview(overview_ws)
    money_events = _read_money_events(workbook[MONEY_SHEET_NAME])

    project_details = []

    for ws in workbook.worksheets:
        if ws.title in IGNORED_SHEETS or not _is_project_sheet(ws):
            continue

        project_details.append(_read_project_sheet(ws))

    project_details.sort(
        key=lambda item: (
            item.get("investmentDate") or "9999-12-31",
            item.get("id") or "",
        )
    )

    overview_by_id = {
        item["id"]: item
        for item in overview["projects"]
    }

    for project in project_details:
        overview_project = overview_by_id.get(project["id"], {})

        for key in (
            "investedDate",
            "returnedDate",
            "durationDays",
            "returnedPrincipal",
            "interest",
            "xirr",
        ):
            if not project.get(key) and overview_project.get(key):
                project[key] = overview_project[key]

    active_projects = [
        project
        for project in project_details
        if project["status"] == "active"
    ]
    completed_projects = [
        project
        for project in project_details
        if project["status"] == "completed"
    ]

    summary = {
        **overview["summary"],
        "activeValue": round(
            sum(item["remaining"] for item in active_projects),
            2,
        ),
        "totalProjects": len(project_details),
        "activeProjects": len(active_projects),
        "completedProjects": len(completed_projects),
        "totalTransactions": sum(
            item["transactionCount"]
            for item in project_details
        ),
        "averagePtv": round(
            sum(item["ptv"] * item["invested"] for item in project_details)
            / sum(item["invested"] for item in project_details)
            if sum(item["invested"] for item in project_details)
            else 0,
            2,
        ),
        "averagePdt": round(
            sum(item["pdt"] * item["invested"] for item in project_details)
            / sum(item["invested"] for item in project_details)
            if sum(item["invested"] for item in project_details)
            else 0,
            2,
        ),
    }

    history = _build_history(
        project_details,
        money_events,
        summary,
    )

    result = build_standard_result(
        platform_name="Indemo",
        source_file=source_path,
        history=history,
        active_positions=active_projects,
        sold_positions=completed_projects,
        summary=summary,
        module_type="npl",
        cashflow={
            "events": money_events,
            "deposited": summary["deposited"],
            "withdrawn": round(
                sum(item["withdrawn"] for item in money_events),
                2,
            ),
            "bonuses": summary["bonuses"],
            "cash": summary["cash"],
            "returnedPrincipal": summary["returnedPrincipal"],
            "interest": summary["interest"],
        },
        updated_at=history[-1]["date"] if history else "",
    )

    result.update(
        {
            "projects": project_details,
            "activeProjects": active_projects,
            "completedProjects": completed_projects,
            "projectCounts": {
                "active": len(active_projects),
                "completed": len(completed_projects),
                "total": len(project_details),
            },
            "transactions": [
                {
                    **transaction,
                    "projectId": project["id"],
                    "projectName": project["name"],
                }
                for project in project_details
                for transaction in project["transactions"]
            ],
            "moneyEvents": money_events,
            "distributions": _build_distributions(project_details),
            "nplSummary": summary,
            "modules": {
                **result.get("modules", {}),
                "projects": True,
                "transactions": True,
                "moneyEvents": bool(money_events),
                "distributions": True,
            },
        }
    )

    return result
