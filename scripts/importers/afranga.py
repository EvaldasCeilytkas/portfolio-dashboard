"""Afranga P2P Excel failo importeris."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base_importer import build_standard_result, format_date, round_number, safe_number


OVERVIEW_SHEET = "Overview"
MONEY_SHEET = "Pinigai"


def _is_date(value: Any) -> bool:
    return isinstance(value, (datetime, date))


def _date(value: Any) -> str:
    return format_date(value) if value else ""


def _percent(value: Any) -> float:
    number = safe_number(value)
    return round(number * 100 if abs(number) <= 1 else number, 2)


def _read_overview_summary(ws) -> dict[str, float]:
    """Overview 3 eilutėje saugoma visos platformos suvestinė."""
    deposited = safe_number(ws["Q3"].value)
    invested_in_loans = safe_number(ws["R3"].value)
    bonuses = safe_number(ws["S3"].value)
    cash = safe_number(ws["T3"].value)
    principal_repaid = safe_number(ws["U3"].value)
    interest = safe_number(ws["V3"].value)
    value = safe_number(ws["W3"].value)
    profit = safe_number(ws["X3"].value)
    return_rate = _percent(ws["Y3"].value)
    xirr = _percent(ws["Z3"].value)
    fees = safe_number(ws["AA3"].value)

    # Investuota platformos lygiu yra realiai įnešti pinigai. Tai sutampa
    # su pagrindinio Investavimas.xlsx logika ir nepadvigubina reinvesticijų.
    invested = deposited

    if value == 0:
        value = cash + invested_in_loans - principal_repaid

    if profit == 0 and value:
        profit = value - invested

    if return_rate == 0 and invested:
        return_rate = profit / invested * 100

    return {
        "invested": round(invested, 2),
        "value": round(value, 2),
        "profit": round(profit, 2),
        "returnRate": round(return_rate, 2),
        "xirr": round(xirr, 2),
        "deposited": round(deposited, 2),
        "withdrawn": 0.0,
        "bonuses": round(bonuses, 2),
        "cash": round(cash, 2),
        "loanInvestments": round(invested_in_loans, 2),
        "principalRepaid": round(principal_repaid, 2),
        "interest": round(interest, 2),
        "fees": round(fees, 2),
    }


def _read_money_events(ws) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        event_date = ws.cell(row, 1).value
        if not _is_date(event_date):
            continue

        values = {
            "deposit": safe_number(ws.cell(row, 2).value),
            "withdrawal": safe_number(ws.cell(row, 3).value),
            "bonus": safe_number(ws.cell(row, 4).value),
        }

        for event_type, amount in values.items():
            if amount:
                events.append(
                    {
                        "date": _date(event_date),
                        "type": event_type,
                        "amount": round(amount, 2),
                        "description": {
                            "deposit": "Įnašas į Afranga",
                            "withdrawal": "Lėšų išėmimas iš Afranga",
                            "bonus": "Afranga premija",
                        }[event_type],
                    }
                )

    return sorted(events, key=lambda item: item["date"])


def _read_loan_sheet(ws) -> dict[str, Any]:
    loan_id = str(ws["B1"].value or ws.title).strip()
    country = str(ws["B2"].value or "").strip()
    originator = str(ws["B3"].value or "").strip()
    loan_type = str(ws["B4"].value or "").strip()
    term = str(ws["B8"].value or "").strip()
    invested_date = ws["B9"].value
    start_date = ws["B10"].value
    planned_end = ws["B11"].value
    interest_rate = _percent(ws["B15"].value)
    invested = safe_number(ws["B16"].value)

    payments: list[dict[str, Any]] = []
    principal_repaid = 0.0
    interest_received = 0.0
    fees = 0.0
    last_actual_date = None

    for row in range(3, 53):
        planned_date = ws.cell(row, 4).value
        actual_date = ws.cell(row, 5).value
        principal = safe_number(ws.cell(row, 7).value)
        interest = safe_number(ws.cell(row, 8).value)
        fee = safe_number(ws.cell(row, 9).value)

        if not (_is_date(planned_date) or _is_date(actual_date) or principal or interest or fee):
            continue

        if _is_date(actual_date):
            last_actual_date = actual_date

        principal_repaid += principal
        interest_received += interest
        fees += fee

        payments.append(
            {
                "plannedDate": _date(planned_date),
                "actualDate": _date(actual_date),
                "delayDays": round_number(ws.cell(row, 6).value, 0),
                "principal": round(principal, 2),
                "interest": round(interest, 2),
                "fee": round(fee, 2),
                "netIncome": round(interest - fee, 2),
                "status": "paid" if _is_date(actual_date) else "scheduled",
            }
        )

    remaining = max(0.0, invested - principal_repaid)
    today = date.today()
    end_date_value = planned_end.date() if isinstance(planned_end, datetime) else planned_end

    if remaining <= 0.005:
        status = "completed"
    elif isinstance(end_date_value, date) and end_date_value < today:
        status = "late"
    else:
        status = "active"

    days_remaining = (
        (end_date_value - today).days
        if isinstance(end_date_value, date) and status != "completed"
        else 0
    )

    return {
        "id": ws.title,
        "loanId": loan_id,
        "country": country,
        "originator": originator,
        "type": loan_type,
        "term": term,
        "investedDate": _date(invested_date),
        "startDate": _date(start_date),
        "plannedEndDate": _date(planned_end),
        "actualEndDate": _date(last_actual_date) if status == "completed" else "",
        "interestRate": interest_rate,
        "invested": round(invested, 2),
        "principalRepaid": round(principal_repaid, 2),
        "remainingPrincipal": round(remaining, 2),
        "interestReceived": round(interest_received, 2),
        "fees": round(fees, 2),
        "netIncome": round(interest_received - fees, 2),
        "daysRemaining": days_remaining,
        "status": status,
        "payments": payments,
    }


def _build_monthly_history(
    loans: list[dict[str, Any]],
    money_events: list[dict[str, Any]],
    summary: dict[str, Any],
) -> list[dict[str, Any]]:
    months: dict[str, dict[str, float]] = defaultdict(
        lambda: {
            "deposits": 0.0,
            "withdrawals": 0.0,
            "bonuses": 0.0,
            "loanInvestments": 0.0,
            "principalRepaid": 0.0,
            "interest": 0.0,
            "fees": 0.0,
        }
    )

    for event in money_events:
        month = event["date"][:7]
        if event["type"] == "deposit":
            months[month]["deposits"] += event["amount"]
        elif event["type"] == "withdrawal":
            months[month]["withdrawals"] += event["amount"]
        elif event["type"] == "bonus":
            months[month]["bonuses"] += event["amount"]

    for loan in loans:
        if loan["investedDate"]:
            months[loan["investedDate"][:7]]["loanInvestments"] += loan["invested"]

        for payment in loan["payments"]:
            if not payment["actualDate"]:
                continue
            month = payment["actualDate"][:7]
            months[month]["principalRepaid"] += payment["principal"]
            months[month]["interest"] += payment["interest"]
            months[month]["fees"] += payment["fee"]

    if not months:
        return []

    first_month = min(months)
    last_month = max(months)
    year, month = map(int, first_month.split("-"))
    last_year, last_month_number = map(int, last_month.split("-"))

    running_deposited = 0.0
    running_withdrawn = 0.0
    running_bonuses = 0.0
    running_interest = 0.0
    running_fees = 0.0
    running_principal_repaid = 0.0
    running_loan_investments = 0.0
    history: list[dict[str, Any]] = []

    while (year, month) <= (last_year, last_month_number):
        key = f"{year:04d}-{month:02d}"
        data = months[key]
        running_deposited += data["deposits"]
        running_withdrawn += data["withdrawals"]
        running_bonuses += data["bonuses"]
        running_interest += data["interest"]
        running_fees += data["fees"]
        running_principal_repaid += data["principalRepaid"]
        running_loan_investments += data["loanInvestments"]

        invested = running_deposited - running_withdrawn
        cash = (
            invested
            + running_bonuses
            + running_principal_repaid
            + running_interest
            - running_fees
            - running_loan_investments
        )
        outstanding = max(0.0, running_loan_investments - running_principal_repaid)
        value = cash + outstanding
        profit = value - invested

        history.append(
            {
                "date": f"{key}-01",
                "invested": round(invested, 2),
                "value": round(value, 2),
                "profit": round(profit, 2),
                "returnRate": round(profit / invested * 100, 2) if invested else 0.0,
                "deposits": round(data["deposits"], 2),
                "withdrawals": round(data["withdrawals"], 2),
                "bonuses": round(data["bonuses"], 2),
                "principalRepaid": round(data["principalRepaid"], 2),
                "interest": round(data["interest"], 2),
                "fees": round(data["fees"], 2),
                "netIncome": round(data["interest"] + data["bonuses"] - data["fees"], 2),
                "cashFlow": round(data["deposits"] - data["withdrawals"], 2),
                "monthlyProfit": round(data["interest"] + data["bonuses"] - data["fees"], 2),
            }
        )

        month += 1
        if month == 13:
            month = 1
            year += 1

    # Excel suvestinė yra galutinis autoritetas. Suvienodiname paskutinį tašką,
    # kad dėl apvalinimų ar dar neįvestų mokėjimų nebūtų skirtumo.
    history[-1].update(
        {
            "invested": summary["invested"],
            "value": summary["value"],
            "profit": summary["profit"],
            "returnRate": summary["returnRate"],
        }
    )
    return history



def _build_income_forecast(loans: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sujungia būsimus paskolų mokėjimus pagal mėnesius."""
    months: dict[str, dict[str, float]] = defaultdict(
        lambda: {"principal": 0.0, "interest": 0.0, "fees": 0.0}
    )

    for loan in loans:
        for payment in loan.get("payments", []):
            planned_date = payment.get("plannedDate", "")
            if payment.get("status") != "scheduled" or not planned_date:
                continue

            month = planned_date[:7]
            months[month]["principal"] += safe_number(payment.get("principal"))
            months[month]["interest"] += safe_number(payment.get("interest"))
            months[month]["fees"] += safe_number(payment.get("fee"))

    # Kai Excel faile nėra būsimų grafiko eilučių, pateikiama atsargi
    # preliminari prognozė pagal likusį pagrindą, palūkanų normą ir terminą.
    # Tai nėra platformos garantuotas mokėjimo grafikas.
    if not months:
        today = date.today()
        for loan in loans:
            remaining = safe_number(loan.get("remainingPrincipal"))
            planned_end = loan.get("plannedEndDate", "")
            if remaining <= 0 or not planned_end:
                continue

            try:
                end_date = datetime.strptime(planned_end, "%Y-%m-%d").date()
            except ValueError:
                continue

            if end_date < today:
                continue

            days = max(1, (end_date - today).days)
            estimated_interest = remaining * safe_number(loan.get("interestRate")) / 100 * days / 365
            month = planned_end[:7]
            months[month]["principal"] += remaining
            months[month]["interest"] += estimated_interest

    result = []
    for month, values in sorted(months.items()):
        total = values["principal"] + values["interest"] - values["fees"]
        result.append(
            {
                "month": month,
                "principal": round(values["principal"], 2),
                "interest": round(values["interest"], 2),
                "fees": round(values["fees"], 2),
                "total": round(total, 2),
                "estimated": True,
            }
        )

    return result


def _build_distributions(loans: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    def group(field: str) -> list[dict[str, Any]]:
        result: dict[str, float] = defaultdict(float)
        for loan in loans:
            result[str(loan.get(field) or "Nežinoma")] += safe_number(loan.get("remainingPrincipal"))
        total = sum(result.values())
        return [
            {
                "name": name,
                "value": round(value, 2),
                "share": round(value / total * 100, 2) if total else 0.0,
            }
            for name, value in sorted(result.items(), key=lambda item: item[1], reverse=True)
        ]

    return {
        "byOriginator": group("originator"),
        "byCountry": group("country"),
        "byType": group("type"),
        "byStatus": group("status"),
    }


def _build_health(loans: list[dict[str, Any]], summary: dict[str, Any]) -> dict[str, Any]:
    active = [loan for loan in loans if loan["status"] in {"active", "late"}]
    outstanding = sum(loan["remainingPrincipal"] for loan in active)
    largest_share = max(
        (loan["remainingPrincipal"] / outstanding * 100 for loan in active),
        default=0.0,
    )
    late_share = (
        sum(loan["remainingPrincipal"] for loan in active if loan["status"] == "late")
        / outstanding
        * 100
        if outstanding
        else 0.0
    )
    cash_share = summary["cash"] / summary["value"] * 100 if summary["value"] else 0.0
    originators = len({loan["originator"] for loan in active if loan["originator"]})

    diversification = max(0, min(20, round(8 + len(active) * 0.7 + originators * 1.5 - max(0, largest_share - 15) * 0.35)))
    loan_quality = max(0, min(20, round(20 - late_share * 0.8)))
    cash_efficiency = max(0, min(20, round(20 - max(0, cash_share - 3) * 1.2)))
    repayment = max(0, min(20, round(20 - late_share * 0.65)))
    concentration = max(0, min(20, round(20 - max(0, largest_share - 10) * 0.55)))
    score = diversification + loan_quality + cash_efficiency + repayment + concentration

    return {
        "score": score,
        "label": "Puiki būklė" if score >= 85 else "Gera būklė" if score >= 70 else "Reikia stebėti",
        "metrics": [
            {"key": "diversification", "label": "Diversifikacija", "score": diversification, "max": 20, "description": f"{len(active)} aktyvių paskolų ir {originators} skolintojai."},
            {"key": "loanQuality", "label": "Paskolų kokybė", "score": loan_quality, "max": 20, "description": f"Vėluojanti portfelio dalis sudaro {late_share:.2f} %."},
            {"key": "cashEfficiency", "label": "Laisvų pinigų efektyvumas", "score": cash_efficiency, "max": 20, "description": f"Laisvi pinigai sudaro {cash_share:.2f} % portfelio."},
            {"key": "repayment", "label": "Grąžinimų stabilumas", "score": repayment, "max": 20, "description": "Vertinama pagal aktyvių vėlavimų dalį."},
            {"key": "concentration", "label": "Koncentracijos rizika", "score": concentration, "max": 20, "description": f"Didžiausia paskola sudaro {largest_share:.2f} % aktyvaus portfelio."},
        ],
        "largestLoanShare": round(largest_share, 2),
        "lateShare": round(late_share, 2),
        "cashShare": round(cash_share, 2),
        "summary": (
            "Portfelis stabilus, mokėjimai nevėluoja, o laisvų pinigų dalis maža."
            if score >= 85 and late_share == 0
            else "Portfelio būklė gera, tačiau verta stebėti koncentraciją ir vėlavimus."
            if score >= 70
            else "Keli rizikos rodikliai reikalauja papildomo dėmesio."
        ),
    }


def load_afranga(source_file: str | Path) -> dict[str, Any]:
    source_path = Path(source_file)
    if not source_path.exists():
        raise FileNotFoundError(f"Afranga failas nerastas: {source_path}")

    workbook = load_workbook(source_path, data_only=True)
    if OVERVIEW_SHEET not in workbook.sheetnames or MONEY_SHEET not in workbook.sheetnames:
        raise ValueError("Afranga faile nerasti privalomi Overview ir Pinigai lapai.")

    overview = workbook[OVERVIEW_SHEET]
    summary = _read_overview_summary(overview)
    money_events = _read_money_events(workbook[MONEY_SHEET])

    loans = [
        _read_loan_sheet(workbook[sheet_name])
        for sheet_name in workbook.sheetnames
        if sheet_name not in {OVERVIEW_SHEET, MONEY_SHEET}
    ]
    loans.sort(key=lambda loan: (loan["status"] == "completed", loan["plannedEndDate"], loan["id"]))

    history = _build_monthly_history(loans, money_events, summary)
    active_loans = [loan for loan in loans if loan["status"] == "active"]
    late_loans = [loan for loan in loans if loan["status"] == "late"]
    completed_loans = [loan for loan in loans if loan["status"] == "completed"]
    weighted_base = sum(loan["remainingPrincipal"] for loan in loans)
    average_interest = (
        sum(loan["interestRate"] * loan["remainingPrincipal"] for loan in loans) / weighted_base
        if weighted_base
        else 0.0
    )

    result = build_standard_result(
        platform_name="Afranga",
        source_file=source_path,
        history=history,
        active_positions=loans,
        sold_positions=[],
        summary=summary,
        module_type="p2p",
        cashflow=summary,
        updated_at=history[-1]["date"] if history else "",
    )

    result.update(
        {
            "type": "p2p",
            "modules": {
                **result.get("modules", {}),
                "positions": False,
                "loans": True,
                "income": True,
                "cashflows": True,
                "health": True,
            },
            "loans": loans,
            "moneyEvents": money_events,
            "distributions": _build_distributions(loans),
            "incomeForecast": _build_income_forecast(loans),
            "health": _build_health(loans, summary),
            "p2pSummary": {
                "totalLoans": len(loans),
                "activeLoans": len(active_loans),
                "lateLoans": len(late_loans),
                "completedLoans": len(completed_loans),
                "averageInterestRate": round(average_interest, 2),
                "outstandingPrincipal": round(weighted_base, 2),
                "cash": summary["cash"],
                "interestReceived": summary["interest"],
                "fees": summary["fees"],
                "bonuses": summary["bonuses"],
            },
        }
    )
    return result
