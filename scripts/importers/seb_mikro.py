"""SEB Mikro Excel failo ETF pozicijų importeris."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

from .base_importer import (
    build_standard_result,
    create_slug,
    format_period,
    safe_number,
)


OVERVIEW_SHEET = "Apžvalga"
OVERVIEW_TICKER_ROW = 2
OVERVIEW_FIRST_TICKER_COLUMN = 5
HISTORY_FIRST_ROW = 5
SUMMARY_ROW = 5
SOLD_GREEN_RGB = "FF92D050"






def get_rgb_color(cell) -> str:
    """Grąžina langelio užpildo RGB spalvą, kai ji pasiekiama."""
    color = cell.fill.fgColor

    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()

    if color.type == "indexed" and color.indexed is not None:
        index = int(color.indexed)
        if 0 <= index < len(COLOR_INDEX):
            return str(COLOR_INDEX[index]).upper()

    return ""


def is_sold_header(cell) -> bool:
    """Žalias ETF pavadinimo langelis Apžvalgoje reiškia parduotą ETF."""
    return cell.fill.fill_type == "solid" and get_rgb_color(cell) == SOLD_GREEN_RGB


def parse_holding_title(value: Any, fallback_ticker: str) -> tuple[str, str]:
    """Iš A1 išskiria tickerį ir pilną ETF pavadinimą."""
    text = " ".join(str(value or "").strip().split())

    if not text:
        return fallback_ticker, fallback_ticker

    parts = text.split(" ", 2)
    ticker = parts[0].strip() or fallback_ticker

    if len(parts) >= 3:
        name = parts[2].strip()
    elif len(parts) == 2:
        name = parts[1].strip()
    else:
        name = ticker

    return ticker, name


def build_holding_history(ws) -> list[dict[str, Any]]:
    """Nuskaito ETF mėnesinę istoriją iš A:M stulpelių."""
    history: list[dict[str, Any]] = []

    for row_number in range(HISTORY_FIRST_ROW, ws.max_row + 1):
        period = ws.cell(row=row_number, column=1).value

        if period in (None, ""):
            continue

        invested = safe_number(ws.cell(row=row_number, column=4).value)
        quantity = safe_number(ws.cell(row=row_number, column=6).value)
        price = safe_number(ws.cell(row=row_number, column=7).value)
        value = safe_number(ws.cell(row=row_number, column=8).value)
        profit = safe_number(ws.cell(row=row_number, column=9).value)
        return_rate = safe_number(ws.cell(row=row_number, column=10).value)

        history.append(
            {
                "date": format_period(period),
                "contribution": round(
                    safe_number(ws.cell(row=row_number, column=2).value), 2
                ),
                "monthlyInvested": round(
                    safe_number(ws.cell(row=row_number, column=3).value), 2
                ),
                "invested": round(invested, 2),
                "monthlyQuantity": round(
                    safe_number(ws.cell(row=row_number, column=5).value), 8
                ),
                "quantity": round(quantity, 8),
                "price": round(price, 6),
                "value": round(value, 2),
                "profit": round(profit, 2),
                "returnRate": round(return_rate * 100, 2),
                "buyFee": round(
                    safe_number(ws.cell(row=row_number, column=11).value), 2
                ),
                "sellFee": round(
                    safe_number(ws.cell(row=row_number, column=12).value), 2
                ),
                "dividends": round(
                    safe_number(ws.cell(row=row_number, column=13).value), 2
                ),
            }
        )

    return history



def to_full_date(period: Any) -> str:
    """Paverčia Excel laikotarpį į YYYY-MM-01 datą."""
    formatted = format_period(period)

    if not formatted:
        return ""

    return f"{formatted}-01"


def build_position_transactions(
    history: list[dict[str, Any]],
    *,
    sold: bool,
    sold_value: float,
    historical_quantity: float,
    sell_fees: float,
) -> list[dict[str, Any]]:
    """
    Sukuria pozicijos sandorių istoriją iš mėnesinių Excel duomenų.

    SEB Mikro faile saugomas mėnesio tikslumas, todėl sandorio data
    pateikiama kaip atitinkamo mėnesio pirmoji diena.
    """
    transactions: list[dict[str, Any]] = []

    for point in history:
        quantity = safe_number(point.get("monthlyQuantity"))
        invested = safe_number(point.get("monthlyInvested"))
        contribution = safe_number(point.get("contribution"))
        buy_fee = safe_number(point.get("buyFee"))

        if quantity <= 0:
            continue

        amount = invested if invested > 0 else max(contribution - buy_fee, 0)
        price = amount / quantity if quantity else 0

        transactions.append(
            {
                "id": f"buy-{point.get('date', '')}-{len(transactions) + 1}",
                "date": to_full_date(point.get("date")),
                "type": "buy",
                "side": "buy",
                "label": "Pirkimas",
                "quantity": round(quantity, 8),
                "price": round(price, 6),
                "amount": round(amount, 2),
                "fee": round(buy_fee, 2),
                "currency": "EUR",
                "estimated": True,
            }
        )

    if sold and sold_value > 0 and historical_quantity > 0:
        last_date = to_full_date(history[-1].get("date")) if history else ""
        sale_price = sold_value / historical_quantity

        transactions.append(
            {
                "id": f"sell-{last_date or 'closed'}",
                "date": last_date,
                "type": "sell",
                "side": "sell",
                "label": "Pardavimas",
                "quantity": round(historical_quantity, 8),
                "price": round(sale_price, 6),
                "amount": round(sold_value, 2),
                "fee": round(sell_fees, 2),
                "currency": "EUR",
                "estimated": True,
            }
        )

    return transactions


def build_position_dividends(
    history: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Sukuria atskirų dividendų mokėjimų sąrašą."""
    dividends: list[dict[str, Any]] = []

    for point in history:
        amount = safe_number(point.get("dividends"))

        if amount <= 0:
            continue

        dividends.append(
            {
                "id": f"dividend-{point.get('date', '')}-{len(dividends) + 1}",
                "date": to_full_date(point.get("date")),
                "amount": round(amount, 2),
                "currency": "EUR",
                "type": "dividend",
                "label": "Dividendai",
            }
        )

    return dividends


def build_position_events(
    history: list[dict[str, Any]],
    transactions: list[dict[str, Any]],
    dividends: list[dict[str, Any]],
    *,
    sold: bool,
) -> list[dict[str, Any]]:
    """Sukuria Investment Profile laiko juostos įvykius."""
    events: list[dict[str, Any]] = []
    buy_transactions = [
        item for item in transactions if item.get("type") == "buy"
    ]

    for index, transaction in enumerate(buy_transactions):
        events.append(
            {
                "id": f"event-{transaction['id']}",
                "date": transaction.get("date", ""),
                "type": "first_buy" if index == 0 else "buy",
                "title": (
                    "Pirmasis pozicijos pirkimas"
                    if index == 0
                    else "Papildytas pozicijos kiekis"
                ),
                "description": (
                    f"Nupirkta {transaction.get('quantity', 0):.4f} vnt. "
                    f"už {transaction.get('amount', 0):.2f} €."
                ),
                "amount": transaction.get("amount", 0),
            }
        )

    for dividend in dividends:
        events.append(
            {
                "id": f"event-{dividend['id']}",
                "date": dividend.get("date", ""),
                "type": "dividend",
                "title": "Gauti dividendai",
                "description": (
                    f"Gauta {safe_number(dividend.get('amount')):.2f} € dividendų."
                ),
                "amount": dividend.get("amount", 0),
            }
        )

    ath_point: dict[str, Any] | None = None

    for point in history:
        if ath_point is None or safe_number(point.get("value")) > safe_number(
            ath_point.get("value")
        ):
            ath_point = point

    if ath_point and safe_number(ath_point.get("value")) > 0:
        events.append(
            {
                "id": f"event-ath-{ath_point.get('date', '')}",
                "date": to_full_date(ath_point.get("date")),
                "type": "ath",
                "title": "Pasiektas pozicijos vertės rekordas",
                "description": (
                    f"Pozicijos vertė pasiekė "
                    f"{safe_number(ath_point.get('value')):.2f} €."
                ),
                "amount": round(safe_number(ath_point.get("value")), 2),
            }
        )

    if sold:
        sell_transaction = next(
            (
                item
                for item in reversed(transactions)
                if item.get("type") == "sell"
            ),
            None,
        )

        if sell_transaction:
            events.append(
                {
                    "id": f"event-{sell_transaction['id']}",
                    "date": sell_transaction.get("date", ""),
                    "type": "sell",
                    "title": "Pozicija uždaryta",
                    "description": (
                        f"Pozicija realizuota už "
                        f"{safe_number(sell_transaction.get('amount')):.2f} €."
                    ),
                    "amount": sell_transaction.get("amount", 0),
                }
            )

    events.sort(key=lambda item: (item.get("date", ""), item.get("id", "")))
    return events


def build_performance_summary(
    history: list[dict[str, Any]],
) -> dict[str, Any]:
    """Sukuria glaustą grafiko statistikos santrauką."""
    if not history:
        return {
            "startValue": 0.0,
            "currentValue": 0.0,
            "change": 0.0,
            "changePercent": 0.0,
            "ath": 0.0,
            "athDate": "",
            "lowestValue": 0.0,
            "lowestValueDate": "",
        }

    first = history[0]
    latest = history[-1]
    ath_point = max(history, key=lambda item: safe_number(item.get("value")))
    low_point = min(history, key=lambda item: safe_number(item.get("value")))

    start_value = safe_number(first.get("value"))
    current_value = safe_number(latest.get("value"))
    change = current_value - start_value
    change_percent = change / start_value * 100 if start_value else 0.0

    return {
        "startValue": round(start_value, 2),
        "currentValue": round(current_value, 2),
        "change": round(change, 2),
        "changePercent": round(change_percent, 2),
        "ath": round(safe_number(ath_point.get("value")), 2),
        "athDate": to_full_date(ath_point.get("date")),
        "lowestValue": round(safe_number(low_point.get("value")), 2),
        "lowestValueDate": to_full_date(low_point.get("date")),
    }


def build_holding(ws, ticker: str, sold: bool) -> dict[str, Any]:
    """Sukuria vieno ETF pilną V10.4 Investment Profile struktūrą."""
    parsed_ticker, name = parse_holding_title(ws["A1"].value, ticker)
    history = build_holding_history(ws)
    latest = history[-1] if history else {}

    cash_contributed = safe_number(ws.cell(SUMMARY_ROW, 16).value)
    summary_quantity = safe_number(ws.cell(SUMMARY_ROW, 17).value)
    sold_value = safe_number(ws.cell(SUMMARY_ROW, 18).value)
    summary_profit = safe_number(ws.cell(SUMMARY_ROW, 19).value)
    summary_return_rate = safe_number(ws.cell(SUMMARY_ROW, 20).value)
    xirr = safe_number(ws.cell(SUMMARY_ROW, 21).value)
    buy_fees = safe_number(ws.cell(SUMMARY_ROW, 22).value)
    sell_fees = safe_number(ws.cell(SUMMARY_ROW, 23).value)
    dividend_income = safe_number(ws.cell(SUMMARY_ROW, 24).value)
    average_price = safe_number(ws.cell(SUMMARY_ROW, 25).value)

    if sold:
        invested = max(cash_contributed - buy_fees, 0)
        current_value = 0.0
        quantity = 0.0
        profit = summary_profit
        return_rate = summary_return_rate * 100
        current_price = safe_number(latest.get("price"))
    else:
        invested = safe_number(latest.get("invested"))
        current_value = safe_number(latest.get("value"))
        quantity = safe_number(latest.get("quantity"))
        profit = current_value - invested + dividend_income
        return_rate = profit / invested * 100 if invested else 0.0
        current_price = safe_number(latest.get("price"))

    transactions = build_position_transactions(
        history,
        sold=sold,
        sold_value=sold_value,
        historical_quantity=summary_quantity,
        sell_fees=sell_fees,
    )
    dividend_history = build_position_dividends(history)
    events = build_position_events(
        history,
        transactions,
        dividend_history,
        sold=sold,
    )
    performance_summary = build_performance_summary(history)

    first_buy = next(
        (
            item
            for item in transactions
            if item.get("type") == "buy"
        ),
        None,
    )
    holding_since = first_buy.get("date", "") if first_buy else ""

    fees = {
        "buy": round(buy_fees, 2),
        "sell": round(sell_fees, 2),
        "total": round(buy_fees + sell_fees, 2),
    }

    summary = {
        "invested": round(invested, 2),
        "cashContributed": round(cash_contributed, 2),
        "value": round(current_value, 2),
        "currentValue": round(current_value, 2),
        "soldValue": round(sold_value, 2),
        "profit": round(profit, 2),
        "returnRate": round(return_rate, 2),
        "xirr": round(xirr * 100, 2),
        "quantity": round(quantity, 8),
        "historicalQuantity": round(summary_quantity, 8),
        "averageBuyPrice": round(average_price, 6),
        "currentPrice": round(current_price, 6),
        "dividendIncome": round(dividend_income, 2),
        "fees": fees,
        "holdingSince": holding_since,
    }

    return {
        "schemaVersion": 8,
        "id": create_slug(parsed_ticker),
        "symbol": parsed_ticker,
        "ticker": parsed_ticker,
        "name": name,
        "type": "ETF",
        "assetType": "ETF",
        "currency": "EUR",
        "active": not sold,
        "status": "sold" if sold else "active",
        "holdingSince": holding_since,
        "firstPurchaseDate": holding_since,
        "invested": summary["invested"],
        "cashContributed": summary["cashContributed"],
        "quantity": summary["quantity"],
        "historicalQuantity": summary["historicalQuantity"],
        "averagePrice": summary["averageBuyPrice"],
        "averageBuyPrice": summary["averageBuyPrice"],
        "lastPrice": summary["currentPrice"],
        "price": summary["currentPrice"],
        "currentPrice": summary["currentPrice"],
        "value": summary["value"],
        "currentValue": summary["currentValue"],
        "soldValue": summary["soldValue"],
        "sold": summary["soldValue"],
        "profit": summary["profit"],
        "returnRate": summary["returnRate"],
        "xirr": summary["xirr"],
        "buyFees": fees["buy"],
        "sellFees": fees["sell"],
        "fees": fees,
        "dividendIncome": summary["dividendIncome"],
        "dividendsTotal": summary["dividendIncome"],
        "summary": summary,
        "performanceSummary": performance_summary,
        "history": history,
        "transactions": transactions,
        "dividends": dividend_history,
        "events": events,
    }



def build_platform_history(
    holdings: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Sudeda visų ETF istoriją į vieną SEB Mikro platformos istoriją.

    Įtraukiamos ir vėliau parduotos pozicijos, todėl istorinis grafikas
    išlieka pilnas iki realizavimo mėnesio.
    """
    monthly: dict[str, dict[str, float | str]] = {}

    for holding in holdings:
        for point in holding.get("history", []):
            date_value = point.get("date")

            if not date_value:
                continue

            item = monthly.setdefault(
                date_value,
                {
                    "date": date_value,
                    "invested": 0.0,
                    "value": 0.0,
                    "profit": 0.0,
                    "dividends": 0.0,
                    "buyFees": 0.0,
                    "sellFees": 0.0,
                },
            )

            item["invested"] += safe_number(point.get("invested"))
            item["value"] += safe_number(point.get("value"))
            item["profit"] += safe_number(point.get("profit"))
            item["dividends"] += safe_number(point.get("dividends"))
            item["buyFees"] += safe_number(point.get("buyFee"))
            item["sellFees"] += safe_number(point.get("sellFee"))

    history: list[dict[str, Any]] = []

    for item in monthly.values():
        invested = safe_number(item["invested"])
        profit = safe_number(item["profit"])

        history.append(
            {
                **item,
                "date": f"{item['date']}-01",
                "invested": round(invested, 2),
                "value": round(safe_number(item["value"]), 2),
                "profit": round(profit, 2),
                "returnRate": round(
                    profit / invested * 100 if invested else 0,
                    2,
                ),
                "dividends": round(
                    safe_number(item["dividends"]),
                    2,
                ),
                "buyFees": round(
                    safe_number(item["buyFees"]),
                    2,
                ),
                "sellFees": round(
                    safe_number(item["sellFees"]),
                    2,
                ),
            }
        )

    history.sort(key=lambda item: item["date"])
    return history


def load_seb_mikro_holdings(excel_file: str | Path) -> dict[str, Any]:
    """
    Nuskaito pilną SEB Mikro platformos informaciją.

    Grąžina vienodo formato platformos rezultatą:
    summary, history, cashflow ir positions.
    """
    excel_path = Path(excel_file)

    if not excel_path.exists():
        raise FileNotFoundError(f"Nerastas SEB Mikro failas: {excel_path}")

    values_workbook = load_workbook(
        excel_path,
        data_only=True,
        read_only=False,
    )
    styles_workbook = load_workbook(
        excel_path,
        data_only=False,
        read_only=False,
    )

    if OVERVIEW_SHEET not in values_workbook.sheetnames:
        raise ValueError(
            f"Nerastas lapas '{OVERVIEW_SHEET}' "
            f"faile {excel_path.name}"
        )

    overview_values = values_workbook[OVERVIEW_SHEET]
    overview_styles = styles_workbook[OVERVIEW_SHEET]

    active_holdings: list[dict[str, Any]] = []
    sold_holdings: list[dict[str, Any]] = []

    for column_number in range(
        OVERVIEW_FIRST_TICKER_COLUMN,
        overview_values.max_column + 1,
    ):
        ticker_value = overview_values.cell(
            row=OVERVIEW_TICKER_ROW,
            column=column_number,
        ).value

        if ticker_value in (None, ""):
            continue

        ticker = str(ticker_value).strip()

        if ticker not in values_workbook.sheetnames:
            continue

        sold = is_sold_header(
            overview_styles.cell(
                row=OVERVIEW_TICKER_ROW,
                column=column_number,
            )
        )

        holding = build_holding(
            values_workbook[ticker],
            ticker,
            sold,
        )

        if sold:
            sold_holdings.append(holding)
        else:
            active_holdings.append(holding)

    active_holdings.sort(
        key=lambda item: item["value"],
        reverse=True,
    )
    sold_holdings.sort(
        key=lambda item: item["soldValue"],
        reverse=True,
    )

    all_holdings = active_holdings + sold_holdings
    history = build_platform_history(all_holdings)
    latest = history[-1] if history else {}

    invested = sum(
        safe_number(item.get("invested"))
        for item in active_holdings
    )
    value = sum(
        safe_number(item.get("value"))
        for item in active_holdings
    )
    profit = value - invested
    return_rate = profit / invested * 100 if invested else 0.0

    dividends = sum(
        safe_number(item.get("dividends"))
        for item in all_holdings
    )
    fees = sum(
        safe_number(item.get("fees", {}).get("total"))
        for item in all_holdings
    )

    result = build_standard_result(
        platform_name="SEB Mikro",
        source_file=excel_path,
        history=history,
        active_positions=active_holdings,
        sold_positions=sold_holdings,
        module_type="etf",
        updated_at=latest.get("date", ""),
        summary={
            "invested": invested,
            "value": value,
            "profit": profit,
            "returnRate": return_rate,
            "dividends": dividends,
            "fees": fees,
        },
        cashflow={
            "dividends": round(dividends, 2),
            "fees": round(fees, 2),
        },
    )

    result["schemaVersion"] = 8
    return result
