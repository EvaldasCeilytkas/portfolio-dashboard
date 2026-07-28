#!/usr/bin/env python3
"""
NT platformų vėlavimų sinchronizatorius V4.

Naudojimas:
    python scripts/update_real_estate_delays_v4.py \
        --excel excel/Nordstreet.xlsx \
        --json public/data/nordstreet.json \
        --as-of 2026-07-27

Vėlavimo taisyklė:
- vertinami tik aktyvūs projektai;
- mokėjimas laikomas šiuo metu vėluojančiu, kai:
  1) planinė data yra ne vėlesnė už --as-of datą;
  2) planinė paskolos arba palūkanų suma yra didesnė už 0;
  3) faktinė mokėjimo data Excel lape neįrašyta;
- istorinis, jau apmokėtas vėlavimas projekto nepažymi kaip šiuo metu vėluojančio.
"""

from __future__ import annotations

import argparse
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


OVERVIEW_SHEETS = {"apžvalga", "overview"}
CASH_SHEETS = {"pinigai", "cash"}


def number(value: Any) -> float:
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def parse_as_of(value: str | None) -> date:
    if value:
        return date.fromisoformat(value)
    return date.today()


def is_project_sheet(sheet_name: str) -> bool:
    normalized = sheet_name.strip().casefold()
    return normalized not in OVERVIEW_SHEETS | CASH_SHEETS


def read_current_delay(ws, as_of: date) -> dict[str, Any]:
    overdue_payments = []
    historical_delay_days = []

    for row in range(2, ws.max_row + 1):
        planned_date = ws.cell(row, 4).value       # D – planinė data
        planned_principal = number(ws.cell(row, 5).value)  # E
        planned_interest = number(ws.cell(row, 6).value)   # F
        actual_date = ws.cell(row, 10).value       # J – faktinė data
        excel_delay = ws.cell(row, 11).value       # K – istorinis vėlavimas

        if isinstance(excel_delay, (int, float)) and excel_delay > 0:
            historical_delay_days.append(int(excel_delay))

        if not isinstance(planned_date, (datetime, date)):
            continue

        planned_day = planned_date.date() if isinstance(planned_date, datetime) else planned_date
        planned_amount = planned_principal + planned_interest

        # Tik reali, jau suėjusi ir vis dar neapmokėta įmoka.
        if planned_day <= as_of and planned_amount > 0 and not isinstance(actual_date, (datetime, date)):
            days = max(0, (as_of - planned_day).days)
            overdue_payments.append(
                {
                    "row": row,
                    "plannedDate": planned_day.isoformat(),
                    "plannedPrincipal": round(planned_principal, 2),
                    "plannedInterest": round(planned_interest, 2),
                    "overdueAmount": round(planned_amount, 2),
                    "delayDays": days,
                }
            )

    return {
        "hasCurrentDelay": bool(overdue_payments),
        "currentDelayDays": max((item["delayDays"] for item in overdue_payments), default=0),
        "overduePayments": overdue_payments,
        "overduePaymentsCount": len(overdue_payments),
        "overdueAmount": round(sum(item["overdueAmount"] for item in overdue_payments), 2),
        "maxHistoricalDelayDays": max(historical_delay_days, default=0),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, help="Excel failo kelias")
    parser.add_argument("--json", required=True, help="JSON failo kelias")
    parser.add_argument("--as-of", help="Vertinimo data YYYY-MM-DD; pagal nutylėjimą šiandien")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    json_path = Path(args.json)
    as_of = parse_as_of(args.as_of)

    if not excel_path.exists():
        raise FileNotFoundError(f"Nerastas Excel failas: {excel_path}")
    if not json_path.exists():
        raise FileNotFoundError(f"Nerastas JSON failas: {json_path}")

    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    project_sheets = [name for name in workbook.sheetnames if is_project_sheet(name)]

    with json_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)

    projects = data.get("projects", [])
    if len(projects) != len(project_sheets):
        raise ValueError(
            "Projektų skaičius Excel ir JSON nesutampa: "
            f"{len(project_sheets)} Excel, {len(projects)} JSON."
        )

    delayed_active = 0

    for index, (sheet_name, project) in enumerate(zip(project_sheets, projects), start=1):
        ws = workbook[sheet_name]
        invested = number(project.get("invested"))
        outstanding = number(project.get("outstanding"))
        completed = outstanding <= 0.005 and invested > 0

        delay = read_current_delay(ws, as_of)

        project["sourceSheet"] = sheet_name
        project["delay"] = delay
        project["delayDays"] = delay["currentDelayDays"]
        project["overduePaymentsCount"] = delay["overduePaymentsCount"]
        project["overdueAmount"] = delay["overdueAmount"]

        if completed:
            project["status"] = "completed"
        elif delay["hasCurrentDelay"]:
            project["status"] = "delayed"
            delayed_active += 1
        else:
            project["status"] = "active"

        # Atnaujiname mokėjimų masyvą, kad projekto puslapyje matytųsi
        # dabartinis neapmokėtų įmokų vėlavimas.
        overdue_by_date = {
            item["plannedDate"]: item["delayDays"]
            for item in delay["overduePayments"]
        }
        for payment in project.get("payments", []):
            planned = payment.get("plannedDate")
            actual = payment.get("actualDate")
            if planned in overdue_by_date and not actual:
                payment["delayDays"] = overdue_by_date[planned]
                payment["isOverdue"] = True
            else:
                payment["isOverdue"] = False

    summary = data.setdefault("summary", {})
    active_count = sum(1 for project in projects if project.get("status") != "completed")
    completed_count = sum(1 for project in projects if project.get("status") == "completed")

    summary["activeProjects"] = active_count
    summary["delayedProjects"] = delayed_active
    summary["repaidProjects"] = completed_count
    summary["completedProjects"] = completed_count
    summary["totalProjects"] = len(projects)

    platform = data.setdefault("platform", {})
    platform["updatedAt"] = as_of.isoformat()

    # Vėlavimų Health dalis tiesiogiai paremta aktyvių projektų dalimi.
    delay_health = round(100 * (1 - delayed_active / active_count)) if active_count else 100
    health_parts = platform.setdefault("healthParts", {})
    health_parts["delays"] = max(0, min(100, delay_health))

    # Perskaičiuojamas bendras Health balas iš penkių rodomų dalių.
    parts = [
        number(health_parts.get("diversification", 80)),
        number(health_parts.get("ratings", health_parts.get("rating", 80))),
        number(health_parts.get("ltv", 80)),
        number(health_parts.get("delays", 80)),
        number(health_parts.get("cash", 80)),
    ]
    platform["healthScore"] = round(sum(parts) / len(parts))

    data["delayCalculation"] = {
        "version": "4.0",
        "asOf": as_of.isoformat(),
        "rule": "Aktyvus projektas vėluoja, jei yra suėjusi planinė ir vis dar neapmokėta įmoka.",
        "historicalPaidDelaysDoNotMarkCurrentStatus": True,
    }

    with json_path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)

    print(f"Atnaujinta: {json_path}")
    print(f"Aktyvūs projektai: {active_count}")
    print(f"Šiuo metu vėluoja: {delayed_active}")
    print(f"Užbaigti projektai: {completed_count}")


if __name__ == "__main__":
    main()
