from __future__ import annotations

import argparse
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHEMA_VERSION = 1
EPSILON = 0.000001
UNIT_TOLERANCE = 0.01

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = PROJECT_ROOT / "excel" / "SEB Fondai.xlsx"
DEFAULT_OUTPUT = (
    PROJECT_ROOT / "public" / "data" / "platforms" / "seb-fondai.json"
)

OVERVIEW_SHEET = "Apžvalga"
MONTH_SHEET_RE = re.compile(r"^\d{4}\.\d{2}$")


def finite_number(value: Any, default: float = 0.0) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if math.isfinite(number) else default


def rounded(value: Any, digits: int = 2) -> float:
    return round(finite_number(value), digits)


def nullable_number(value: Any, digits: int = 4) -> float | None:
    if value is None:
        return None
    number = finite_number(value, float("nan"))
    if not math.isfinite(number):
        return None
    return round(number, digits)


def month_key(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return f"{value.year:04d}-{value.month:02d}"

    if isinstance(value, (int, float)):
        text = f"{float(value):.2f}"
    else:
        text = str(value or "").strip().replace(",", ".")

    match = re.fullmatch(r"(\d{4})\.(\d{1,2})", text)
    if not match:
        return None

    year = int(match.group(1))
    month = int(match.group(2))
    if not 1 <= month <= 12:
        return None
    return f"{year:04d}-{month:02d}"


def month_start(value: Any) -> str | None:
    key = month_key(value)
    return f"{key}-01" if key else None


def month_end(value: Any) -> str | None:
    key = month_key(value)
    if not key:
        return None

    year, month = map(int, key.split("-"))
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)

    return date.fromordinal(next_month.toordinal() - 1).isoformat()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def find_data_rows(ws, start_row: int = 3) -> list[int]:
    return [
        row
        for row in range(start_row, ws.max_row + 1)
        if month_key(ws.cell(row, 1).value)
    ]


def get_fund_sheets(workbook) -> list[str]:
    return [
        name
        for name in workbook.sheetnames
        if name != OVERVIEW_SHEET and not MONTH_SHEET_RE.fullmatch(name)
    ]


def scan_monthly_activity(
    workbook,
    fund_sheets: list[str],
) -> tuple[dict[str, dict[str, float]], dict[str, dict[str, float]]]:
    """Nuskaito pirkimus, pardavimus ir dividendus iš mėnesinių lapų."""
    full_name_to_sheet = {
        str(workbook[name]["A1"].value or "").strip(): name
        for name in fund_sheets
    }

    totals: dict[str, dict[str, float]] = {
        name: defaultdict(float) for name in fund_sheets
    }
    monthly: dict[str, dict[str, float]] = defaultdict(
        lambda: defaultdict(float)
    )

    for sheet_name in workbook.sheetnames:
        if not MONTH_SHEET_RE.fullmatch(sheet_name):
            continue

        ws = workbook[sheet_name]
        month = month_key(sheet_name)
        if not month:
            continue

        for row in range(1, ws.max_row + 1):
            title = str(ws.cell(row, 1).value or "").strip()
            fund_sheet = full_name_to_sheet.get(title)
            if not fund_sheet:
                continue

            total_row = None
            for offset in range(1, 12):
                candidate = row + offset
                if candidate > ws.max_row:
                    break
                if str(ws.cell(candidate, 1).value or "").strip() in {
                    "Viso:",
                    "Eur.",
                }:
                    total_row = candidate
                    break

            if not total_row:
                continue

            purchased = finite_number(ws.cell(total_row, 2).value)
            invested = finite_number(ws.cell(total_row, 3).value)
            purchase_fees = finite_number(ws.cell(total_row, 4).value)
            purchased_units = finite_number(ws.cell(total_row, 6).value)

            sold = finite_number(ws.cell(total_row, 9).value)
            sale_fees = finite_number(ws.cell(total_row, 10).value)
            sold_units = finite_number(ws.cell(total_row, 12).value)
            dividends = finite_number(ws.cell(total_row, 15).value)

            item = totals[fund_sheet]
            item["purchased"] += purchased
            item["invested"] += invested
            item["purchaseFees"] += purchase_fees
            item["purchasedUnits"] += purchased_units
            item["sold"] += sold
            item["saleFees"] += sale_fees
            item["soldUnits"] += sold_units
            item["dividends"] += dividends

            month_item = monthly[month]
            month_item["purchased"] += purchased
            month_item["invested"] += invested
            month_item["purchaseFees"] += purchase_fees
            month_item["sold"] += sold
            month_item["saleFees"] += sale_fees
            month_item["dividends"] += dividends

    return totals, monthly


def read_investment(
    ws,
    fund_sheet: str,
    activity: dict[str, float],
) -> dict[str, Any]:
    rows = find_data_rows(ws, start_row=5)
    if not rows:
        raise ValueError(
            f"{fund_sheet}: fondo lape nėra mėnesinių duomenų."
        )

    first_row = rows[0]
    last_row = rows[-1]

    full_name = str(ws["A1"].value or fund_sheet).strip()
    code = full_name.split()[0] if full_name else fund_sheet
    name = fund_sheet

    total_purchased = finite_number(activity.get("purchased"))
    total_invested = finite_number(activity.get("invested"))
    purchase_fees = finite_number(activity.get("purchaseFees"))
    purchased_units = finite_number(activity.get("purchasedUnits"))
    total_sold = finite_number(activity.get("sold"))
    sale_fees = finite_number(activity.get("saleFees"))
    sold_units = finite_number(activity.get("soldUnits"))
    dividends = finite_number(activity.get("dividends"))

    remaining_units = purchased_units - sold_units
    is_active = remaining_units > UNIT_TOLERANCE

    # Fondo lape paskutinė vertė gali likti istorinė ir po pilno pardavimo.
    # Dabartinė vertė skaičiuojama tik tada, kai realiai liko vienetų.
    sheet_quantity = finite_number(ws.cell(last_row, 6).value)
    price = finite_number(ws.cell(last_row, 7).value)
    sheet_value = finite_number(ws.cell(last_row, 8).value)

    quantity = remaining_units if is_active else 0.0
    current_value = (
        sheet_value
        if is_active and sheet_quantity > UNIT_TOLERANCE
        else 0.0
    )

    contributed = total_purchased + purchase_fees
    net_invested = max(total_invested - total_sold, 0.0)
    profit = (
        current_value
        + total_sold
        + dividends
        - contributed
        - sale_fees
    )
    return_rate = (
        profit / contributed * 100
        if contributed > EPSILON
        else None
    )

    return {
        "id": slugify(fund_sheet),
        "slug": slugify(fund_sheet),
        "ticker": code,
        "name": name,
        "fullName": full_name,
        "type": "fund",
        "status": "active" if is_active else "completed",
        "currency": "EUR",
        "startDate": month_start(ws.cell(first_row, 1).value),
        "endDate": (
            None if is_active else month_end(ws.cell(last_row, 1).value)
        ),
        "invested": rounded(contributed),
        "netInvested": rounded(net_invested if is_active else 0.0),
        "currentValue": rounded(current_value),
        "profit": rounded(profit),
        "returnRate": nullable_number(return_rate),
        "quantity": round(quantity, 8),
        "purchasedUnits": round(purchased_units, 8),
        "soldUnits": round(sold_units, 8),
        "price": rounded(price if is_active else 0.0, 6),
        "realizedProceeds": rounded(total_sold),
        "dividends": rounded(dividends),
        "fees": rounded(purchase_fees + sale_fees),
        "purchaseFees": rounded(purchase_fees),
        "saleFees": rounded(sale_fees),
    }


def build_history(
    overview,
    investments: list[dict[str, Any]],
    monthly_activity: dict[str, dict[str, float]],
) -> list[dict[str, Any]]:
    rows = find_data_rows(overview, start_row=3)
    if not rows:
        raise ValueError("Apžvalgos lape nėra mėnesinių duomenų.")

    investment_by_start = defaultdict(list)
    investment_by_end = defaultdict(list)
    for item in investments:
        investment_by_start[item["startDate"][:7]].append(item)
        if item["endDate"]:
            investment_by_end[item["endDate"][:7]].append(item)

    history: list[dict[str, Any]] = []
    cumulative_contributed = 0.0
    cumulative_withdrawn = 0.0

    for row in rows:
        month = month_key(overview.cell(row, 1).value)
        if not month:
            continue

        month_activity = monthly_activity.get(month, {})
        contributed_month = finite_number(
            month_activity.get("purchased")
        ) + finite_number(month_activity.get("purchaseFees"))
        withdrawn_month = finite_number(month_activity.get("sold"))
        dividends_month = finite_number(month_activity.get("dividends"))
        fees_month = (
            finite_number(month_activity.get("purchaseFees"))
            + finite_number(month_activity.get("saleFees"))
        )

        cumulative_contributed += contributed_month
        cumulative_withdrawn += withdrawn_month

        # Iki pilno pardavimo naudojame Apžvalgos rinkos vertę.
        raw_value = finite_number(overview.cell(row, 13).value)

        completed_by_month = sum(
            1
            for item in investments
            if item["endDate"] and item["endDate"][:7] <= month
        )
        active_by_month = sum(
            1
            for item in investments
            if item["startDate"][:7] <= month
            and (not item["endDate"] or item["endDate"][:7] > month)
        )

        # Jei mėnesio gale visos pozicijos parduotos, vertė turi būti nulis.
        current_value = raw_value if active_by_month > 0 else 0.0
        invested_now = max(
            cumulative_contributed - cumulative_withdrawn,
            0.0,
        )
        realized_profit = (
            cumulative_withdrawn
            + sum(
                finite_number(point["income"])
                for point in history
            )
            + dividends_month
            - cumulative_contributed
        )
        profit = (
            current_value - invested_now
            if active_by_month > 0
            else realized_profit
        )

        history.append({
            "date": f"{month}-01",
            "month": month,
            "invested": rounded(invested_now),
            "totalContributed": rounded(cumulative_contributed),
            "currentValue": rounded(current_value),
            "profit": rounded(profit),
            "returnRate": (
                round(
                    profit / cumulative_contributed * 100,
                    4,
                )
                if cumulative_contributed > EPSILON
                else None
            ),
            "cash": 0.0,
            "income": rounded(dividends_month),
            "fees": rounded(fees_month),
            "contributions": rounded(contributed_month),
            "withdrawals": rounded(withdrawn_month),
            "activeInvestments": active_by_month,
            "delayedInvestments": 0,
            "completedInvestments": completed_by_month,
        })

    return history


def build_document(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        input_path,
        data_only=True,
        read_only=False,
    )

    if OVERVIEW_SHEET not in workbook.sheetnames:
        raise ValueError(f"Nerastas privalomas lapas „{OVERVIEW_SHEET}“.")

    fund_sheets = get_fund_sheets(workbook)
    if not fund_sheets:
        raise ValueError("Nerasti fondų lapai.")

    activity, monthly_activity = scan_monthly_activity(
        workbook,
        fund_sheets,
    )

    investments = [
        read_investment(
            workbook[fund_sheet],
            fund_sheet,
            activity.get(fund_sheet, {}),
        )
        for fund_sheet in fund_sheets
    ]
    investments.sort(
        key=lambda item: (-item["currentValue"], item["name"])
    )

    active = [
        item for item in investments if item["status"] == "active"
    ]
    completed = [
        item for item in investments if item["status"] == "completed"
    ]

    history = build_history(
        workbook[OVERVIEW_SHEET],
        investments,
        monthly_activity,
    )

    current_invested = rounded(
        sum(item["netInvested"] for item in active)
    )
    current_value = rounded(
        sum(item["currentValue"] for item in active)
    )

    overview = workbook[OVERVIEW_SHEET]
    overview_rows = find_data_rows(overview, start_row=3)
    latest_overview_row = overview_rows[-1]

    # Apžvalgos B ir N stulpeliai yra galutinis apskaitos etalonas.
    # Pardavus visus fondus dabartinė investuota suma ir vertė yra 0,
    # o istorinis įnašas bei realizuotas pelnas išlieka atskirai.
    total_contributed = rounded(
        overview.cell(latest_overview_row, 2).value
    )
    realized_profit = rounded(
        overview.cell(latest_overview_row, 14).value
    )
    total_dividends = rounded(
        sum(item["dividends"] for item in investments)
    )
    total_fees = rounded(
        sum(item["fees"] for item in investments)
    )

    profit = rounded(
        current_value - current_invested
        if active
        else realized_profit
    )
    overview_return_raw = overview.cell(
        latest_overview_row, 15
    ).value
    return_rate = (
        round(finite_number(overview_return_raw) * 100, 4)
        if overview_return_raw is not None
        else (
            round(realized_profit / total_contributed * 100, 4)
            if total_contributed > EPSILON
            else None
        )
    )

    latest = history[-1]
    latest["invested"] = current_invested
    latest["currentValue"] = current_value
    latest["profit"] = profit
    latest["activeInvestments"] = len(active)
    latest["completedInvestments"] = len(completed)

    document = {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAt": datetime.now().astimezone().isoformat(
            timespec="seconds"
        ),
        "platform": {
            "id": "seb-fondai",
            "slug": "seb-fondai",
            "name": "SEB Fondai",
            "group": "funds",
            "type": "funds",
            "category": "Investiciniai fondai",
            "currency": "EUR",
            "active": bool(active),
            "startDate": history[0]["date"],
            "updatedAt": month_end(history[-1]["month"]),
            "website": "https://www.seb.lt",
        },
        "summary": {
            "invested": current_invested,
            "netInvested": current_invested,
            "totalContributed": total_contributed,
            "currentValue": current_value,
            "profit": profit,
            "realizedProfit": realized_profit,
            "returnRate": return_rate,
            "xirr": None,
            "cash": 0.0,
            "incomeReceived": total_dividends,
            "fees": total_fees,
            "activeInvestments": len(active),
            "delayedInvestments": 0,
            "completedInvestments": len(completed),
            "averageRate": None,
            "averageLtv": None,
            "totalInvestments": len(investments),
        },
        "history": history,
        "investments": investments,
        "distributions": {
            "status": [
                {
                    "key": "active",
                    "label": "Aktyvūs",
                    "count": len(active),
                    "value": current_value,
                },
                {
                    "key": "completed",
                    "label": "Parduoti",
                    "count": len(completed),
                    "value": rounded(
                        sum(
                            item["realizedProceeds"]
                            for item in completed
                        )
                    ),
                },
            ],
            "holdings": [
                {
                    "key": item["ticker"],
                    "label": item["ticker"],
                    "name": item["name"],
                    "value": item["currentValue"],
                    "weight": (
                        round(
                            item["currentValue"]
                            / current_value
                            * 100,
                            4,
                        )
                        if current_value > EPSILON
                        else 0.0
                    ),
                }
                for item in active
            ],
        },
        "latestMonth": latest,
        "largestInvestment": (
            {
                "id": active[0]["id"],
                "ticker": active[0]["ticker"],
                "name": active[0]["name"],
                "currentValue": active[0]["currentValue"],
            }
            if active
            else None
        ),
        "source": {
            "file": input_path.name,
            "overviewSheet": OVERVIEW_SHEET,
            "instrumentSheets": fund_sheets,
            "monthlySheets": [
                name
                for name in workbook.sheetnames
                if MONTH_SHEET_RE.fullmatch(name)
            ],
        },
    }

    validate_document(document)
    return document


def validate_document(document: dict[str, Any]) -> None:
    summary = document["summary"]
    investments = document["investments"]
    errors: list[str] = []

    if not document["history"]:
        errors.append("History yra tuščias.")
    if not investments:
        errors.append("Investicijų sąrašas yra tuščias.")

    active_count = sum(
        item["status"] == "active" for item in investments
    )
    completed_count = sum(
        item["status"] == "completed" for item in investments
    )

    if active_count != summary["activeInvestments"]:
        errors.append("Nesutampa aktyvių fondų skaičius.")
    if completed_count != summary["completedInvestments"]:
        errors.append("Nesutampa parduotų fondų skaičius.")

    current_value = rounded(
        sum(
            item["currentValue"]
            for item in investments
            if item["status"] == "active"
        )
    )
    if abs(current_value - summary["currentValue"]) > 0.05:
        errors.append("Nesutampa aktyvių fondų dabartinė vertė.")

    ids = [item["id"] for item in investments]
    if len(ids) != len(set(ids)):
        errors.append("Rasti dubliuoti fondų ID.")

    if errors:
        raise ValueError("\n".join(errors))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="SEB Fondai Excel importeris į Portfolio V2 JSON."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Excel failas. Numatyta: {DEFAULT_INPUT}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"JSON failas. Numatyta: {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    print("=" * 64)
    print("SEB FONDAI IMPORTER V1")
    print("=" * 64)
    print(f"Excel failas: {input_path}")

    if not input_path.is_file():
        raise FileNotFoundError(f"Excel failas nerastas: {input_path}")

    document = build_document(input_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        json.dump(document, file, ensure_ascii=False, indent=2)

    summary = document["summary"]
    print(f"✅ Nuskaityta fondų: {summary['totalInvestments']}")
    print(f"✅ Aktyvių: {summary['activeInvestments']}")
    print(f"✅ Parduotų: {summary['completedInvestments']}")
    print(f"✅ Investuota dabar: {summary['invested']:.2f} EUR")
    print(f"✅ Istoriškai įnešta: {summary['totalContributed']:.2f} EUR")
    print(f"✅ Fondų vertė: {summary['currentValue']:.2f} EUR")
    print(f"✅ Realizuotas pelnas: {summary['realizedProfit']:.2f} EUR")
    print(f"✅ JSON sukurtas: {output_path}")


if __name__ == "__main__":
    main()
