from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "excel" / "Investavimas.xlsx"
DEFAULT_OUTPUT = ROOT / "public" / "data" / "platform_history.json"


PLATFORM_ALIASES = {
    "rontgen": "rontgen",
    "rontgen-platform": "rontgen",
    "revolut-brokerage": "revolut-brokerage",
    "revolut-robo": "revolut-robo",
    "seb-fondai": "seb-fondai",
    "seb-mikro": "seb-mikro",
    "seb-robo": "seb-robo",
}


def slugify(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.translate(
        str.maketrans(
            {
                "ą": "a",
                "č": "c",
                "ę": "e",
                "ė": "e",
                "į": "i",
                "š": "s",
                "ų": "u",
                "ū": "u",
                "ž": "z",
                "ö": "o",
            }
        )
    )
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return PLATFORM_ALIASES.get(text, text)


def to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0

    if isinstance(value, bool):
        return float(value)

    if isinstance(value, (int, float)):
        return round(float(value), 4)

    text = (
        str(value)
        .strip()
        .replace("\u00a0", "")
        .replace("€", "")
        .replace("%", "")
        .replace(" ", "")
        .replace(",", ".")
    )

    try:
        return round(float(text), 4)
    except ValueError:
        return 0.0


def to_iso_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if value in (None, ""):
        return None

    text = str(value).strip()

    for fmt in (
        "%Y-%m-%d",
        "%Y.%m.%d",
        "%d.%m.%Y",
        "%Y/%m/%d",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    return None


def merged_value(ws, row: int, column: int) -> Any:
    cell = ws.cell(row=row, column=column)

    if cell.value not in (None, ""):
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value

    return None


def find_history_sheet(workbook):
    preferred_names = (
        "Investavimas",
        "investavimas",
        "Istorija",
        "istorija",
        "History",
        "history",
    )

    for name in preferred_names:
        if name in workbook.sheetnames:
            return workbook[name]

    return workbook[workbook.sheetnames[0]]


def find_date_column(ws) -> int:
    possible_names = {"data", "date", "menuo", "mėnuo", "periodas"}

    for row in range(1, min(ws.max_row, 15) + 1):
        for column in range(1, min(ws.max_column, 15) + 1):
            if slugify(ws.cell(row=row, column=column).value) in {
                slugify(name) for name in possible_names
            }:
                return column

    return 1


def metric_kind(value: Any) -> str | None:
    raw = str(value or "").strip()
    key = slugify(raw)

    if key in {
        "inesta",
        "investuota",
        "invested",
        "kapitalas",
        "ideta",
        "įdėta",
    }:
        return "invested"

    if key in {
        "verte",
        "value",
        "dabartine-verte",
        "portfelio-verte",
    }:
        return "value"

    if key in {
        "graza",
        "return",
        "return-rate",
        "roi",
        "proc",
        "procentai",
    } or raw == "%":
        return "returnRate"

    return None


def find_header_rows(ws) -> tuple[int, int]:
    for metric_row in range(1, min(ws.max_row, 15) + 1):
        metric_matches = 0

        for column in range(1, ws.max_column + 1):
            if metric_kind(merged_value(ws, metric_row, column)):
                metric_matches += 1

        if metric_matches >= 2:
            platform_row = max(1, metric_row - 1)
            return platform_row, metric_row

    raise RuntimeError(
        "Nepavyko rasti platformų antraščių. "
        "Tikėtasi platformos pavadinimo eilutės ir po ja esančių "
        "„Investuota / Vertė / %“ stulpelių."
    )


def calculate_return_rate(invested: float, value: float) -> float:
    if invested == 0:
        return 0.0

    return round(((value - invested) / invested) * 100, 4)


def collect_platform_columns(ws) -> tuple[int, int, int, dict[str, dict[str, Any]]]:
    platform_row, metric_row = find_header_rows(ws)
    date_column = find_date_column(ws)

    platforms: dict[str, dict[str, Any]] = {}
    current_platform_name: str | None = None

    for column in range(1, ws.max_column + 1):
        platform_value = merged_value(ws, platform_row, column)

        if platform_value not in (None, ""):
            current_platform_name = str(platform_value).strip()

        metric = metric_kind(merged_value(ws, metric_row, column))

        if not current_platform_name or not metric:
            continue

        platform_slug = slugify(current_platform_name)

        if platform_slug in {"data", "date", "menuo", "periodas"}:
            continue

        platform_meta = platforms.setdefault(
            platform_slug,
            {
                "name": current_platform_name,
                "columns": {},
            },
        )
        platform_meta["columns"][metric] = column

    return platform_row, metric_row, date_column, platforms


def build_platform_history(excel_path: Path) -> dict[str, Any]:
    workbook = load_workbook(
        excel_path,
        data_only=True,
        read_only=False,
    )
    ws = find_history_sheet(workbook)

    _, metric_row, date_column, platform_columns = collect_platform_columns(ws)
    first_data_row = metric_row + 1

    platforms: dict[str, Any] = {}

    for platform_slug, platform_meta in platform_columns.items():
        columns = platform_meta["columns"]
        history: list[dict[str, Any]] = []

        invested_column = columns.get("invested")
        value_column = columns.get("value")

        if not invested_column or not value_column:
            continue

        for row in range(first_data_row, ws.max_row + 1):
            row_date = to_iso_date(ws.cell(row=row, column=date_column).value)

            if not row_date:
                continue

            invested = to_number(
                ws.cell(row=row, column=invested_column).value
            )
            value = to_number(
                ws.cell(row=row, column=value_column).value
            )

            if invested == 0 and value == 0:
                continue

            profit = round(value - invested, 4)
            return_rate = calculate_return_rate(invested, value)

            history.append(
                {
                    "date": row_date,
                    "invested": invested,
                    "value": value,
                    "profit": profit,
                    "returnRate": return_rate,
                    "source": excel_path.name,
                }
            )

        history.sort(key=lambda item: item["date"])

        if history:
            platforms[platform_slug] = {
                "name": platform_meta["name"],
                "history": history,
            }

    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")

    return {
        "schemaVersion": 1,
        "type": "platform_history",
        "currency": "EUR",
        "generatedAt": generated_at,
        "source": {
            "file": excel_path.name,
            "sheet": ws.title,
        },
        "platforms": platforms,
    }


def resolve_path(argument: str | None, default_path: Path) -> Path:
    if not argument:
        return default_path.resolve()

    path = Path(argument)

    if not path.is_absolute():
        path = ROOT / path

    return path.resolve()


def main() -> int:
    excel_path = resolve_path(
        sys.argv[1] if len(sys.argv) > 1 else None,
        DEFAULT_EXCEL,
    )
    output_path = resolve_path(
        sys.argv[2] if len(sys.argv) > 2 else None,
        DEFAULT_OUTPUT,
    )

    if not excel_path.exists():
        print()
        print("KLAIDA: nerastas istorinis Excel failas:")
        print(f"  {excel_path}")
        print()
        print("Numatytoji vieta:")
        print(f"  {DEFAULT_EXCEL}")
        return 1

    try:
        payload = build_platform_history(excel_path)
    except Exception as error:
        print()
        print("KLAIDA generuojant platformų istoriją:")
        print(f"  {error}")
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    platform_count = len(payload["platforms"])
    history_points = sum(
        len(platform["history"])
        for platform in payload["platforms"].values()
    )

    print()
    print("Platformų istorija atnaujinta.")
    print(f"Excel:      {excel_path}")
    print(f"JSON:       {output_path}")
    print(f"Platformų:  {platform_count}")
    print(f"Įrašų:      {history_points}")
    print()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
