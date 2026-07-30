from __future__ import annotations

import argparse
import json
import math
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_NAME = "Investavimas"
DEFAULT_INPUT_CANDIDATES = (
    Path("excel/Investavimas.xlsx"),
    Path("Investavimas.xlsx"),
)
DEFAULT_OUTPUT_DIR = Path("public/data")


@dataclass(frozen=True)
class HistorySection:
    name: str
    output_filename: str
    date_column: str
    invested_column: str
    monthly_contribution_column: str
    value_column: str
    profit_column: str
    return_rate_column: str
    monthly_result_column: str


SECTIONS = (
    HistorySection(
        name="Fondai",
        output_filename="funds_history.json",
        date_column="BN",
        invested_column="BO",
        monthly_contribution_column="BP",
        value_column="BQ",
        profit_column="BR",
        return_rate_column="BS",
        monthly_result_column="BT",
    ),
    HistorySection(
        name="P2P",
        output_filename="p2p_history.json",
        date_column="BV",
        invested_column="BW",
        monthly_contribution_column="BX",
        value_column="BY",
        profit_column="BZ",
        return_rate_column="CA",
        monthly_result_column="CB",
    ),
    HistorySection(
        name="Visas portfelis",
        output_filename="portfolio_history.json",
        date_column="CD",
        invested_column="CE",
        monthly_contribution_column="CF",
        value_column="CG",
        profit_column="CH",
        return_rate_column="CI",
        monthly_result_column="CJ",
    ),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Importuoja Fondų, P2P ir viso portfelio mėnesinę istoriją "
            "iš Excel lapo 'Investavimas'."
        )
    )
    parser.add_argument(
        "input",
        nargs="?",
        type=Path,
        help="Kelias iki Investavimas.xlsx. Jei nenurodytas, ieškoma excel/Investavimas.xlsx.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="JSON failų aplankas (numatyta: public/data).",
    )
    return parser.parse_args()


def resolve_input_path(explicit_path: Path | None) -> Path:
    if explicit_path is not None:
        path = explicit_path.expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Excel failas nerastas: {path}")
        return path

    for candidate in DEFAULT_INPUT_CANDIDATES:
        if candidate.is_file():
            return candidate.resolve()

    searched = ", ".join(str(path) for path in DEFAULT_INPUT_CANDIDATES)
    raise FileNotFoundError(
        "Investavimas.xlsx nerastas. Ieškota: " + searched
    )


def clean_number(value: Any, *, field: str, row: int) -> float:
    if value in (None, ""):
        return 0.0

    if isinstance(value, bool):
        raise ValueError(f"{field}, eilutė {row}: rasta loginė reikšmė {value!r}.")

    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{field}, eilutė {row}: ne skaičius {value!r}. "
            "Patikrink, ar Excel formulės perskaičiuotos."
        ) from exc

    if not math.isfinite(number):
        raise ValueError(f"{field}, eilutė {row}: netinkamas skaičius {number!r}.")

    # Pašalina Excel slankaus kablelio triukšmą, pvz. 12166.860000000001.
    return round(number, 2)


def clean_rate(value: Any, *, field: str, row: int) -> float:
    if value in (None, ""):
        return 0.0

    try:
        rate = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"{field}, eilutė {row}: ne skaičius {value!r}."
        ) from exc

    if not math.isfinite(rate):
        raise ValueError(f"{field}, eilutė {row}: netinkama reikšmė {rate!r}.")

    # Excel saugo 13,39 % kaip 0.1339; JSON saugome procentiniais punktais.
    return round(rate * 100, 2)


def clean_date(value: Any, *, row: int, section: HistorySection) -> str | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            parsed = datetime.fromisoformat(text).date()
        except ValueError as exc:
            raise ValueError(
                f"{section.name}, eilutė {row}: neatpažinta data {value!r}."
            ) from exc
    else:
        raise ValueError(
            f"{section.name}, eilutė {row}: neatpažinta data {value!r}."
        )

    return parsed.isoformat()


def read_section(ws: Any, section: HistorySection) -> list[dict[str, Any]]:
    rows_by_date: dict[str, dict[str, Any]] = {}

    for row in range(3, ws.max_row + 1):
        date_value = clean_date(
            ws[f"{section.date_column}{row}"].value,
            row=row,
            section=section,
        )
        if date_value is None:
            continue

        invested = clean_number(
            ws[f"{section.invested_column}{row}"].value,
            field=f"{section.name} / įnešta",
            row=row,
        )
        monthly_contribution = clean_number(
            ws[f"{section.monthly_contribution_column}{row}"].value,
            field=f"{section.name} / per mėnesį įnešta",
            row=row,
        )
        current_value = clean_number(
            ws[f"{section.value_column}{row}"].value,
            field=f"{section.name} / vertė",
            row=row,
        )
        profit = clean_number(
            ws[f"{section.profit_column}{row}"].value,
            field=f"{section.name} / prieaugis",
            row=row,
        )
        return_rate = clean_rate(
            ws[f"{section.return_rate_column}{row}"].value,
            field=f"{section.name} / grąža",
            row=row,
        )
        monthly_result = clean_number(
            ws[f"{section.monthly_result_column}{row}"].value,
            field=f"{section.name} / mėnesio rezultatas",
            row=row,
        )

        # Tuščios būsimos formulės eilutės nepridedamos.
        if (
            invested == 0
            and monthly_contribution == 0
            and current_value == 0
            and profit == 0
            and monthly_result == 0
        ):
            continue

        rows_by_date[date_value] = {
            "date": date_value,
            "invested": invested,
            "monthlyContribution": monthly_contribution,
            "value": current_value,
            "profit": profit,
            "returnRate": return_rate,
            "monthlyResult": monthly_result,
        }

    history = [rows_by_date[key] for key in sorted(rows_by_date)]

    if not history:
        raise ValueError(f"Skyriuje '{section.name}' nerasta istorinių eilučių.")

    return history


def build_payload(
    *,
    section: HistorySection,
    history: list[dict[str, Any]],
    source_path: Path,
) -> dict[str, Any]:
    return {
        "schemaVersion": 1,
        "type": "monthlyPortfolioHistory",
        "section": section.name,
        "currency": "EUR",
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": {
            "file": source_path.name,
            "sheet": SHEET_NAME,
        },
        "period": {
            "from": history[0]["date"],
            "to": history[-1]["date"],
            "months": len(history),
        },
        "latest": history[-1],
        "history": history,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")

    temporary_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary_path.replace(path)


def main() -> int:
    args = parse_args()

    try:
        input_path = resolve_input_path(args.input)
        output_dir = args.output_dir.expanduser().resolve()

        print("=" * 66)
        print("PORTFELIO ISTORIJOS IMPORTERIS V1")
        print("=" * 66)
        print(f"Excel:  {input_path}")
        print(f"Išvestis: {output_dir}")
        print()

        workbook = load_workbook(
            input_path,
            data_only=True,
            read_only=True,
        )

        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Lapas '{SHEET_NAME}' nerastas. Rasti lapai: "
                + ", ".join(workbook.sheetnames)
            )

        worksheet = workbook[SHEET_NAME]

        for section in SECTIONS:
            history = read_section(worksheet, section)
            payload = build_payload(
                section=section,
                history=history,
                source_path=input_path,
            )
            output_path = output_dir / section.output_filename
            write_json(output_path, payload)

            latest = history[-1]
            print(f"[OK] {section.name}")
            print(f"     Failas: {output_path}")
            print(
                f"     Laikotarpis: {history[0]['date']} – {latest['date']} "
                f"({len(history)} taškai)"
            )
            print(
                f"     Naujausia vertė: {latest['value']:.2f} EUR | "
                f"Įnešta: {latest['invested']:.2f} EUR | "
                f"Grąža: {latest['returnRate']:.2f} %"
            )
            print()

        print("Importas baigtas sėkmingai.")
        return 0

    except Exception as exc:
        print(f"KLAIDA: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
