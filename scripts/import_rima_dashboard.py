from __future__ import annotations

import json
import math
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

PROJECT_ROOT = Path(__file__).resolve().parents[1]
EXCEL_FILE = PROJECT_ROOT / "excel" / "Investavimas Rima.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "public" / "data" / "rima"
SHEET_NAME = "Investavimas"


def as_number(value: Any, *, field: str, row: int) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, bool):
        raise ValueError(f"{field}, eilutė {row}: rasta loginė reikšmė {value!r}.")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field}, eilutė {row}: ne skaičius {value!r}.") from exc
    if not math.isfinite(number):
        raise ValueError(f"{field}, eilutė {row}: netinkamas skaičius {number!r}.")
    return round(number, 2)


def as_rate(value: Any, *, field: str, row: int) -> float:
    if value in (None, ""):
        return 0.0
    try:
        rate = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field}, eilutė {row}: ne skaičius {value!r}.") from exc
    if not math.isfinite(rate):
        raise ValueError(f"{field}, eilutė {row}: netinkama reikšmė {rate!r}.")
    return round(rate * 100.0, 2)


def as_date(value: Any, epoch) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and value > 0:
        converted = from_excel(value, epoch)
        if isinstance(converted, datetime):
            return converted.date().isoformat()
        if isinstance(converted, date):
            return converted.isoformat()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass
    raise ValueError(f"Neatpažinta data: {value!r}")


def find_section_start(ws, title: str) -> int:
    wanted = title.strip().casefold()
    for cell in ws[1]:
        if isinstance(cell.value, str) and cell.value.strip().casefold() == wanted:
            second = ws.cell(row=2, column=cell.column).value
            if isinstance(second, str) and second.strip().casefold() == "data":
                return cell.column
    raise ValueError(f"Nerastas suvestinės blokas „{title}“.")


def read_history(ws, title: str, epoch) -> list[dict[str, Any]]:
    start = find_section_start(ws, title)
    history_by_date: dict[str, dict[str, Any]] = {}
    for row in range(3, ws.max_row + 1):
        period = as_date(ws.cell(row=row, column=start).value, epoch)
        if not period:
            continue
        invested = as_number(ws.cell(row=row, column=start + 1).value, field=f"{title} / įnešta", row=row)
        monthly_contribution = as_number(ws.cell(row=row, column=start + 2).value, field=f"{title} / per mėn. įnešta", row=row)
        value = as_number(ws.cell(row=row, column=start + 3).value, field=f"{title} / vertė", row=row)
        profit = as_number(ws.cell(row=row, column=start + 4).value, field=f"{title} / prieaugis", row=row)
        return_rate = as_rate(ws.cell(row=row, column=start + 5).value, field=f"{title} / grąža", row=row)
        monthly_result = as_number(ws.cell(row=row, column=start + 6).value, field=f"{title} / mėnesio rezultatas", row=row)
        if invested == 0 and monthly_contribution == 0 and value == 0 and profit == 0 and monthly_result == 0:
            continue
        history_by_date[period] = {
            "date": period,
            "invested": invested,
            "monthlyContribution": monthly_contribution,
            "value": value,
            "profit": profit,
            "returnRate": return_rate,
            "monthlyResult": monthly_result,
        }
    history = [history_by_date[key] for key in sorted(history_by_date)]
    if not history:
        raise ValueError(f"Skyriuje „{title}“ nerasta istorinių eilučių.")
    return history


def build_payload(section: str, history: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "schemaVersion": 1,
        "type": "monthlyPortfolioHistory",
        "section": section,
        "currency": "EUR",
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": {"file": EXCEL_FILE.name, "sheet": SHEET_NAME},
        "period": {"from": history[0]["date"], "to": history[-1]["date"], "months": len(history)},
        "latest": history[-1],
        "history": history,
    }


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    temporary.replace(path)


def print_result(label: str, history: list[dict[str, Any]]) -> None:
    latest = history[-1]
    print(f"[OK] {label}: {len(history)} mėn. {latest['date']}, {latest['value']:.2f} €")


def main() -> int:
    print("=" * 66)
    print("RIMOS DASHBOARD IMPORTERIS V2.5.4.1")
    print("=" * 66)
    print(f"Excel:    {EXCEL_FILE}")
    print(f"Išvestis: {OUTPUT_DIR}")
    if not EXCEL_FILE.is_file():
        print(f"KLAIDA: nerastas Excel failas: {EXCEL_FILE}")
        return 1
    workbook = load_workbook(EXCEL_FILE, data_only=True, read_only=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Excel faile nerastas lapas „{SHEET_NAME}“.")
        ws = workbook[SHEET_NAME]
        sections = (
            ("Fondai", "funds_history.json"),
            ("P2P", "p2p_history.json"),
            ("Viso", "portfolio_history.json"),
        )
        for title, filename in sections:
            history = read_history(ws, title, workbook.epoch)
            write_json(OUTPUT_DIR / filename, build_payload("Visas portfelis" if title == "Viso" else title, history))
            print_result("Visas portfelis" if title == "Viso" else title, history)
        print("\nRimos Dashboard duomenys atnaujinti sėkmingai.")
        return 0
    finally:
        workbook.close()


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"KLAIDA: {error}", file=sys.stderr)
        raise SystemExit(1)
