from pathlib import Path
import json
import statistics
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXCEL_FILE = ROOT / "excel" / "Crowdpear.xlsx"
OUTPUT_FILE = ROOT / "public" / "data" / "crowdpear.json"

def iso(value):
    return value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else None

def number(value):
    return float(value) if isinstance(value, (int, float)) else 0.0

def main():
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Nerastas failas: {EXCEL_FILE}")

    workbook = load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    overview = workbook["Overview"]

    project_meta = {}
    for sheet_name in workbook.sheetnames[2:]:
        sheet = workbook[sheet_name]
        name = str(sheet["B2"].value or sheet_name)
        project_meta[name] = {
            "loanCode": str(sheet["B1"].value or ""),
            "ltv": round(number(sheet["B6"].value) * 100, 2),
            "maxLtv": round(number(sheet["B7"].value) * 100, 2),
            "paymentFrequency": str(sheet["B9"].value or "—"),
        }

    loans = []
    for row in range(3, overview.max_row + 1):
        loan_id = overview.cell(row, 1).value
        name = overview.cell(row, 3).value
        if not loan_id or not name:
            continue

        invested = number(overview.cell(row, 7).value)
        actual_repayment = overview.cell(row, 9).value
        delayed_days = number(overview.cell(row, 11).value)
        repaid_principal = number(overview.cell(row, 12).value)
        interest_received = number(overview.cell(row, 13).value)
        meta = project_meta.get(str(name), {})

        if actual_repayment:
            status = "repaid"
        elif delayed_days > 0:
            status = "delayed"
        else:
            status = "active"

        loans.append({
            "id": int(loan_id),
            "name": str(name),
            "loanCode": meta.get("loanCode", ""),
            "investmentDate": iso(overview.cell(row, 2).value),
            "durationMonths": int(number(overview.cell(row, 4).value)),
            "rating": str(overview.cell(row, 5).value or "—"),
            "interestRate": round(number(overview.cell(row, 6).value) * 100, 2),
            "ltv": meta.get("ltv", 0),
            "maxLtv": meta.get("maxLtv", 0),
            "paymentFrequency": meta.get("paymentFrequency", "—"),
            "invested": round(invested, 2),
            "outstanding": round(max(invested - repaid_principal, 0), 2),
            "plannedRepayment": iso(overview.cell(row, 8).value),
            "actualRepayment": iso(actual_repayment),
            "remainingDays": overview.cell(row, 10).value,
            "delayedDays": int(delayed_days),
            "repaidPrincipal": round(repaid_principal, 2),
            "interestReceived": round(interest_received, 2),
            "status": status,
        })

    weighted_base = sum(item["invested"] for item in loans) or 1
    summary = {
        "deposited": round(number(overview["O3"].value), 2),
        "invested": round(number(overview["P3"].value), 2),
        "bonuses": round(number(overview["Q3"].value), 2),
        "cash": round(number(overview["R3"].value), 2),
        "repaidPrincipal": round(number(overview["S3"].value), 2),
        "interestReceived": round(number(overview["T3"].value), 2),
        "portfolioValue": round(number(overview["U3"].value), 2),
        "profit": round(number(overview["V3"].value), 2),
        "roi": round(number(overview["W3"].value) * 100, 2),
        "xirr": round(number(overview["X3"].value) * 100, 2),
        "activeLoans": sum(item["status"] == "active" for item in loans),
        "delayedLoans": sum(item["status"] == "delayed" for item in loans),
        "repaidLoans": sum(item["status"] == "repaid" for item in loans),
        "averageInterest": round(
            sum(item["interestRate"] * item["invested"] for item in loans) / weighted_base, 2
        ),
        "averageLtv": round(
            sum(item["ltv"] * item["invested"] for item in loans) / weighted_base, 2
        ),
        "averageDuration": round(statistics.mean(item["durationMonths"] for item in loans), 1),
        "outstandingPrincipal": round(sum(item["outstanding"] for item in loans), 2),
    }

    ratings = {}
    for item in loans:
        ratings[item["rating"]] = round(
            ratings.get(item["rating"], 0) + item["invested"], 2
        )

    result = {
        "platform": {
            "name": "Crowdpear",
            "type": "NT sutelktinis finansavimas",
            "status": "active",
            "currency": "EUR",
            "website": "https://crowdpear.com",
        },
        "summary": summary,
        "ratingAllocation": [
            {"name": name, "value": amount}
            for name, amount in sorted(ratings.items())
        ],
        "loans": loans,
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Sukurta: {OUTPUT_FILE}")
    print(f"Projektų: {len(loans)}")

if __name__ == "__main__":
    main()
