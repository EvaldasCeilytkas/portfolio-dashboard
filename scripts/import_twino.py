from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SCHEMA_VERSION = 1
EPSILON = 0.005
NON_LOAN_SHEETS = {"overview", "pinigai"}
LOAN_SHEET_RE = re.compile(r"^.+$")


def finite_number(value: Any, default: float = 0.0) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError):
        if isinstance(value, str):
            text = value.strip().replace("\u00a0", "").replace(" ", "")
            text = text.replace(",", ".")
            try:
                result = float(text)
            except ValueError:
                return default
        else:
            return default
    return result if math.isfinite(result) else default


def rounded(value: Any, digits: int = 2) -> float:
    return round(finite_number(value), digits)


def parse_date(value: Any, epoch: datetime) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                continue
    return None


def iso_date(value: Any, epoch: datetime) -> str | None:
    parsed = parse_date(value, epoch)
    return parsed.isoformat() if parsed else None


def month_end(value: date) -> date:
    if value.month == 12:
        return date(value.year, 12, 31)
    first_next = date(value.year, value.month + 1, 1)
    return date.fromordinal(first_next.toordinal() - 1)


def month_range(start: date, end: date) -> Iterable[date]:
    current = date(start.year, start.month, 1)
    final = date(end.year, end.month, 1)
    while current <= final:
        yield month_end(current)
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)


def find_input(project_root: Path, explicit: str | None) -> Path:
    if explicit:
        path = Path(explicit).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Nerastas Excel failas: {path}")
        return path

    excel_dir = project_root / "excel"
    preferred = [
        excel_dir / "TWINO.xlsx",
        project_root / "TWINO.xlsx",
    ]
    for path in preferred:
        if path.is_file():
            return path

    candidates = sorted(excel_dir.glob("TWINO*.xlsx")) if excel_dir.exists() else []
    if candidates:
        return candidates[-1]

    raise FileNotFoundError(
        "Nerastas TWINO Excel failas. Įdėk jį į excel/TWINO.xlsx "
        "arba naudok --input."
    )


def read_overview(ws, epoch: datetime) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for row in range(3, ws.max_row + 1):
        loan_code = str(ws.cell(row, 3).value or "").strip()
        if not loan_code:
            continue

        rows.append({
            "number": int(finite_number(ws.cell(row, 1).value)),
            "investmentDate": iso_date(ws.cell(row, 2).value, epoch),
            "loanCode": loan_code,
            "country": str(ws.cell(row, 4).value or "").strip() or None,
            "loanType": str(ws.cell(row, 5).value or "").strip() or None,
            "term": str(ws.cell(row, 6).value or "").strip() or None,
            "rate": rounded(finite_number(ws.cell(row, 7).value) * 100, 4),
            "amount": rounded(ws.cell(row, 8).value),
            "plannedEndDate": iso_date(ws.cell(row, 9).value, epoch),
            "actualEndDate": iso_date(ws.cell(row, 10).value, epoch),
            "daysRemaining": int(finite_number(ws.cell(row, 11).value)),
            "delayDays": int(finite_number(ws.cell(row, 12).value)),
            "principalReturned": rounded(ws.cell(row, 13).value),
            "interestReceived": rounded(ws.cell(row, 14).value),
            "fees": rounded(ws.cell(row, 15).value),
        })

    summary = {
        "deposited": rounded(ws["Q3"].value),
        "lifetimeInvested": rounded(ws["R3"].value),
        "bonuses": rounded(ws["S3"].value),
        "cash": rounded(ws["T3"].value),
        "principalReturned": rounded(ws["U3"].value),
        "interestReceived": rounded(ws["V3"].value),
        "currentValue": rounded(ws["W3"].value),
        "profit": rounded(ws["X3"].value),
        "returnRate": rounded(finite_number(ws["Y3"].value) * 100, 4),
        "xirr": rounded(finite_number(ws["Z3"].value) * 100, 4),
        "fees": rounded(ws["AA3"].value),
        "overviewDelayDays": int(finite_number(ws["AC3"].value)),
    }
    return rows, summary


def read_money_sheet(ws, epoch: datetime) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        tx_date = parse_date(ws.cell(row, 1).value, epoch)
        deposited = rounded(ws.cell(row, 2).value)
        withdrawn = rounded(ws.cell(row, 3).value)
        bonus = rounded(ws.cell(row, 4).value)

        if not tx_date and all(abs(v) <= EPSILON for v in (deposited, withdrawn, bonus)):
            continue

        result.append({
            "date": tx_date.isoformat() if tx_date else None,
            "deposited": deposited,
            "withdrawn": withdrawn,
            "bonus": bonus,
        })
    return result


def read_loan_sheet(ws, epoch: datetime, today: date) -> dict[str, Any]:
    if str(ws["A1"].value or "").strip().lower() != "loan id":
        raise ValueError(f"{ws.title}: A1 nėra 'Loan ID'.")

    loan_code = str(ws["B1"].value or ws.title).strip()
    country = str(ws["B2"].value or "").strip() or None
    lender = str(ws["B3"].value or "").strip() or None
    loan_type = str(ws["B4"].value or "").strip() or None
    external_id = str(ws["B5"].value or "").strip() or None
    trust_score = str(ws["B6"].value or "").strip() or None
    term = str(ws["B8"].value or "").strip() or None
    investment_date = parse_date(ws["B9"].value, epoch)
    start_date = parse_date(ws["B10"].value, epoch)
    planned_end = parse_date(ws["B11"].value, epoch)
    nominal_rate = rounded(finite_number(ws["B15"].value) * 100, 4)
    invested = rounded(ws["B16"].value)

    schedule: list[dict[str, Any]] = []
    overdue_unpaid: list[dict[str, Any]] = []

    # D:I – planas ir faktiniai mokėjimai.
    for row in range(3, ws.max_row + 1):
        planned_date = parse_date(ws.cell(row, 4).value, epoch)
        actual_date = parse_date(ws.cell(row, 5).value, epoch)
        timing_days = int(finite_number(ws.cell(row, 6).value))
        principal = rounded(ws.cell(row, 7).value)
        interest = rounded(ws.cell(row, 8).value)
        fee = rounded(ws.cell(row, 9).value)

        # Eilutės be planuotos ir faktinės datos yra Excel suvestinės,
        # o ne realūs mokėjimai.
        if not planned_date and not actual_date:
            continue

        item = {
            "plannedDate": planned_date.isoformat() if planned_date else None,
            "actualDate": actual_date.isoformat() if actual_date else None,
            "timingDays": timing_days,
            "principal": principal,
            "interest": interest,
            "fee": fee,
            "paid": actual_date is not None,
        }
        schedule.append(item)

        if (
            planned_date is not None
            and planned_date < today
            and actual_date is None
        ):
            overdue_unpaid.append(item)

    # K:Q – bendra grąžinimo suvestinė.
    principal_returned = rounded(ws["K3"].value)
    interest_received = rounded(ws["L3"].value)
    fees = rounded(ws["M3"].value)
    xirr_value = ws["N3"].value
    xirr = (
        rounded(finite_number(xirr_value) * 100, 4)
        if xirr_value not in (None, "")
        else None
    )
    actual_end = parse_date(ws["O3"].value, epoch)
    overview_delay = int(finite_number(ws["P3"].value))
    early_late_days = int(finite_number(ws["Q3"].value))

    # X:Y – realūs pinigų srautai.
    cash_flows: list[dict[str, Any]] = []
    for row in range(3, ws.max_row + 1):
        flow_date = parse_date(ws.cell(row, 24).value, epoch)
        amount = rounded(ws.cell(row, 25).value)
        delay_days = int(finite_number(ws.cell(row, 26).value))
        if not flow_date and abs(amount) <= EPSILON:
            continue
        cash_flows.append({
            "date": flow_date.isoformat() if flow_date else None,
            "amount": amount,
            "delayDays": delay_days,
        })

    completed = actual_end is not None or (
        principal_returned >= invested - EPSILON and invested > EPSILON
    )
    current_principal = 0.0 if completed else rounded(max(invested - principal_returned, 0))
    calculated_delay = max(
        [
            today.toordinal()
            - date.fromisoformat(item["plannedDate"]).toordinal()
            for item in overdue_unpaid
            if item["plannedDate"]
        ]
        + [0]
    )
    actual_delay_days = max(
        calculated_delay,
        overview_delay if overdue_unpaid else 0,
    )
    status = "completed" if completed else ("delayed" if actual_delay_days > 0 else "active")

    return {
        "id": loan_code,
        "slug": loan_code.lower(),
        "loanCode": loan_code,
        "externalId": external_id,
        "sourceSheet": ws.title,
        "country": country,
        "lender": lender,
        "loanType": loan_type,
        "trustScore": trust_score,
        "term": term,
        "investmentDate": investment_date.isoformat() if investment_date else None,
        "startDate": start_date.isoformat() if start_date else None,
        "plannedEndDate": planned_end.isoformat() if planned_end else None,
        "actualEndDate": actual_end.isoformat() if actual_end else None,
        "rate": nominal_rate,
        "invested": invested,
        "currentValue": current_principal,
        "outstandingPrincipal": current_principal,
        "principalReturned": principal_returned,
        "interestReceived": interest_received,
        "fees": fees,
        "profit": rounded(interest_received - fees),
        "xirr": xirr,
        "status": status,
        "delayDays": actual_delay_days,
        "earlyLateDays": early_late_days,
        "schedule": schedule,
        "cashFlows": cash_flows,
    }


def merge_overview(
    investments: list[dict[str, Any]],
    overview_rows: list[dict[str, Any]],
) -> None:
    overview_by_code = {row["loanCode"]: row for row in overview_rows}
    for investment in investments:
        overview = overview_by_code.get(investment["loanCode"])
        if not overview:
            continue

        # Overview naudojamas kaip papildoma kontrolė ir informacijos šaltinis.
        for key in ("country", "loanType", "term", "investmentDate", "plannedEndDate"):
            if not investment.get(key) and overview.get(key):
                investment[key] = overview[key]

        if investment["rate"] == 0 and overview["rate"]:
            investment["rate"] = overview["rate"]

        if abs(investment["invested"]) <= EPSILON:
            investment["invested"] = overview["amount"]

        if abs(investment["principalReturned"]) <= EPSILON:
            investment["principalReturned"] = overview["principalReturned"]

        if abs(investment["interestReceived"]) <= EPSILON:
            investment["interestReceived"] = overview["interestReceived"]

        if investment["actualEndDate"] is None and overview["actualEndDate"]:
            investment["actualEndDate"] = overview["actualEndDate"]
            investment["status"] = "completed"
            investment["currentValue"] = 0.0
            investment["outstandingPrincipal"] = 0.0

        has_unpaid_overdue = any(
            item.get("plannedDate")
            and not item.get("paid")
            and date.fromisoformat(item["plannedDate"]) < date.today()
            for item in investment["schedule"]
        )

        if (
            investment["status"] != "completed"
            and has_unpaid_overdue
        ):
            investment["status"] = "delayed"
            investment["delayDays"] = max(
                investment["delayDays"],
                overview["delayDays"],
            )


def build_history(
    money_rows: list[dict[str, Any]],
    investments: list[dict[str, Any]],
    summary: dict[str, Any],
) -> list[dict[str, Any]]:
    dated_money = [row for row in money_rows if row["date"]]
    dated_flows: list[dict[str, Any]] = []

    for investment in investments:
        for flow in investment["cashFlows"]:
            if flow["date"]:
                dated_flows.append(flow)

    all_dates = [
        date.fromisoformat(row["date"]) for row in dated_money
    ] + [
        date.fromisoformat(row["date"]) for row in dated_flows
    ]

    if not all_dates:
        return []

    start = min(all_dates)
    end = max(max(all_dates), date.today())

    monthly_money: dict[tuple[int, int], dict[str, float]] = defaultdict(
        lambda: {"deposited": 0.0, "withdrawn": 0.0, "bonus": 0.0}
    )
    for row in dated_money:
        d = date.fromisoformat(row["date"])
        bucket = monthly_money[(d.year, d.month)]
        bucket["deposited"] += row["deposited"]
        bucket["withdrawn"] += row["withdrawn"]
        bucket["bonus"] += row["bonus"]

    monthly_flows: dict[tuple[int, int], float] = defaultdict(float)
    for row in dated_flows:
        d = date.fromisoformat(row["date"])
        if row["amount"] > 0:
            monthly_flows[(d.year, d.month)] += row["amount"]

    cumulative_deposited = 0.0
    cumulative_withdrawn = 0.0
    cumulative_bonus = 0.0
    cumulative_positive_flows = 0.0
    previous_value = 0.0
    history: list[dict[str, Any]] = []

    for period_end in month_range(start, end):
        key = (period_end.year, period_end.month)
        money = monthly_money[key]
        income = monthly_flows[key]

        cumulative_deposited += money["deposited"]
        cumulative_withdrawn += money["withdrawn"]
        cumulative_bonus += money["bonus"]
        cumulative_positive_flows += income

        net_invested = cumulative_deposited - cumulative_withdrawn
        # Tarpiniams mėnesiams naudojama pinigų srautų aproksimacija.
        # Galutinis taškas žemiau tiksliai sulyginamas su Overview.
        value = net_invested + cumulative_bonus
        profit = value - net_invested
        monthly_profit = money["bonus"]
        base = previous_value + money["deposited"] - money["withdrawn"]
        monthly_return = (
            monthly_profit / base * 100 if abs(base) > EPSILON else 0.0
        )

        history.append({
            "date": period_end.isoformat(),
            "invested": rounded(net_invested),
            "value": rounded(value),
            "profit": rounded(profit),
            "cash": 0.0,
            "income": rounded(income + money["bonus"]),
            "deposits": rounded(money["deposited"]),
            "withdrawals": rounded(money["withdrawn"]),
            "bonuses": rounded(money["bonus"]),
            "monthlyProfit": rounded(monthly_profit),
            "monthlyReturn": rounded(monthly_return),
        })
        previous_value = value

    if history:
        history[-1]["invested"] = summary["invested"]
        history[-1]["value"] = summary["currentValue"]
        history[-1]["profit"] = summary["profit"]
        history[-1]["cash"] = summary["cash"]

    return history


def build_distributions(investments: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    active = [item for item in investments if item["status"] != "completed"]

    def make_distribution(field: str) -> list[dict[str, Any]]:
        grouped: dict[str, dict[str, float | int]] = defaultdict(
            lambda: {"count": 0, "value": 0.0}
        )
        for item in active:
            label = str(item.get(field) or "Nenurodyta")
            grouped[label]["count"] = int(grouped[label]["count"]) + 1
            grouped[label]["value"] = float(grouped[label]["value"]) + item["currentValue"]

        return [
            {
                "label": label,
                "count": int(values["count"]),
                "value": rounded(values["value"]),
            }
            for label, values in sorted(grouped.items())
        ]

    return {
        "country": make_distribution("country"),
        "lender": make_distribution("lender"),
        "loanType": make_distribution("loanType"),
        "trustScore": make_distribution("trustScore"),
        "status": make_distribution("status"),
    }


def validate(data: dict[str, Any], expected_count: int) -> list[str]:
    errors: list[str] = []
    investments = data.get("investments", [])
    summary = data.get("summary", {})

    required_top = {
        "schemaVersion", "generatedAt", "platform", "summary",
        "history", "investments",
    }
    missing_top = required_top - set(data)
    if missing_top:
        errors.append("Trūksta viršutinių laukų: " + ", ".join(sorted(missing_top)))

    required_summary = {
        "invested", "currentValue", "profit", "returnRate", "xirr", "cash",
        "incomeReceived", "activeInvestments", "delayedInvestments",
        "completedInvestments",
    }
    missing_summary = required_summary - set(summary)
    if missing_summary:
        errors.append("summary trūksta laukų: " + ", ".join(sorted(missing_summary)))

    if len(investments) != expected_count:
        errors.append(
            f"Paskolų skaičius nesutampa: Excel {expected_count}, JSON {len(investments)}."
        )

    codes = [item["loanCode"] for item in investments]
    duplicates = [code for code, count in Counter(codes).items() if count > 1]
    if duplicates:
        errors.append("Dubliuoti paskolų kodai: " + ", ".join(sorted(duplicates)))

    active = sum(item["status"] == "active" for item in investments)
    delayed = sum(item["status"] == "delayed" for item in investments)
    completed = sum(item["status"] == "completed" for item in investments)

    if summary.get("activeInvestments") != active:
        errors.append("summary.activeInvestments nesutampa.")
    if summary.get("delayedInvestments") != delayed:
        errors.append("summary.delayedInvestments nesutampa.")
    if summary.get("completedInvestments") != completed:
        errors.append("summary.completedInvestments nesutampa.")

    total_interest = rounded(sum(item["interestReceived"] for item in investments))
    total_fees = rounded(sum(item["fees"] for item in investments))
    net_interest = rounded(total_interest - total_fees)
    expected_interest = rounded(summary.get("interestReceived"))
    if abs(net_interest - expected_interest) > 0.03:
        errors.append(
            f"Grynosios palūkanos nesutampa: paskolos {net_interest:.2f} "
            f"(bruto {total_interest:.2f} - mokesčiai {total_fees:.2f}), "
            f"Overview {expected_interest:.2f}."
        )

    total_principal = rounded(sum(item["principalReturned"] for item in investments))
    expected_principal = rounded(summary.get("principalReturned"))
    if abs(total_principal - expected_principal) > 0.03:
        errors.append(
            f"Grąžinta paskola nesutampa: paskolos {total_principal:.2f}, "
            f"Overview {expected_principal:.2f}."
        )

    expected_fees = rounded(summary.get("fees"))
    if abs(total_fees - expected_fees) > 0.03:
        errors.append(
            f"Mokesčiai nesutampa: paskolos {total_fees:.2f}, "
            f"Overview {expected_fees:.2f}."
        )

    return errors


def load_twino(excel_file: Path) -> dict[str, Any]:
    workbook = load_workbook(excel_file, data_only=True, read_only=False)
    try:
        lookup = {ws.title.strip().lower(): ws for ws in workbook.worksheets}
        if "overview" not in lookup:
            raise ValueError("Nerastas privalomas lapas 'Overview'.")
        if "pinigai" not in lookup:
            raise ValueError("Nerastas privalomas lapas 'Pinigai'.")

        today = date.today()
        overview_rows, overview = read_overview(lookup["overview"], workbook.epoch)
        money_rows = read_money_sheet(lookup["pinigai"], workbook.epoch)

        loan_sheets = [
            ws for ws in workbook.worksheets
            if ws.title.strip().lower() not in NON_LOAN_SHEETS
            and LOAN_SHEET_RE.match(ws.title.strip())
        ]
        skipped_sheets = [
            ws.title for ws in workbook.worksheets
            if ws.title.strip().lower() not in NON_LOAN_SHEETS
            and not LOAN_SHEET_RE.match(ws.title.strip())
        ]

        investments = [
            read_loan_sheet(ws, workbook.epoch, today) for ws in loan_sheets
        ]
        merge_overview(investments, overview_rows)
        investments.sort(
            key=lambda item: (
                item["investmentDate"] or "9999-12-31",
                item["loanCode"],
            )
        )

        active_count = sum(item["status"] == "active" for item in investments)
        delayed_count = sum(item["status"] == "delayed" for item in investments)
        completed_count = sum(item["status"] == "completed" for item in investments)

        current_principal = rounded(
            sum(item["currentValue"] for item in investments)
        )
        weighted_rate_denominator = sum(
            item["currentValue"] for item in investments
            if item["status"] != "completed"
        )
        average_rate = (
            rounded(
                sum(item["rate"] * item["currentValue"] for item in investments)
                / weighted_rate_denominator,
                4,
            )
            if weighted_rate_denominator > EPSILON
            else None
        )

        withdrawn = rounded(sum(row["withdrawn"] for row in money_rows))
        invested = rounded(overview["deposited"] - withdrawn)
        # Overview interestReceived yra grynosios palūkanos:
        # bruto palūkanos - mokesčiai.
        income_received = rounded(
            overview["interestReceived"] + overview["bonuses"]
        )

        summary = {
            # Portfolio V2 privalomi laukai.
            "invested": invested,
            "currentValue": overview["currentValue"],
            "profit": overview["profit"],
            "returnRate": overview["returnRate"],
            "xirr": overview["xirr"],
            "cash": overview["cash"],
            "incomeReceived": income_received,
            "activeInvestments": active_count,
            "delayedInvestments": delayed_count,
            "completedInvestments": completed_count,
            "averageRate": average_rate,
            "averageLtv": None,

            # TWINO specifiniai laukai.
            "deposited": overview["deposited"],
            "withdrawn": withdrawn,
            "lifetimeInvested": overview["lifetimeInvested"],
            "outstandingPrincipal": current_principal,
            "principalReturned": overview["principalReturned"],
            # Vienoda visos P2P grupės palūkanų struktūra.
            "interestReceived": overview["interestReceived"],
            "grossInterestReceived": rounded(
                overview["interestReceived"] + overview["fees"]
            ),
            "bonuses": overview["bonuses"],
            "fees": overview["fees"],
            "totalInvestments": len(investments),
            "roi": overview["returnRate"],
        }

        start_dates = [
            date.fromisoformat(row["date"])
            for row in money_rows if row["date"]
        ]
        start_date = min(start_dates).isoformat() if start_dates else None

        history = build_history(money_rows, investments, summary)
        updated_at = history[-1]["date"] if history else today.isoformat()
        generated_at = datetime.now().astimezone().isoformat(timespec="seconds")

        data = {
            "schemaVersion": SCHEMA_VERSION,
            "generatedAt": generated_at,
            "moduleType": "p2p_loans",
            "platform": {
                "id": "twino",
                "slug": "twino",
                "name": "TWINO",
                "group": "p2p",
                "type": "p2p_loans",
                "category": "P2P paskolos",
                "currency": "EUR",
                "active": True,
                "status": "active",
                "startDate": start_date,
                "updatedAt": updated_at,
                "website": "https://www.twino.eu",
            },
            "summary": summary,
            "history": history,
            "latestMonth": history[-1] if history else None,
            "largestInvestment": max(
                investments,
                key=lambda item: item["currentValue"],
                default=None,
            ),
            "distributions": build_distributions(investments),
            "cashFlows": money_rows,
            "investments": investments,
            "source": {
                "file": excel_file.name,
                "overviewSheet": "Overview",
                "moneySheet": "Pinigai",
                "loanSheets": len(loan_sheets),
                "skippedSheets": skipped_sheets,
                "generatedAt": generated_at,
            },
        }

        errors = validate(data, len(loan_sheets))
        if errors:
            raise ValueError("\n".join(errors))

        return data
    finally:
        workbook.close()


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")

    with temporary.open("w", encoding="utf-8", newline="\n") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.write("\n")
        file.flush()
        os.fsync(file.fileno())

    last_error: PermissionError | None = None

    for attempt in range(12):
        try:
            os.replace(temporary, path)
            return
        except PermissionError as error:
            last_error = error
            if attempt < 11:
                time.sleep(0.5)

    raise PermissionError(
        f"Nepavyko atnaujinti JSON failo: {path}. "
        "Failą gali būti trumpam užrakinęs OneDrive arba Vite."
    ) from last_error


def main() -> int:
    parser = argparse.ArgumentParser(
        description="TWINO Excel → Portfolio V2 platformos JSON."
    )
    parser.add_argument("--input", help="TWINO Excel failo kelias.")
    parser.add_argument("--output", help="TWINO JSON išvesties kelias.")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent
    excel_file = find_input(project_root, args.input)
    output_file = (
        Path(args.output).expanduser().resolve()
        if args.output
        else project_root / "public" / "data" / "platforms" / "twino.json"
    )

    print("=" * 64)
    print("TWINO IMPORTER V1")
    print("=" * 64)
    print(f"Excel failas: {excel_file}")

    data = load_twino(excel_file)
    write_json(output_file, data)

    summary = data["summary"]
    print(f"✅ Nuskaityta paskolų: {summary['totalInvestments']}")
    print(f"✅ Aktyvių: {summary['activeInvestments']}")
    print(f"✅ Vėluojančių: {summary['delayedInvestments']}")
    print(f"✅ Užbaigtų: {summary['completedInvestments']}")
    print(f"✅ Investuota: {summary['invested']:.2f} EUR")
    print(f"✅ Portfelio vertė: {summary['currentValue']:.2f} EUR")
    print(f"✅ Pelnas: {summary['profit']:.2f} EUR")
    print(f"✅ JSON sukurtas: {output_file}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"❌ KLAIDA: {error}", file=sys.stderr)
        raise SystemExit(1)
