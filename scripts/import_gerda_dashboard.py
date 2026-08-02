from __future__ import annotations
import argparse, json, math
from datetime import date, datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
DEFAULT_INPUT=ROOT/"excel"/"Investavimas Gerda.xlsx"
DEFAULT_OUTPUT=ROOT/"public"/"data"/"gerda"
SHEET="Investavimas"
SECTIONS=(("P2P","p2p_history.json","N","O","P","Q","R","S","T"),("Visas portfelis","portfolio_history.json","V","W","X","Y","Z","AA","AB"))
def num(v,percent=False):
    if v in (None,""): return 0.0
    x=float(v); x=x*100 if percent else x
    if not math.isfinite(x): raise ValueError(f"Netinkama reikšmė: {v!r}")
    return round(x,2)
def iso(v):
    if isinstance(v,datetime): return v.date().isoformat()
    if isinstance(v,date): return v.isoformat()
    return None
def payload(name,rows,source):
    return {"schemaVersion":1,"type":"monthlyPortfolioHistory","section":name,"currency":"EUR","generatedAt":datetime.now(timezone.utc).isoformat(timespec="seconds"),"source":{"file":source.name,"sheet":SHEET,"owner":"Gerda"},"period":{"from":rows[0]["date"],"to":rows[-1]["date"],"months":len(rows)},"latest":rows[-1],"history":rows}
def main():
    ap=argparse.ArgumentParser(); ap.add_argument("input",nargs="?",type=Path,default=DEFAULT_INPUT); ap.add_argument("--output-dir",type=Path,default=DEFAULT_OUTPUT); a=ap.parse_args()
    wb=load_workbook(a.input,data_only=True,read_only=True); ws=wb[SHEET]; a.output_dir.mkdir(parents=True,exist_ok=True)
    all_rows=None
    for name,file,dc,ic,mc,vc,pc,rc,resc in SECTIONS:
        rows=[]
        for r in range(3,ws.max_row+1):
            d=iso(ws[f"{dc}{r}"].value)
            if not d: continue
            item={"date":d,"invested":num(ws[f"{ic}{r}"].value),"monthlyContribution":num(ws[f"{mc}{r}"].value),"value":num(ws[f"{vc}{r}"].value),"profit":num(ws[f"{pc}{r}"].value),"returnRate":num(ws[f"{rc}{r}"].value,True),"monthlyResult":num(ws[f"{resc}{r}"].value)}
            if any(item[k] for k in ("invested","monthlyContribution","value","profit","monthlyResult")): rows.append(item)
        (a.output_dir/file).write_text(json.dumps(payload(name,rows,a.input),ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
        if name=="P2P": all_rows=rows
    zeros=[{**r,"invested":0.0,"monthlyContribution":0.0,"value":0.0,"profit":0.0,"returnRate":0.0,"monthlyResult":0.0} for r in all_rows]
    (a.output_dir/"funds_history.json").write_text(json.dumps(payload("Fondai",zeros,a.input),ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    ph={"schemaVersion":1,"type":"platform_history","currency":"EUR","generatedAt":datetime.now(timezone.utc).isoformat(timespec="seconds"),"source":{"file":a.input.name,"sheet":SHEET,"owner":"Gerda"},"platforms":{"profitus":{"name":"Profitus","history":[{"date":r["date"],"invested":r["invested"],"value":r["value"],"profit":r["profit"],"returnRate":r["returnRate"],"source":a.input.name} for r in all_rows]}}}
    (a.output_dir/"platform_history.json").write_text(json.dumps(ph,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    print(f"[OK] Gerdos istorija: {len(all_rows)} mėn.")
    return 0
if __name__=="__main__": raise SystemExit(main())
