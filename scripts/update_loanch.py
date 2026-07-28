from __future__ import annotations

import json
import math
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "excel" / "Loanch.xlsx"
PORTFOLIO_PATH = ROOT / "public" / "data" / "portfolio.json"
PREVIEW_PATH = ROOT / "public" / "data" / "loanch-platform.json"


def number(value: Any, default: float = 0.0) -> float:
    try:
        result = float(value)
        return result if math.isfinite(result) else default
    except (TypeError, ValueError):
        return default


def money(value: Any) -> float:
    return round(number(value), 2)


def excel_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


def percent(value: Any) -> float:
    result = number(value)
    return result * 100 if abs(result) <= 1 else result


def find_value(ws, label: str) -> Any:
    wanted = label.strip().lower()
    for row in ws.iter_rows(min_col=1, max_col=2):
        current = str(row[0].value or "").strip().lower()
        if current == wanted:
            return row[1].value
    return None


def read_overview(ws) -> tuple[list[dict[str, Any]], dict[str, float]]:
    loans: list[dict[str, Any]] = []

    for row in range(3, ws.max_row + 1):
        loan_id = ws.cell(row, 3).value
        if loan_id in (None, ""):
            continue

        loans.append(
            {
                "loanId": str(loan_id),
                "investedDate": ws.cell(row, 2).value,
                "country": ws.cell(row, 4).value,
                "collateral": ws.cell(row, 5).value,
                "term": ws.cell(row, 6).value,
                "interestRate": ws.cell(row, 7).value,
                "invested": ws.cell(row, 8).value,
                "plannedEndDate": ws.cell(row, 9).value,
                "actualEndDate": ws.cell(row, 10).value,
                "daysRemaining": ws.cell(row, 11).value,
                "delayDays": ws.cell(row, 12).value,
                "principalRepaid": ws.cell(row, 13).value,
                "interestReceived": ws.cell(row, 14).value,
                "fees": ws.cell(row, 15).value,
            }
        )

    summary_row = 3
    summary = {
        "deposited": money(ws.cell(summary_row, 17).value),
        "invested": money(ws.cell(summary_row, 18).value),
        "bonuses": money(ws.cell(summary_row, 19).value),
        "cash": money(ws.cell(summary_row, 20).value),
        "principalRepaid": money(ws.cell(summary_row, 21).value),
        "interestReceived": money(ws.cell(summary_row, 22).value),
        "value": money(ws.cell(summary_row, 23).value),
        "profit": money(ws.cell(summary_row, 24).value),
        "returnRate": percent(ws.cell(summary_row, 25).value),
        "xirr": percent(ws.cell(summary_row, 26).value),
        "fees": money(ws.cell(summary_row, 27).value),
    }
    return loans, summary


def read_payments(ws) -> list[dict[str, Any]]:
    payments: list[dict[str, Any]] = []

    for row in range(3, ws.max_row + 1):
        planned_raw = ws.cell(row, 4).value

        if str(planned_raw or "").strip().lower() == "viso:":
            continue

        planned = excel_date(planned_raw)
        actual = excel_date(ws.cell(row, 5).value)
        principal = money(ws.cell(row, 7).value)
        interest = money(ws.cell(row, 8).value)
        fee = money(ws.cell(row, 9).value)

        if not planned and not actual and principal == 0 and interest == 0 and fee == 0:
            continue

        delay = 0
        if planned and actual:
            delay = (actual - planned).days

        payments.append(
            {
                "plannedDate": planned.isoformat() if planned else None,
                "actualDate": actual.isoformat() if actual else None,
                "delayDays": delay,
                "principal": principal,
                "interest": interest,
                "fee": fee,
                "netIncome": round(interest - fee, 2),
                "status": "paid" if actual else "scheduled",
            }
        )

    return payments


def read_loan(ws, overview: dict[str, Any], today: date) -> dict[str, Any]:
    loan_id = str(find_value(ws, "Loan ID") or ws.title)
    country = str(find_value(ws, "Šalis") or overview.get("country") or "—")
    borrower = str(find_value(ws, "Skolintojas") or "—")
    collateral = "—"
    ltv = 0.0
    term = str(find_value(ws, "Terminas") or overview.get("term") or "—")

    invested_date = excel_date(find_value(ws, "Investuota"))
    start_date = excel_date(find_value(ws, "Pradžia"))
    planned_end = excel_date(find_value(ws, "Pabaiga") or overview.get("plannedEndDate"))
    actual_end = excel_date(overview.get("actualEndDate"))

    interest_rate = percent(find_value(ws, "Palūkanos") or overview.get("interestRate"))
    invested = money(find_value(ws, "Investuota, Eur") or overview.get("invested"))
    principal_repaid = money(overview.get("principalRepaid"))
    interest_received = money(overview.get("interestReceived"))
    fees = money(overview.get("fees"))
    remaining = max(0.0, round(invested - principal_repaid, 2))
    delay_days = int(number(overview.get("delayDays")))

    if actual_end or remaining <= 0.005:
        status = "completed"
    elif delay_days > 0:
        status = "late"
    else:
        status = "active"

    days_remaining = (
        max(0, (planned_end - today).days)
        if planned_end and status != "completed"
        else 0
    )

    return {
        "id": loan_id,
        "loanId": loan_id,
        "country": country,
        "originator": borrower,
        "borrower": borrower,
        "type": str(find_value(ws, "Tipas") or "Consumer loan"),
        "collateral": collateral,
        "ltv": round(ltv, 2),
        "term": term,
        "investedDate": invested_date.isoformat() if invested_date else None,
        "startDate": start_date.isoformat() if start_date else None,
        "plannedEndDate": planned_end.isoformat() if planned_end else None,
        "actualEndDate": actual_end.isoformat() if actual_end else None,
        "daysRemaining": days_remaining,
        "delayDays": delay_days,
        "interestRate": round(interest_rate, 2),
        "invested": invested,
        "principalRepaid": principal_repaid,
        "remainingPrincipal": remaining,
        "interestReceived": interest_received,
        "fees": fees,
        "netIncome": round(interest_received - fees, 2),
        "status": status,
        "payments": read_payments(ws),
    }


def read_cash_events(ws) -> list[dict[str, Any]]:
    events = []

    for row in range(2, ws.max_row + 1):
        event_date = excel_date(ws.cell(row, 1).value)
        deposited = money(ws.cell(row, 2).value)
        withdrawn = money(ws.cell(row, 3).value)
        bonus = money(ws.cell(row, 4).value)

        if not event_date:
            continue

        events.append(
            {
                "date": event_date,
                "deposited": deposited,
                "withdrawn": withdrawn,
                "bonus": bonus,
            }
        )

    return sorted(events, key=lambda item: item["date"])


def build_platform(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    today = date.today()

    overview_rows, summary = read_overview(workbook["Overview"])
    overview_by_id = {item["loanId"]: item for item in overview_rows}

    loans = []
    for sheet_name in workbook.sheetnames:
        if sheet_name in {"Overview", "Pinigai"}:
            continue
        # Visi kiti lapai yra atskiros Loanch paskolos.
        loans.append(
            read_loan(
                workbook[sheet_name],
                overview_by_id.get(str(sheet_name), {}),
                today,
            )
        )

    loans.sort(key=lambda item: item.get("investedDate") or "")

    active = [item for item in loans if item["status"] == "active"]
    late = [item for item in loans if item["status"] == "late"]
    completed = [item for item in loans if item["status"] == "completed"]

    deposited = summary["deposited"]
    withdrawn = 0.0
    bonuses = summary["bonuses"]
    cash = summary["cash"]
    current_value = summary["value"]
    profit = summary["profit"] or round(current_value + withdrawn - deposited, 2)
    return_rate = summary["returnRate"] or (
        profit / deposited * 100 if deposited else 0
    )

    outstanding = round(sum(item["remainingPrincipal"] for item in loans), 2)
    active_outstanding = sum(item["remainingPrincipal"] for item in active)
    average_rate = (
        sum(item["interestRate"] * item["remainingPrincipal"] for item in active)
        / active_outstanding
        if active_outstanding
        else 0
    )

    largest_share = (
        max((item["remainingPrincipal"] for item in active), default=0)
        / current_value
        * 100
        if current_value
        else 0
    )
    late_share = (
        sum(item["remainingPrincipal"] for item in late) / current_value * 100
        if current_value
        else 0
    )
    cash_share = cash / current_value * 100 if current_value else 0

    forecast = {}
    for loan in active:
        end_date = loan.get("plannedEndDate")
        if not end_date:
            continue
        month = end_date[:7]
        row = forecast.setdefault(
            month,
            {
                "month": month,
                "principal": 0.0,
                "interest": 0.0,
                "fees": 0.0,
                "total": 0.0,
            },
        )
        row["principal"] += loan["remainingPrincipal"]
        row["total"] += loan["remainingPrincipal"]

    income_forecast = []
    for row in sorted(forecast.values(), key=lambda item: item["month"]):
        income_forecast.append(
            {
                **row,
                "principal": money(row["principal"]),
                "interest": money(row["interest"]),
                "fees": money(row["fees"]),
                "total": money(row["total"]),
            }
        )

    income_history = []
    for loan in loans:
        for payment in loan["payments"]:
            if payment["status"] == "paid" and payment["actualDate"]:
                income_history.append(
                    {
                        "date": payment["actualDate"],
                        "principal": payment["principal"],
                        "interest": payment["interest"],
                        "fees": payment["fee"],
                        "netIncome": payment["netIncome"],
                    }
                )
    income_history.sort(key=lambda item: item["date"])

    cash_events = read_cash_events(workbook["Pinigai"])
    history = []
    cumulative = 0.0

    for event in cash_events:
        cumulative += event["deposited"] - event["withdrawn"]
        history.append(
            {
                "date": event["date"].isoformat(),
                "invested": money(cumulative),
                "value": money(cumulative + event["bonus"]),
            }
        )

    today_iso = today.isoformat()
    if not history or history[-1]["date"] != today_iso:
        history.append(
            {
                "date": today_iso,
                "invested": deposited,
                "value": current_value,
            }
        )
    else:
        history[-1]["value"] = current_value

    diversification_score = min(20, 5 + len(loans) * 2)
    quality_score = max(0, round(20 - late_share))
    cash_score = max(0, round(20 - cash_share * 0.6))
    repayment_score = 16 if not income_history else 19
    concentration_score = max(0, round(20 - max(0, largest_share - 20) * 0.13))
    health_score = int(
        diversification_score
        + quality_score
        + cash_score
        + repayment_score
        + concentration_score
    )

    details = {
        "type": "p2p",
        "modules": {"loans": True},
        "p2pSummary": {
            "totalLoans": len(loans),
            "activeLoans": len(active),
            "completedLoans": len(completed),
            "lateLoans": len(late),
            "averageInterestRate": round(average_rate, 2),
            "cash": cash,
            "interestReceived": money(sum(item["interestReceived"] for item in loans)),
            "principalRepaid": money(sum(item["principalRepaid"] for item in loans)),
            "outstandingPrincipal": outstanding,
        },
        "cashflow": {
            "deposited": deposited,
            "withdrawn": withdrawn,
            "bonuses": bonuses,
            "cash": cash,
            "invested": summary["invested"],
            "principalRepaid": money(sum(item["principalRepaid"] for item in loans)),
            "interestReceived": money(sum(item["interestReceived"] for item in loans)),
            "fees": money(sum(item["fees"] for item in loans)),
            "netIncome": money(
                sum(item["netIncome"] for item in loans) + bonuses
            ),
        },
        "health": {
            "score": health_score,
            "label": (
                "Puiki būklė"
                if health_score >= 85
                else "Gera būklė"
                if health_score >= 70
                else "Reikia stebėti"
            ),
            "summary": (
                "Paskolos nevėluoja, tačiau verta stebėti skolintojų koncentraciją."
                if not late
                else "Portfelyje yra vėluojančių paskolų."
            ),
            "largestLoanShare": round(largest_share, 2),
            "lateShare": round(late_share, 2),
            "cashShare": round(cash_share, 2),
            "metrics": [
                {
                    "key": "diversification",
                    "label": "Diversifikacija",
                    "score": diversification_score,
                    "max": 20,
                    "description": f"Portfelyje yra {len(loans)} paskola(-os).",
                },
                {
                    "key": "loanQuality",
                    "label": "Paskolų kokybė",
                    "score": quality_score,
                    "max": 20,
                    "description": f"Aktyvių vėluojančių paskolų: {len(late)}.",
                },
                {
                    "key": "cashEfficiency",
                    "label": "Pinigų efektyvumas",
                    "score": cash_score,
                    "max": 20,
                    "description": f"Laisvų pinigų likutis – {cash:.2f} €.",
                },
                {
                    "key": "repaymentStability",
                    "label": "Grąžinimų stabilumas",
                    "score": repayment_score,
                    "max": 20,
                    "description": (
                        "Grąžinimų istorijos dar nėra."
                        if not income_history
                        else "Vertinama pagal faktinius grąžinimus."
                    ),
                },
                {
                    "key": "concentration",
                    "label": "Koncentracijos rizika",
                    "score": concentration_score,
                    "max": 20,
                    "description": (
                        f"Didžiausia paskola sudaro {largest_share:.1f} % portfelio."
                    ),
                },
            ],
        },
        "history": income_history,
        "incomeForecast": income_forecast,
        "loans": loans,
    }

    return {
        "name": "Loanch",
        "slug": "loanch",
        "category": "P2P paskolos",
        "assetClass": "p2p",
        "currency": "EUR",
        "active": True,
        "website": "https://loanch.com",
        "invested": deposited,
        "value": current_value,
        "profit": profit,
        "returnRate": round(return_rate, 6),
        "xirr": round(summary["xirr"], 6),
        "history": history,
        "analytics": {
            "startDate": history[0]["date"] if history else "",
            "months": max(1, len({item["date"][:7] for item in history})),
            "highestValue": max(
                [number(item["value"]) for item in history],
                default=current_value,
            ),
            "averageMonthlyReturn": round(return_rate, 6),
            "maxDrawdown": 0,
            "winningRate": 100 if profit >= 0 else 0,
            "volatility": 0,
            "monthlyPerformance": [
                {
                    "date": today.strftime("%Y-%m"),
                    "previousValue": deposited,
                    "value": current_value,
                    "monthlyProfit": profit,
                    "monthlyReturn": round(return_rate, 6),
                }
            ],
        },
        "details": details,
    }


def merge_into_portfolio(platform: dict[str, Any]) -> None:
    if not PORTFOLIO_PATH.exists():
        raise FileNotFoundError(
            f"Nerastas {PORTFOLIO_PATH}. Patikrink projekto aplankų struktūrą."
        )

    portfolio = json.loads(PORTFOLIO_PATH.read_text(encoding="utf-8"))
    platforms = portfolio.setdefault("platforms", [])

    index = next(
        (
            i
            for i, item in enumerate(platforms)
            if item.get("slug") == "loanch"
            or str(item.get("name", "")).strip().lower() == "loanch"
        ),
        None,
    )

    if index is None:
        platforms.append(platform)
        action = "pridėta"
    else:
        platforms[index] = platform
        action = "atnaujinta"

    backup = PORTFOLIO_PATH.with_suffix(".before-loanch.json")
    shutil.copy2(PORTFOLIO_PATH, backup)

    PORTFOLIO_PATH.write_text(
        json.dumps(portfolio, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    PREVIEW_PATH.write_text(
        json.dumps(platform, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Loanch platforma {action}.")
    print(f"Atnaujintas failas: {PORTFOLIO_PATH}")
    print(f"Atsarginė kopija: {backup}")
    print("Atidaryk: http://localhost:5173/platforms/loanch")


def main() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Nerastas {EXCEL_PATH}. Įkelk Loanch.xlsx į projekto excel aplanką."
        )

    platform = build_platform(EXCEL_PATH)
    merge_into_portfolio(platform)


if __name__ == "__main__":
    main()
