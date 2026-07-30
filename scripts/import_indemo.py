from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


PROJECT_SHEET_RE = re.compile(r"^[A-Z]\d{3}$", re.IGNORECASE)
NON_PROJECT_SHEETS = {"overview", "pinigai"}
EPSILON = 0.005


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        text = str(value).strip().replace("\u00a0", "").replace(" ", "")
        text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0


def rounded(value: Any, digits: int = 2) -> float:
    return round(number(value), digits)


def excel_date(value: Any, epoch: datetime) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, pattern).date()
            except ValueError:
                pass
    return None


def iso(value: Any, epoch: datetime) -> str | None:
    parsed = excel_date(value, epoch)
    return parsed.isoformat() if parsed else None


def month_end(value: date) -> date:
    if value.month == 12:
        return date(value.year, 12, 31)
    first_next = date(value.year + (value.month == 12), value.month % 12 + 1, 1)
    return first_next.fromordinal(first_next.toordinal() - 1)


def month_range(start: date, end: date) -> Iterable[date]:
    current = date(start.year, start.month, 1)
    final = date(end.year, end.month, 1)
    while current <= final:
        yield month_end(current)
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)


def find_input(root: Path, explicit: str | None) -> Path:
    if explicit:
        path = Path(explicit).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Nerastas Excel failas: {path}")
        return path

    candidates = [
        root / "excel" / "Indemo.xlsx",
        root / "excel" / "Indemo(3).xlsx",
        root / "Indemo.xlsx",
    ]
    candidates.extend(sorted((root / "excel").glob("Indemo*.xlsx")) if (root / "excel").exists() else [])

    for path in candidates:
        if path.is_file():
            return path

    raise FileNotFoundError(
        "Nerastas Indemo Excel failas. Įdėk jį į excel/Indemo.xlsx "
        "arba paleisk su --input kelias\\iki\\failo.xlsx"
    )


def read_project(ws, epoch: datetime) -> dict[str, Any]:
    marker = str(ws["A1"].value or "").strip().lower()
    if marker != "mortgage loan id":
        raise ValueError(f"{ws.title}: A1 turi būti 'Mortgage loan ID'.")

    loan_code = str(ws["B1"].value or ws.title).strip()
    name = str(ws["B2"].value or "").strip()
    ptv = rounded(ws["B3"].value)
    pdt = rounded(ws["B4"].value)

    # Indemo projekto suvestinė turi du vienodai pavadintus
    # „Investuota“ laukus, todėl ją skaitome pagal stabilias eilučių pozicijas.
    weighted_investment_date = excel_date(ws["B7"].value, epoch)
    loan_start = excel_date(ws["B8"].value, epoch)
    actual_repayment = excel_date(ws["B9"].value, epoch)
    invested_total = rounded(ws["B10"].value)
    repaid_principal = rounded(ws["B11"].value)
    interest_received = rounded(ws["B12"].value)
    xirr = rounded(number(ws["B13"].value) * 100, 2)
    term_value = ws["B14"].value

    investments: list[dict[str, Any]] = []
    repayments: list[dict[str, Any]] = []

    # Investavimo lentelė: D = nr., E = Note, F = Data, G = Suma.
    for row in range(3, ws.max_row + 1):
        note = str(ws.cell(row, 5).value or "").strip()
        investment_date = excel_date(ws.cell(row, 6).value, epoch)
        amount = rounded(ws.cell(row, 7).value)
        if note or investment_date or abs(amount) > EPSILON:
            investments.append({
                "sequence": int(number(ws.cell(row, 4).value)) or len(investments) + 1,
                "note": note or None,
                "date": investment_date.isoformat() if investment_date else None,
                "amount": amount,
            })

        # Grąžinimo lentelė: H = data, I = paskola, J = palūkanos,
        # K = pelnas/papildoma suma, L = laikas.
        repayment_date = excel_date(ws.cell(row, 8).value, epoch)
        principal = rounded(ws.cell(row, 9).value)
        interest = rounded(ws.cell(row, 10).value)
        extra = rounded(ws.cell(row, 11).value)
        if repayment_date or any(abs(v) > EPSILON for v in (principal, interest, extra)):
            repayments.append({
                "date": repayment_date.isoformat() if repayment_date else None,
                "principal": principal,
                "interest": interest,
                "extra": extra,
                "days": int(number(ws.cell(row, 12).value)),
            })

    # Apsaugai perskaičiuojame sumas iš operacijų, bet Excel suvestinė lieka autoritetas.
    tx_invested = rounded(sum(item["amount"] for item in investments))
    tx_principal = rounded(sum(item["principal"] for item in repayments))
    tx_interest = rounded(sum(item["interest"] + item["extra"] for item in repayments))

    if invested_total <= EPSILON:
        invested_total = tx_invested
    if repaid_principal <= EPSILON and tx_principal > EPSILON:
        repaid_principal = tx_principal
    if interest_received <= EPSILON and tx_interest > EPSILON:
        interest_received = tx_interest

    completed = actual_repayment is not None
    outstanding = 0.0 if completed else rounded(max(invested_total - repaid_principal, 0))

    duration_days = None
    if loan_start and actual_repayment:
        duration_days = (actual_repayment - loan_start).days
    elif number(term_value):
        duration_days = int(number(term_value))

    return {
        "id": loan_code,
        "slug": loan_code.lower(),
        "loanCode": loan_code,
        "sourceSheet": ws.title,
        "name": name,
        "loanGroup": loan_code[:1].upper(),
        "ptv": ptv,
        "pdt": pdt,
        "investmentDate": weighted_investment_date.isoformat() if weighted_investment_date else (
            investments[0]["date"] if investments else None
        ),
        "loanStartDate": loan_start.isoformat() if loan_start else None,
        "actualRepayment": actual_repayment.isoformat() if actual_repayment else None,
        "durationDays": duration_days,
        "invested": invested_total,
        "repaidPrincipal": repaid_principal,
        "outstanding": outstanding,
        "interestReceived": interest_received,
        "extraReceived": rounded(sum(item["extra"] for item in repayments)),
        "xirr": xirr,
        "status": "repaid" if completed else "active",
        "delayDays": 0,
        "investments": investments,
        "repayments": repayments,
    }


def read_money_sheet(ws, epoch: datetime) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        tx_date = excel_date(ws.cell(row, 1).value, epoch)
        deposited = rounded(ws.cell(row, 2).value)
        withdrawn = rounded(ws.cell(row, 3).value)
        bonuses = rounded(ws.cell(row, 4).value)
        if not tx_date and all(abs(v) <= EPSILON for v in (deposited, withdrawn, bonuses)):
            continue
        rows.append({
            "date": tx_date.isoformat() if tx_date else None,
            "deposited": deposited,
            "withdrawn": withdrawn,
            "bonuses": bonuses,
        })
    return rows


def read_overview_summary(ws) -> dict[str, float]:
    # Overview K:T suvestinė yra tiesioginis Excel autoritetas.
    return {
        "deposited": rounded(ws["K3"].value),
        "totalInvested": rounded(ws["L3"].value),
        "bonuses": rounded(ws["M3"].value),
        "cash": rounded(ws["N3"].value),
        "repaidPrincipal": rounded(ws["O3"].value),
        "interestReceived": rounded(ws["P3"].value),
        "portfolioValue": rounded(ws["Q3"].value),
        "profit": rounded(ws["R3"].value),
        "roi": rounded(number(ws["S3"].value) * 100, 2),
        "xirr": rounded(number(ws["T3"].value) * 100, 2),
    }


def build_history(
    money_rows: list[dict[str, Any]],
    projects: list[dict[str, Any]],
    final_summary: dict[str, float],
) -> list[dict[str, Any]]:
    dated_money = [row for row in money_rows if row["date"]]
    repayment_events: list[dict[str, Any]] = []
    for project in projects:
        for repayment in project["repayments"]:
            if repayment["date"]:
                repayment_events.append(repayment)

    all_dates = [
        date.fromisoformat(row["date"]) for row in dated_money
    ] + [
        date.fromisoformat(row["date"]) for row in repayment_events
    ]

    if not all_dates:
        return []

    start = min(all_dates)
    end = max(max(all_dates), date.today())

    monthly_money: dict[tuple[int, int], dict[str, float]] = defaultdict(
        lambda: {"deposited": 0.0, "withdrawn": 0.0, "bonuses": 0.0}
    )
    for row in dated_money:
        d = date.fromisoformat(row["date"])
        bucket = monthly_money[(d.year, d.month)]
        for key in ("deposited", "withdrawn", "bonuses"):
            bucket[key] += row[key]

    monthly_income: dict[tuple[int, int], dict[str, float]] = defaultdict(
        lambda: {"principal": 0.0, "interest": 0.0, "extra": 0.0}
    )
    for row in repayment_events:
        d = date.fromisoformat(row["date"])
        bucket = monthly_income[(d.year, d.month)]
        for key in ("principal", "interest", "extra"):
            bucket[key] += row[key]

    cumulative_deposited = 0.0
    cumulative_withdrawn = 0.0
    cumulative_bonuses = 0.0
    cumulative_interest = 0.0
    previous_value = 0.0
    history: list[dict[str, Any]] = []

    for period_end in month_range(start, end):
        key = (period_end.year, period_end.month)
        cash = monthly_money[key]
        income = monthly_income[key]

        cumulative_deposited += cash["deposited"]
        cumulative_withdrawn += cash["withdrawn"]
        cumulative_bonuses += cash["bonuses"]
        monthly_interest = income["interest"] + income["extra"]
        cumulative_interest += monthly_interest

        net_contributions = cumulative_deposited - cumulative_withdrawn
        value = net_contributions + cumulative_bonuses + cumulative_interest
        profit = value - net_contributions
        monthly_profit = cash["bonuses"] + monthly_interest
        cash_flow = cash["deposited"] - cash["withdrawn"]
        base = previous_value + cash_flow
        monthly_return = (monthly_profit / base * 100) if abs(base) > EPSILON else 0.0

        history.append({
            "date": period_end.isoformat(),
            "invested": rounded(net_contributions),
            "value": rounded(value),
            "profit": rounded(profit),
            "cash": 0.0,
            "income": rounded(monthly_interest),
            "deposits": rounded(cash["deposited"]),
            "withdrawals": rounded(cash["withdrawn"]),
            "bonuses": rounded(cash["bonuses"]),
            "principalRepaid": rounded(income["principal"]),
            "interest": rounded(income["interest"]),
            "extraIncome": rounded(income["extra"]),
            "monthlyProfit": rounded(monthly_profit),
            "monthlyReturn": rounded(monthly_return),
        })
        previous_value = value

    # Paskutinis taškas turi tiksliai sutapti su Excel Overview suvestine.
    if history:
        history[-1]["invested"] = rounded(
            final_summary["deposited"] - final_summary.get("withdrawn", 0)
        )
        history[-1]["value"] = final_summary["currentValue"]
        history[-1]["profit"] = final_summary["profit"]
        history[-1]["cash"] = final_summary["cash"]

    return history


def build_distributions(projects: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    active = [p for p in projects if p["status"] != "repaid"]

    groups: dict[str, list[dict[str, Any]]] = {}
    for key, getter in {
        "loanGroup": lambda p: p["loanGroup"],
        "ptv": lambda p: (
            "≤50 %" if p["ptv"] <= 0.50 else
            "51–60 %" if p["ptv"] <= 0.60 else
            "61–70 %" if p["ptv"] <= 0.70 else
            ">70 %"
        ),
        "pdt": lambda p: (
            "≤40 %" if p["pdt"] <= 0.40 else
            "41–50 %" if p["pdt"] <= 0.50 else
            "51–60 %" if p["pdt"] <= 0.60 else
            ">60 %"
        ),
    }.items():
        buckets: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "value": 0.0})
        for project in active:
            label = getter(project)
            buckets[label]["count"] += 1
            buckets[label]["value"] += project["outstanding"]
        groups[key] = [
            {
                "label": label,
                "count": values["count"],
                "value": rounded(values["value"]),
                "display": f'{rounded(values["value"]):.2f} €',
            }
            for label, values in sorted(buckets.items())
        ]
    return groups


def validate(data: dict[str, Any], expected_projects: int) -> list[str]:
    errors: list[str] = []
    projects = data.get("investments", [])
    summary = data.get("summary", {})

    if len(projects) != expected_projects:
        errors.append(
            f"Projektų skaičius nesutampa: Excel {expected_projects}, JSON {len(projects)}."
        )

    codes = [str(p.get("loanCode") or "") for p in projects]
    duplicates = [code for code, count in Counter(codes).items() if code and count > 1]
    if duplicates:
        errors.append("Dubliuoti projektų kodai: " + ", ".join(sorted(duplicates)))

    allowed = {"active", "repaid"}
    unknown = sorted({str(p.get("status")) for p in projects} - allowed)
    if unknown:
        errors.append("Neatpažinti statusai: " + ", ".join(unknown))

    active = sum(p["status"] == "active" for p in projects)
    repaid = sum(p["status"] == "repaid" for p in projects)
    if summary.get("activeInvestments") != active:
        errors.append("summary.activeInvestments nesutampa su projektais.")
    if summary.get("completedInvestments") != repaid:
        errors.append("summary.completedInvestments nesutampa su projektais.")

    project_outstanding = rounded(sum(p["outstanding"] for p in projects))
    if abs(project_outstanding - number(summary.get("totalInvested"))) > 0.03:
        errors.append(
            f"Aktyvus paskolų likutis nesutampa: projektai {project_outstanding:.2f}, "
            f"Overview {number(summary.get('totalInvested')):.2f}."
        )

    project_principal = rounded(sum(p["repaidPrincipal"] for p in projects))
    if abs(project_principal - number(summary.get("repaidPrincipal"))) > 0.03:
        errors.append(
            f"Grąžinta paskola nesutampa: projektai {project_principal:.2f}, "
            f"Overview {number(summary.get('repaidPrincipal')):.2f}."
        )

    project_interest = rounded(sum(p["interestReceived"] for p in projects))
    if abs(project_interest - number(summary.get("incomeReceived"))) > 0.03:
        errors.append(
            f"Palūkanos nesutampa: projektai {project_interest:.2f}, "
            f"Overview {number(summary.get('incomeReceived')):.2f}."
        )

    return errors


def load_indemo(excel_file: Path) -> dict[str, Any]:
    workbook = load_workbook(excel_file, data_only=True, read_only=False)
    try:
        sheet_lookup = {sheet.title.strip().lower(): sheet for sheet in workbook.worksheets}
        if "overview" not in sheet_lookup:
            raise ValueError("Nerastas privalomas lapas 'Overview'.")
        if "pinigai" not in sheet_lookup:
            raise ValueError("Nerastas privalomas lapas 'Pinigai'.")

        project_sheets = [
            sheet for sheet in workbook.worksheets
            if sheet.title.strip().lower() not in NON_PROJECT_SHEETS
            and PROJECT_SHEET_RE.match(sheet.title.strip())
        ]
        skipped = [
            sheet.title for sheet in workbook.worksheets
            if sheet.title.strip().lower() not in NON_PROJECT_SHEETS
            and not PROJECT_SHEET_RE.match(sheet.title.strip())
        ]

        projects = [read_project(sheet, workbook.epoch) for sheet in project_sheets]
        projects.sort(key=lambda item: item["loanCode"])

        money_rows = read_money_sheet(sheet_lookup["pinigai"], workbook.epoch)
        overview = read_overview_summary(sheet_lookup["overview"])
        overview["withdrawn"] = rounded(sum(row["withdrawn"] for row in money_rows))

        active_count = sum(project["status"] == "active" for project in projects)
        repaid_count = sum(project["status"] == "repaid" for project in projects)
        outstanding = rounded(sum(project["outstanding"] for project in projects))

        start_dates = [
            date.fromisoformat(row["date"])
            for row in money_rows if row["date"]
        ]
        start_date = min(start_dates).isoformat() if start_dates else None

        average_ptv = rounded(
            sum(p["ptv"] * p["outstanding"] for p in projects) / outstanding
            if outstanding > EPSILON else 0,
            4,
        )
        average_pdt = rounded(
            sum(p["pdt"] * p["outstanding"] for p in projects) / outstanding
            if outstanding > EPSILON else 0,
            4,
        )

        summary = {
            # Privalomi Portfolio V2 laukai.
            "invested": overview["totalInvested"],
            "currentValue": overview["portfolioValue"],
            "profit": overview["profit"],
            "returnRate": overview["roi"],
            "xirr": overview["xirr"],
            "cash": overview["cash"],
            "incomeReceived": overview["interestReceived"],
            "activeInvestments": active_count,
            "delayedInvestments": 0,
            "completedInvestments": repaid_count,
            "averageRate": None,
            "averageLtv": average_ptv,

            # Papildomi Indemo puslapiui reikalingi laukai.
            "deposited": overview["deposited"],
            "withdrawn": overview["withdrawn"],
            "totalInvested": overview["totalInvested"],
            "lifetimeInvested": rounded(sum(project["invested"] for project in projects)),
            "outstanding": outstanding,
            "bonuses": overview["bonuses"],
            "repaidPrincipal": overview["repaidPrincipal"],
            "interestReceived": overview["interestReceived"],
            "extraReceived": rounded(sum(p["extraReceived"] for p in projects)),
            "portfolioValue": overview["portfolioValue"],
            "roi": overview["roi"],
            "totalProjects": len(projects),
            "activeProjects": active_count,
            "delayedProjects": 0,
            "repaidProjects": repaid_count,
            "completedProjects": repaid_count,
            "averagePtv": average_ptv,
            "averagePdt": average_pdt,
        }

        history = build_history(money_rows, projects, summary)
        updated_at = history[-1]["date"] if history else date.today().isoformat()

        generated_at = datetime.now().astimezone().isoformat(timespec="seconds")

        data = {
            "schemaVersion": 1,
            "generatedAt": generated_at,
            "moduleType": "indemo",
            "platform": {
                "id": "indemo",
                "name": "Indemo",
                "slug": "indemo",
                "group": "p2p",
                "type": "indemo",
                "category": "Hipoteka užtikrintos paskolos",
                "currency": "EUR",
                "active": True,
                "website": "https://indemo.eu",
                "status": "active",
                "startDate": start_date,
                "updatedAt": updated_at,
                "healthScore": 100,
                "healthParts": {
                    "diversification": min(100, round(active_count / 50 * 100)),
                    "delays": 100,
                    "cash": 100 if overview["portfolioValue"] <= EPSILON else round(
                        min(100, overview["cash"] / overview["portfolioValue"] * 100 * 5)
                    ),
                },
            },
            "summary": summary,
            "history": history,
            "latestMonth": history[-1] if history else None,
            "largestProject": max(projects, key=lambda item: item["outstanding"], default=None),
            "distributions": build_distributions(projects),
            "cashFlows": money_rows,
            "investments": projects,
            "projects": projects,
            "source": {
                "file": excel_file.name,
                "overviewSheet": "Overview",
                "moneySheet": "Pinigai",
                "projectSheets": len(project_sheets),
                "skippedSheets": skipped,
                "generatedAt": generated_at,
            },
        }

        errors = validate(data, len(project_sheets))
        if errors:
            raise ValueError("\n".join(errors))

        return data
    finally:
        workbook.close()


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=2)
    temporary.replace(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Indemo Excel → platform JSON importeris")
    parser.add_argument("--input", help="Indemo Excel failo kelias")
    parser.add_argument("--output", help="JSON išvesties kelias")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    root = script_dir.parent
    excel_file = find_input(root, args.input)
    output_file = (
        Path(args.output).expanduser().resolve()
        if args.output
        else root / "public" / "data" / "platforms" / "indemo.json"
    )

    print("=" * 64)
    print("INDEMO IMPORTER V1")
    print("=" * 64)
    print(f"Excel failas: {excel_file}")

    data = load_indemo(excel_file)
    write_json(output_file, data)

    summary = data["summary"]
    print(f"✅ Nuskaityta projektų: {summary['totalProjects']}")
    print(f"✅ Aktyvių: {summary['activeInvestments']}")
    print(f"✅ Grąžintų: {summary['completedInvestments']}")
    print(f"✅ Investuota į projektus: {summary['invested']:.2f} EUR")
    print(f"✅ Aktyvus paskolų likutis: {summary['outstanding']:.2f} EUR")
    print(f"✅ Portfelio vertė: {summary['currentValue']:.2f} EUR")
    print(f"✅ Pelnas: {summary['profit']:.2f} EUR")
    print(f"✅ JSON sukurtas: {output_file}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"❌ KLAIDA: {error}", file=sys.stderr)
        raise SystemExit(1)
