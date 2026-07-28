import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from importers.seb_mikro import load_seb_mikro_holdings
from importers.revolut_brokerage import load_revolut_brokerage
from importers.afranga import load_afranga
from importers.debitum import load_debitum
from importers.indemo import load_indemo
from importers.base_importer import normalize_name
from analytics import (
    calculate_platform_analytics,
    calculate_portfolio_analytics,
)


ROOT = Path(__file__).resolve().parent.parent

EXCEL_FILE = ROOT / "excel" / "Investavimas.xlsx"
SEB_MIKRO_FILE = ROOT / "excel" / "Mikro VPLT038884.xlsx"
REVOLUT_BROKERAGE_FILE = ROOT / "excel" / "Revolut Brokerage.xlsx"
AFRANGA_FILE = ROOT / "excel" / "Afranga.xlsx"
DEBITUM_FILE = ROOT / "excel" / "Debitum.xlsx"
INDEMO_FILE = ROOT / "excel" / "Indemo.xlsx"
OUTPUT_FILE = ROOT / "public" / "data" / "portfolio.json"
SHEET_NAME = "Investavimas"

FIRST_DATA_ROW = 3
FIRST_PLATFORM_COLUMN = 2
PLATFORM_BLOCK_SIZE = 3


PLATFORM_METADATA = {
    "seb fondai": {
        "assetClass": "fund",
        "category": "Investiciniai fondai",
        "domain": "seb.lt",
    },
    "seb mikro": {
        "assetClass": "fund",
        "category": "Investiciniai fondai",
        "domain": "seb.lt",
    },
    "seb robo": {
        "assetClass": "robo",
        "category": "Robo Advisor",
        "domain": "seb.lt",
    },
    "revolut brokerage": {
        "assetClass": "broker",
        "category": "Akcijos ir ETF",
        "domain": "revolut.com",
    },
    "revolut robo": {
        "assetClass": "robo",
        "category": "Robo Advisor",
        "domain": "revolut.com",
    },
    "lightyear": {
        "assetClass": "broker",
        "category": "Akcijos ir ETF",
        "domain": "lightyear.com",
    },
    "synergy": {
        "assetClass": "fund",
        "category": "Investiciniai fondai",
        "domain": "synergy-finance.com",
    },
    "profitus": {
        "assetClass": "real_estate",
        "category": "NT sutelktinis finansavimas",
        "domain": "profitus.lt",
    },
    "nordstreet": {
        "assetClass": "real_estate",
        "category": "NT sutelktinis finansavimas",
        "domain": "nordstreet.com",
    },
    "crowdpear": {
        "assetClass": "real_estate",
        "category": "NT sutelktinis finansavimas",
        "domain": "crowdpear.com",
    },
    "estateguru": {
        "assetClass": "real_estate",
        "category": "NT sutelktinis finansavimas",
        "domain": "estateguru.co",
    },
    "indemo": {
        "assetClass": "npl",
        "category": "NPL investicijos",
        "domain": "indemo.eu",
    },
    "peerberry": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "peerberry.com",
    },
    "fintown": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "fintown.eu",
    },
    "income": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "getincome.com",
    },
    "mintos": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "mintos.com",
    },
    "twino": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "twino.eu",
    },
    "esketit": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "esketit.com",
    },
    "robocash": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "robo.cash",
    },
    "lendermarket": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "lendermarket.com",
    },
    "loanch": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "loanch.com",
    },
    "nectaro": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "nectaro.eu",
    },
    "debitum": {
        "assetClass": "private_credit",
        "category": "Verslo finansavimas",
        "domain": "debituminvestments.com",
    },
    "viainvest": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "viainvest.com",
    },
    "rontgen": {
        "assetClass": "real_estate",
        "category": "NT sutelktinis finansavimas",
        "domain": "rontgen.lt",
    },
    "afranga": {
        "assetClass": "p2p",
        "category": "P2P paskolos",
        "domain": "afranga.com",
    },
    "lande": {
        "assetClass": "p2p",
        "category": "Žemės ūkio paskolos",
        "domain": "lande.finance",
    },
    "scramble": {
        "assetClass": "private_credit",
        "category": "Verslo finansavimas",
        "domain": "scrambleup.com",
    },
}



def safe_number(value):
    """Paverčia Excel reikšmę į skaičių arba grąžina 0."""
    if value is None:
        return 0.0

    if isinstance(value, bool):
        return float(value)

    if isinstance(value, (int, float)):
        return float(value)

    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def format_date(value):
    """Paverčia Excel datą į YYYY-MM-DD formatą."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    return str(value) if value is not None else ""


def normalize_platform_name(value):
    """Normalizuoja platformos pavadinimą palyginimams."""
    return " ".join(str(value or "").strip().lower().split())


def create_slug(value):
    """Sukuria URL tinkamą platformos identifikatorių."""
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.lower().strip()
    ascii_text = re.sub(r"[^a-z0-9]+", "-", ascii_text)
    return ascii_text.strip("-")


def get_platform_metadata(platform_name):
    """
    Grąžina platformos turto klasę, kategoriją, logotipo URL ir svetainę.

    Jei platforma nėra žinoma, paliekama neutrali turto klasė ir kategorija.
    """
    normalized_name = normalize_platform_name(platform_name)

    metadata = PLATFORM_METADATA.get(
        normalized_name,
        {
            "assetClass": "other",
            "category": "Kita investicija",
            "domain": "",
        },
    )

    domain = metadata.get("domain", "")

    return {
        "assetClass": metadata.get("assetClass", "other"),
        "category": metadata.get("category", "Kita investicija"),
        "website": f"https://{domain}" if domain else "",
        "logoUrl": (
            f"https://www.google.com/s2/favicons?domain={domain}&sz=128"
            if domain
            else ""
        ),
    }


def find_last_data_row(ws):
    """Randa paskutinę eilutę, kur A stulpelyje yra data."""
    for row_number in range(ws.max_row, FIRST_DATA_ROW - 1, -1):
        if ws.cell(row=row_number, column=1).value is not None:
            return row_number

    return None


def is_column_hidden(ws, column_number):
    """
    Patikrina, ar stulpelis paslėptas.

    Veikia ir tada, kai Excel stulpeliai paslėpti arba sugrupuoti
    kaip vienas stulpelių intervalas.
    """
    column_letter = get_column_letter(column_number)

    if ws.column_dimensions[column_letter].hidden:
        return True

    for dimension in ws.column_dimensions.values():
        min_col = getattr(dimension, "min", None)
        max_col = getattr(dimension, "max", None)

        if (
            dimension.hidden
            and min_col is not None
            and max_col is not None
            and min_col <= column_number <= max_col
        ):
            return True

    return False


def is_platform_block_hidden(ws, start_col):
    """
    Platforma laikoma neaktyvia, kai paslėpti visi trys jos stulpeliai:
    investuota, vertė ir pelningumas.
    """
    return all(
        is_column_hidden(ws, column_number)
        for column_number in range(
            start_col,
            start_col + PLATFORM_BLOCK_SIZE,
        )
    )


def trim_leading_zero_history(history):
    """
    Pašalina laikotarpius iki pirmos realios investicijos.

    Vėlesnės nulinės reikšmės paliekamos, nes jos gali reikšti,
    kad platforma buvo uždaryta arba lėšos išimtos.
    """
    first_non_zero_index = None

    for index, point in enumerate(history):
        if point["invested"] != 0 or point["value"] != 0:
            first_non_zero_index = index
            break

    if first_non_zero_index is None:
        return []

    return history[first_non_zero_index:]


def build_platform_history(ws, start_col, last_row):
    """
    Sukuria vienos platformos mėnesinę istoriją.

    A stulpelis = data
    start_col = investuota
    start_col + 1 = vertė
    start_col + 2 = pelningumas
    """
    history = []

    for row_number in range(FIRST_DATA_ROW, last_row + 1):
        history_date = ws.cell(row=row_number, column=1).value

        if history_date is None:
            continue

        invested = safe_number(
            ws.cell(row=row_number, column=start_col).value
        )
        value = safe_number(
            ws.cell(row=row_number, column=start_col + 1).value
        )
        return_rate = safe_number(
            ws.cell(row=row_number, column=start_col + 2).value
        )

        history.append(
            {
                "date": format_date(history_date),
                "invested": round(invested, 2),
                "value": round(value, 2),
                "profit": round(value - invested, 2),
                "returnRate": round(return_rate * 100, 2),
            }
        )

    return trim_leading_zero_history(history)


def normalize_header(value):
    """Normalizuoja Excel antraštę platformų blokų aptikimui."""
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return " ".join(ascii_text.strip().lower().split())


def is_platform_block(ws, start_col):
    """
    Patikrina, ar nuo start_col prasideda platformos 3 stulpelių blokas.

    Tikimasi tokios 2 eilutės struktūros:
    - Įnešta arba Investuota;
    - Vertė;
    - %.

    Dėl šio tikrinimo portfelio suvestinės blokai (Fondai, P2P, Viso)
    nėra palaikomi platformomis.
    """
    first_header = normalize_header(ws.cell(row=2, column=start_col).value)
    second_header = normalize_header(ws.cell(row=2, column=start_col + 1).value)
    third_header = normalize_header(ws.cell(row=2, column=start_col + 2).value)

    valid_first_headers = {"inesta", "investuota"}
    valid_second_headers = {"verte"}
    valid_third_headers = {"%", "pelningumas", "graza"}

    return (
        bool(ws.cell(row=1, column=start_col).value)
        and first_header in valid_first_headers
        and second_header in valid_second_headers
        and third_header in valid_third_headers
    )


def find_platform_start_columns(ws):
    """
    Automatiškai suranda visų platformų pradžios stulpelius.

    Paieška vykdoma per visą naudojamą lapo plotį, todėl pridėjus naują
    platformą į Excel failą nebereikia keisti Python konstantų.
    """
    platform_columns = []
    column_number = FIRST_PLATFORM_COLUMN

    while column_number <= ws.max_column - (PLATFORM_BLOCK_SIZE - 1):
        if is_platform_block(ws, column_number):
            platform_columns.append(column_number)
            column_number += PLATFORM_BLOCK_SIZE
        else:
            column_number += 1

    return platform_columns


def build_platforms(ws, row_number, last_row):
    """
    Nuskaito visas platformas iš trijų stulpelių blokų.

    Kiekvienai platformai sugeneruojama:
    - slug;
    - turto klasė;
    - kategorija;
    - logotipo URL;
    - svetainė;
    - paskutinės reikšmės;
    - visa mėnesinė istorija.
    """
    platforms = []

    platform_start_columns = find_platform_start_columns(ws)

    for start_col in platform_start_columns:
        platform_name = ws.cell(row=1, column=start_col).value

        if not platform_name:
            continue

        platform_name = str(platform_name).strip()

        invested = safe_number(
            ws.cell(row=row_number, column=start_col).value
        )
        value = safe_number(
            ws.cell(row=row_number, column=start_col + 1).value
        )
        return_rate = safe_number(
            ws.cell(row=row_number, column=start_col + 2).value
        )

        metadata = get_platform_metadata(platform_name)
        history = build_platform_history(
            ws,
            start_col,
            last_row,
        )
        block_hidden = is_platform_block_hidden(ws, start_col)

        platforms.append(
            {
                "name": platform_name,
                "slug": create_slug(platform_name),
                "assetClass": metadata["assetClass"],
                "category": metadata["category"],
                "logoUrl": metadata["logoUrl"],
                "website": metadata["website"],
                "currency": "EUR",
                "invested": round(invested, 2),
                "value": round(value, 2),
                "profit": round(value - invested, 2),
                "returnRate": round(return_rate * 100, 2),
                "active": not block_hidden,
                "history": history,
                "analytics": calculate_platform_analytics(history),
            }
        )

    return platforms


def find_summary_block(ws, block_name):
    """
    Suranda suvestinės bloko pradžios stulpelį pagal 1 eilutės pavadinimą.

    Tikimasi tokios struktūros:
    Data | Įnešta | Per mėn. | Vertė | Prieaugis | % | Per mėn.
    """
    target = normalize_platform_name(block_name)

    for column_number in range(1, ws.max_column + 1):
        header = normalize_platform_name(
            ws.cell(row=1, column=column_number).value
        )

        if header == target:
            return column_number

    raise ValueError(
        f"Nerastas suvestinės blokas '{block_name}' Excel 1 eilutėje."
    )


def get_summary_columns(ws, block_name):
    """Grąžina konkretaus suvestinės bloko stulpelių numerius."""
    start_col = find_summary_block(ws, block_name)

    return {
        "date": start_col,
        "invested": start_col + 1,
        "monthly_invested": start_col + 2,
        "value": start_col + 3,
        "profit": start_col + 4,
        "return_rate": start_col + 5,
        "monthly_change": start_col + 6,
    }


def build_history(ws, last_row, total_columns):
    """Sukuria viso portfelio istoriją iš automatiškai aptikto „Viso“ bloko."""
    history = []

    for row_number in range(FIRST_DATA_ROW, last_row + 1):
        history_date = ws.cell(
            row=row_number,
            column=total_columns["date"],
        ).value

        if history_date is None:
            continue

        portfolio_value = safe_number(
            ws.cell(
                row=row_number,
                column=total_columns["value"],
            ).value
        )
        invested = safe_number(
            ws.cell(
                row=row_number,
                column=total_columns["invested"],
            ).value
        )

        history.append(
            {
                "date": format_date(history_date),
                "value": round(portfolio_value, 2),
                "invested": round(invested, 2),
                "profit": round(portfolio_value - invested, 2),
            }
        )

    return history




def validate_imported_platform(platform_name, imported):
    """Patikrina, ar atskiro platformos failo duomenys tikrai prijungti."""
    history = imported.get("history") or []
    summary = imported.get("summary") or {}

    if not history:
        raise ValueError(
            f"{platform_name}: importeris negrąžino istorijos."
        )

    latest = history[-1]
    latest_date = str(latest.get("date") or "")

    if not latest_date:
        raise ValueError(
            f"{platform_name}: paskutinis istorijos įrašas neturi datos."
        )

    for field in ("invested", "value", "profit", "returnRate"):
        if field not in summary:
            raise ValueError(
                f"{platform_name}: summary trūksta lauko '{field}'."
            )

    return latest_date


def apply_platform_import(
    platforms,
    platform_name,
    imported,
):
    """
    Prijungia bet kurio importeriaus rezultatą prie platformos.

    Platformos profilis, istorija ir analitika visada atnaujinami
    iš atskiro platformos failo. Pagrindinis Investavimas.xlsx
    lieka bendro portfelio suvestinės šaltinis.
    """
    target_name = normalize_name(platform_name)

    platform = next(
        (
            item
            for item in platforms
            if normalize_name(item.get("name")) == target_name
        ),
        None,
    )

    if platform is None:
        print(
            "⚠️ Pagrindiniame faile nerasta platforma: "
            f"{platform_name}"
        )
        return False

    summary = imported.get("summary", {})
    history = imported.get("history", [])

    platform["invested"] = round(
        safe_number(summary.get("invested")),
        2,
    )
    platform["value"] = round(
        safe_number(summary.get("value")),
        2,
    )
    platform["profit"] = round(
        safe_number(summary.get("profit")),
        2,
    )
    platform["returnRate"] = round(
        safe_number(summary.get("returnRate")),
        2,
    )
    platform["history"] = history
    platform["analytics"] = calculate_platform_analytics(history)
    platform["details"] = imported

    return True


def main():
    print("=" * 56)
    print("PORTFOLIO IMPORTER V8")
    print("=" * 56)

    print(f"Excel failas: {EXCEL_FILE}")

    if not EXCEL_FILE.exists():
        print("❌ Nerastas pagrindinis Excel failas!")
        raise SystemExit(1)

    print("✅ Pagrindinis Excel failas rastas")

    if not SEB_MIKRO_FILE.exists():
        print(f"❌ Nerastas SEB Mikro failas: {SEB_MIKRO_FILE}")
        raise SystemExit(1)

    print("✅ SEB Mikro Excel failas rastas")

    if not REVOLUT_BROKERAGE_FILE.exists():
        print(
            "❌ Nerastas Revolut Brokerage failas: "
            f"{REVOLUT_BROKERAGE_FILE}"
        )
        raise SystemExit(1)

    print("✅ Revolut Brokerage Excel failas rastas")

    workbook = load_workbook(
        EXCEL_FILE,
        data_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        print(f"❌ Nerastas lapas '{SHEET_NAME}'")
        raise SystemExit(1)

    ws = workbook[SHEET_NAME]

    last_row = find_last_data_row(ws)

    if last_row is None:
        print("❌ Nerasta nė viena duomenų eilutė")
        raise SystemExit(1)

    funds_columns = get_summary_columns(ws, "Fondai")
    p2p_columns = get_summary_columns(ws, "P2P")
    total_columns = get_summary_columns(ws, "Viso")

    portfolio_date = ws.cell(
        row=last_row,
        column=total_columns["date"],
    ).value
    if portfolio_date is None:
        portfolio_date = ws.cell(row=last_row, column=1).value

    invested = safe_number(
        ws.cell(
            row=last_row,
            column=total_columns["invested"],
        ).value
    )
    portfolio_value = safe_number(
        ws.cell(
            row=last_row,
            column=total_columns["value"],
        ).value
    )
    profit = safe_number(
        ws.cell(
            row=last_row,
            column=total_columns["profit"],
        ).value
    )
    return_rate = safe_number(
        ws.cell(
            row=last_row,
            column=total_columns["return_rate"],
        ).value
    )

    funds_value = safe_number(
        ws.cell(
            row=last_row,
            column=funds_columns["value"],
        ).value
    )
    p2p_value = safe_number(
        ws.cell(
            row=last_row,
            column=p2p_columns["value"],
        ).value
    )

    platform_start_columns = find_platform_start_columns(ws)
    platforms = build_platforms(
        ws,
        last_row,
        last_row,
    )
    print(f"✅ Automatiškai aptikta platformų: {len(platform_start_columns)}")
    history = build_history(ws, last_row, total_columns)


    importer_registry = [
        {
            "platformName": "SEB Mikro",
            "file": SEB_MIKRO_FILE,
            "loader": load_seb_mikro_holdings,
        },
        {
            "platformName": "Revolut Brokerage",
            "file": REVOLUT_BROKERAGE_FILE,
            "loader": load_revolut_brokerage,
        },
        {
            "platformName": "Afranga",
            "file": AFRANGA_FILE,
            "loader": load_afranga,
        },
        {
            "platformName": "Debitum",
            "file": DEBITUM_FILE,
            "loader": load_debitum,
        },
        {
            "platformName": "Indemo",
            "file": INDEMO_FILE,
            "loader": load_indemo,
        },
    ]

    imported_platforms = {}

    for importer in importer_registry:
        platform_name = importer["platformName"]
        source_file = importer["file"]
        loader = importer["loader"]

        print(f"⏳ Skaitomi {platform_name} duomenys...")
        imported = loader(source_file)
        latest_import_date = validate_imported_platform(
            platform_name,
            imported,
        )
        imported_platforms[platform_name] = imported

        applied = apply_platform_import(
            platforms,
            platform_name,
            imported,
        )

        if not applied:
            raise ValueError(
                f"Nepavyko prijungti {platform_name} prie platforms masyvo."
            )

        counts = imported.get("counts", {})
        print(
            f"✅ {platform_name}: "
            f"{counts.get('active', 0)} aktyvios, "
            f"{counts.get('sold', 0)} parduotos pozicijos; "
            f"atskiro failo paskutinis mėnuo {latest_import_date}"
        )

    seb_mikro = imported_platforms["SEB Mikro"]
    revolut_brokerage = imported_platforms["Revolut Brokerage"]
    afranga = imported_platforms["Afranga"]
    debitum = imported_platforms["Debitum"]
    indemo = imported_platforms["Indemo"]

    active_platforms = [
        platform
        for platform in platforms
        if platform["active"]
    ]

    inactive_platforms = [
        platform
        for platform in platforms
        if not platform["active"]
    ]

    allocation = [
        {
            "name": "Fondai",
            "value": round(funds_value, 2),
        },
        {
            "name": "P2P",
            "value": round(p2p_value, 2),
        },
    ]

    portfolio_data = {
        "schemaVersion": 9,
        "updatedAt": format_date(portfolio_date),
        "portfolioValue": round(portfolio_value, 2),
        "invested": round(invested, 2),
        "profit": round(profit, 2),
        "returnRate": round(return_rate * 100, 2),
        "xirr": 0,
        "passiveIncome": 0,
        "monthlyChange": round(
            safe_number(
                ws.cell(
                    row=last_row,
                    column=total_columns["monthly_change"],
                ).value
            ),
            2,
        ),
        "platformCounts": {
            "active": len(active_platforms),
            "inactive": len(inactive_platforms),
            "total": len(platforms),
        },
        "allocation": allocation,
        "platforms": platforms,
        "history": history,
        "portfolioAnalytics": calculate_portfolio_analytics(history),
        "sebMikro": seb_mikro,
        "revolutBrokerage": revolut_brokerage,
        "afranga": afranga,
        "debitum": debitum,
        "indemo": indemo,
    }

    # Galutinė apsauga: platformos su atskirais importeriais negali
    # būti perrašytos pagrindinio Investavimas.xlsx istorija.
    for platform_name, imported in imported_platforms.items():
        imported_latest = imported["history"][-1]["date"]
        platform = next(
            item
            for item in platforms
            if normalize_name(item.get("name")) == normalize_name(platform_name)
        )
        platform_latest = platform["history"][-1]["date"]

        if platform_latest != imported_latest:
            raise ValueError(
                f"{platform_name}: platformos istorija baigiasi "
                f"{platform_latest}, bet atskiras failas baigiasi "
                f"{imported_latest}. JSON nebus rašomas."
            )

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with OUTPUT_FILE.open(
        "w",
        encoding="utf-8",
    ) as json_file:
        json.dump(
            portfolio_data,
            json_file,
            ensure_ascii=False,
            indent=2,
        )

    history_points = sum(
        len(platform["history"])
        for platform in platforms
    )

    print(f"✅ Paskutinė duomenų eilutė: {last_row}")
    print(f"✅ Data: {format_date(portfolio_date)}")
    print(f"✅ Investuota: {invested:.2f} EUR")
    print(f"✅ Portfelio vertė: {portfolio_value:.2f} EUR")
    print(f"✅ Pelnas: {profit:.2f} EUR")
    print(f"✅ Pelningumas: {return_rate * 100:.2f} %")
    print(f"✅ Aktyvių platformų: {len(active_platforms)}")
    print(f"✅ Neaktyvių platformų: {len(inactive_platforms)}")
    print(f"✅ Visų platformų: {len(platforms)}")
    print(f"✅ Platformų istorijos taškų: {history_points}")
    print(
        "✅ SEB Mikro ETF iš viso: "
        f"{seb_mikro['counts']['total']}"
    )
    print(
        "✅ Revolut Brokerage pozicijų iš viso: "
        f"{revolut_brokerage['counts']['total']}"
    )
    print(
        "✅ Afranga paskolų iš viso: "
        f"{afranga.get('p2pSummary', {}).get('totalLoans', 0)}"
    )
    print(
        "✅ Debitum paskolų iš viso: "
        f"{debitum.get('p2pSummary', {}).get('totalLoans', 0)}"
    )
    print(
        "✅ Indemo projektų iš viso: "
        f"{indemo.get('projectCounts', {}).get('total', 0)}"
    )
    print(f"✅ JSON schema: V{portfolio_data['schemaVersion']}")
    print(f"✅ JSON sukurtas: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()